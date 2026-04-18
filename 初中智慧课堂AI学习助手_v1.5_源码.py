# ======================================================
# 名称：初中智慧课堂AI学习助手
# 作者：邵长超
# 单位：沂水县第二实验中学
# 日期：2026.04
# ------------------------------------------------------
# 功能说明：
#   本程序针对初中人教版六三制教材开发，集成AI辅助教学功能，
#   支持全学科AI答疑、针对性随堂测试、学情分析、本地学习记录及离线语音朗读，
#   旨在提升学习效率，辅助智慧课堂教学。
# ======================================================

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import sqlite3
import os
import sys
import json
import threading
import random
from datetime import datetime
import re
import time
import queue

try:
    import requests
except ImportError:
    requests = None

# 离线语音合成
try:
    import pyttsx3
    TTS_AVAILABLE = True
except:
    TTS_AVAILABLE = False
    pyttsx3 = None

# ---------------------- 全局配置 ----------------------
WINDOW_TITLE = "初中智慧课堂AI学习助手_v1.5"
DB_FILE = "smart_classroom.db"
CONFIG_FILE = "ai_config.json"
DATA_EXCEL_FILE = "smart_classroom_data.xlsx"

ALL_SUBJECTS = ["全部", "语文", "数学", "英语", "物理", "化学", "生物", "道德与法治", "历史", "地理"]
ALL_GRADES = ["七年级", "八年级", "九年级"]
ALL_DIFFICULTIES = ["简单", "中等", "困难"]
ALL_QTYPES = ["选择题", "填空题", "简答题"]
ALL_EXAM_NUMS = ["3", "5", "10", "15"]

SUPPORT_MODELS = ["glm-4.7-flash", "glm-4-flash", "glm-4-flash-250414", "qwen3.5-flash", "qwen3.5-plus"]

# 触摸优化配色
COLORS = {
    "nav_bg": "#0F172A", "nav_fg": "#F8FAFC", "nav_hover": "#1E293B",
    "primary": "#1D4ED8", "secondary": "#0284C7", "success": "#059669",
    "warning": "#D97706", "danger": "#DC2626", "bg_main": "#F8FAFC",
    "bg_card": "#FFFFFF", "text_main": "#0F172A", "text_secondary": "#475569"
}

# 触摸优化字体（AI输出文字统一16号）
FONTS = {
    "title": ("微软雅黑", 20, "bold"),
    "nav": ("Segoe UI Emoji", 14, "bold"),
    "btn": ("Segoe UI Emoji", 13, "bold"),
    "content": ("微软雅黑", 14),
    "content_bold": ("微软雅黑", 14, "bold"),
    "small": ("微软雅黑", 11),
    "combo": ("微软雅黑", 13),
    "placeholder": ("微软雅黑", 13),
    "ai_output": ("微软雅黑", 16)      # AI输出内容字体统一16号
}

ENCOURAGE_TEXTS = ["太棒了！完全正确", "厉害啦！答对了", "完美！答案正确", "太优秀了！继续保持", "正确！做得很好",
                   "恭喜你答对了！"]

# ---------------------- 资源文件处理 ----------------------
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def release_data_excel():
    target_path = os.path.join(os.path.dirname(sys.argv[0]), DATA_EXCEL_FILE)
    if not os.path.exists(target_path):
        try:
            # 尝试从打包的资源中读取内置的完整 data.xlsx
            source_path = resource_path(DATA_EXCEL_FILE)
            if os.path.exists(source_path):
                import shutil
                shutil.copy2(source_path, target_path)
                print(f"已释放数据文件: {target_path}")
                return True
            else:
                # 如果资源中没有（例如开发环境），则创建默认数据
                import openpyxl
                wb = openpyxl.Workbook()
                ws_chapter = wb.active
                ws_chapter.title = "章节"
                ws_chapter.append(["科目", "年级", "单元", "课时"])
                ws_chapter.append(["语文", "七年级", "七年级上册 第一单元 四季美景", "整个章节"])
                ws_chapter.append(["语文", "七年级", "七年级上册 第一单元 四季美景", "春"])
                ws_chapter.append(["语文", "七年级", "七年级上册 第一单元 四季美景", "济南的冬天"])
                ws_chapter.append(["数学", "七年级", "七年级上册 第一章 有理数", "整个章节"])
                ws_chapter.append(["数学", "七年级", "七年级上册 第一章 有理数", "正数和负数"])
                ws_kp = wb.create_sheet("考点")
                ws_kp.append(["科目", "年级", "单元", "课时", "考点编号", "考点名称", "考点描述"])
                ws_kp.append(["语文", "七年级", "七年级上册 第一单元 四季美景", "春", "YW-7-1-1", "比喻修辞", "分析文中比喻句的表达效果"])
                wb.save(target_path)
                print(f"已创建默认数据文件: {target_path}")
                return True
        except Exception as e:
            print(f"释放数据文件失败: {e}")
            return False
    return True

def load_all_data_from_excel():
    chapters = {}
    knowledge_points = {}
    excel_path = os.path.join(os.path.dirname(sys.argv[0]), DATA_EXCEL_FILE)
    if not os.path.exists(excel_path):
        release_data_excel()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        # 加载章节表
        if "章节" in wb.sheetnames:
            ws = wb["章节"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0] or not row[1] or not row[2]:
                    continue
                subject = str(row[0]).strip()
                grade = str(row[1]).strip()
                chapter = str(row[2]).strip()
                lesson = str(row[3]).strip() if row[3] else "整个章节"
                if subject not in chapters:
                    chapters[subject] = {}
                if grade not in chapters[subject]:
                    chapters[subject][grade] = {}
                if chapter not in chapters[subject][grade]:
                    chapters[subject][grade][chapter] = []
                if lesson not in chapters[subject][grade][chapter]:
                    chapters[subject][grade][chapter].append(lesson)
        # 加载考点表
        if "考点" in wb.sheetnames:
            ws = wb["考点"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0] or not row[4]:
                    continue
                subject = str(row[0]).strip()
                grade = str(row[1]).strip()
                chapter = str(row[2]).strip() if row[2] else ""
                lesson = str(row[3]).strip() if row[3] else "整个章节"
                kp_id = str(row[4]).strip()
                kp_name = str(row[5]).strip()
                kp_desc = str(row[6]).strip() if row[6] else ""
                knowledge_points[kp_id] = {
                    "name": kp_name,
                    "desc": kp_desc,
                    "subject": subject,
                    "grade": grade,
                    "chapter": chapter,
                    "lesson": lesson
                }
        wb.close()
    except Exception as e:
        print(f"读取数据文件失败: {e}")
    return chapters, knowledge_points

# 加载章节和考点数据
CHAPTERS, KNOWLEDGE_POINTS = load_all_data_from_excel()
if not CHAPTERS:
    CHAPTERS = {}

# ==================== 智能语音管理类（离线+在线，自动降级） ====================
import asyncio
import threading
import time
import tempfile
import os
from edge_tts import Communicate

class SmartTTS:
    """智能语音管理类，离线使用 SAPI（无延迟，可立即停止），在线使用 edge-tts"""
    def __init__(self, mode='offline', voice='zh-CN-XiaoxiaoNeural'):
        self.mode = mode
        self.voice = voice
        self.offline_engine = None
        self.stop_flag = False
        self._init_offline_engine()
        self.fallback_callback = None

    def _init_offline_engine(self):
        """初始化离线语音引擎（使用 Windows SAPI，无延迟，可立即停止）"""
        try:
            import win32com.client
            self.offline_engine = win32com.client.Dispatch("SAPI.SpVoice")
            self.offline_engine.Rate = 1
            # 尝试设置为中文语音
            for voice in self.offline_engine.GetVoices():
                if 'chinese' in voice.GetDescription().lower() or 'zh' in voice.GetDescription().lower():
                    self.offline_engine.Voice = voice
                    break
            # ========== 预热：激活引擎，避免第一次朗读报错 ==========
            try:
                self.offline_engine.Speak("", 0)  # 同步播放空字符串，不发声
            except:
                pass
            # ==================================================
            print("离线语音引擎(SAPI)初始化成功")
        except Exception as e:
            print(f"离线引擎初始化失败: {e}")
            self.offline_engine = None

    def set_mode(self, mode):
        if mode in ['offline', 'online']:
            self.mode = mode
            print(f"语音模式已切换为: {mode}")

    def set_voice(self, voice):
        self.voice = voice
        print(f"在线音色已切换为: {voice}")

    def _speak_online(self, text, callback):
        """在线语音：生成临时mp3并播放（不卡界面，可停止）"""
        import pygame
        async def _async_speak():
            tmp_file = None
            try:
                with tempfile.NamedTemporaryFile(suffix=".mp3", delete=False) as f:
                    tmp_file = f.name
                communicate = Communicate(text, voice=self.voice)
                await communicate.save(tmp_file)
                if not pygame.mixer.get_init():
                    pygame.mixer.init()
                pygame.mixer.music.stop()
                pygame.mixer.music.load(tmp_file)
                pygame.mixer.music.play()
                while pygame.mixer.music.get_busy():
                    if self.stop_flag:
                        pygame.mixer.music.stop()
                        break
                    pygame.time.wait(100)
                if callback:
                    if self.stop_flag:
                        callback(False, "用户停止")
                    else:
                        callback(True, "在线语音播放完成")
            except Exception as e:
                print(f"在线语音错误: {e}")
                if self.offline_engine is not None:
                    self._speak_offline(text, callback)
                    if self.fallback_callback:
                        self.fallback_callback()
                else:
                    if callback:
                        callback(False, f"在线语音失败且离线不可用: {e}")
            finally:
                if tmp_file and os.path.exists(tmp_file):
                    try:
                        os.remove(tmp_file)
                    except:
                        pass
                self.stop_flag = False
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        loop.run_until_complete(_async_speak())
        loop.close()

    def _speak_offline(self, text, callback):
        """离线语音：使用 SAPI 异步播放，可立即停止（无弹窗）"""
        if self.offline_engine is None:
            print("离线引擎未初始化")
            if callback:
                callback(False, "离线引擎未初始化")
            return
        try:
            # 异步播放
            self.offline_engine.Speak(text, 1)
            # 循环检查播放状态，支持停止
            while True:
                try:
                    if not self.offline_engine.IsSpeaking():
                        break
                    if self.stop_flag:
                        self.offline_engine.Speak("", 2)  # 立即停止
                        break
                except Exception as e:
                    # 如果 IsSpeaking 出错，直接退出循环，不弹窗
                    print(f"检查播放状态异常: {e}")
                    break
                time.sleep(0.05)
            # 播放完成或停止后的回调
            if not self.stop_flag and callback:
                callback(True, "离线语音播放完成")
        except Exception as e:
            print(f"离线播放错误: {e}")
            if callback:
                callback(False, str(e))
        finally:
            self.stop_flag = False

    def speak(self, text, callback=None):
        """统一朗读接口"""
        if not text:
            return
        if self.mode == 'online':
            threading.Thread(target=self._speak_online, args=(text, callback), daemon=True).start()
        else:
            if self.offline_engine is not None:
                threading.Thread(target=self._speak_offline, args=(text, callback), daemon=True).start()
            else:
                print("离线引擎未初始化")
                if callback:
                    callback(False, "离线引擎未初始化")

    def stop(self):
        """停止当前朗读（离线或在线）"""
        self.stop_flag = True
        # 停止在线语音（pygame）
        try:
            import pygame
            if pygame.mixer.get_init():
                pygame.mixer.music.stop()
        except:
            pass
        # 停止离线语音（SAPI）
        if self.offline_engine is not None:
            try:
                self.offline_engine.Speak("", 2)
            except Exception as e:
                print(f"停止离线语音异常: {e}")

# ---------------------- 1. 数据库核心类 ----------------------
class LearningDB:
    def __init__(self, db_file):
        self.db_file = db_file
        self.init_db()

    def _connect(self):
        conn = sqlite3.connect(self.db_file)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        return conn, cursor

    def _close(self, conn):
        conn.commit()
        conn.close()

    def init_db(self):
        conn, cursor = self._connect()

        # 兼容旧表
        try:
            cursor.execute("PRAGMA table_info(knowledge_base)")
            columns = [row[1] for row in cursor.fetchall()]
            if 'lesson' not in columns:
                cursor.execute("DROP TABLE IF EXISTS knowledge_base")
                cursor.execute("DROP INDEX IF EXISTS idx_chapter_lesson")
                cursor.execute("DROP INDEX IF EXISTS idx_subject_grade")
        except:
            pass

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS knowledge_base (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                subject TEXT NOT NULL,
                grade TEXT NOT NULL,
                chapter TEXT NOT NULL,
                lesson TEXT NOT NULL,
                content TEXT NOT NULL,
                create_time TEXT DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_subject_grade ON knowledge_base (subject, grade)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_chapter_lesson ON knowledge_base (chapter, lesson)')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS ai_chat_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ask_time TEXT NOT NULL,
                subject TEXT NOT NULL,
                grade TEXT NOT NULL,
                scene TEXT NOT NULL,
                question TEXT NOT NULL,
                answer TEXT NOT NULL
            )
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS exam_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                exam_time TEXT NOT NULL,
                subject TEXT NOT NULL,
                grade TEXT NOT NULL,
                chapter TEXT NOT NULL,
                lesson TEXT NOT NULL,
                total_num INTEGER NOT NULL,
                correct_num INTEGER NOT NULL,
                accuracy TEXT NOT NULL
            )
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS exam_details (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                exam_record_id INTEGER,
                class_name TEXT NOT NULL,
                subject TEXT NOT NULL,
                grade TEXT NOT NULL,
                chapter TEXT NOT NULL,
                lesson TEXT NOT NULL,
                question TEXT NOT NULL,
                correct_answer TEXT,
                user_answer TEXT,
                is_correct INTEGER,
                knowledge_point TEXT,
                knowledge_point_id TEXT,
                question_type TEXT,
                create_time TEXT DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # 创建模型配置表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS model_config (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                model_name TEXT UNIQUE NOT NULL,
                api_base TEXT DEFAULT '',
                api_key TEXT,
                quota_type TEXT DEFAULT 'daily',
                quota_limit INTEGER DEFAULT 1000000,
                used_tokens INTEGER DEFAULT 0,
                period_start TEXT,
                is_enabled INTEGER DEFAULT 1,
                priority INTEGER DEFAULT 0
            )
        ''')
        # 插入默认模型（如果还没有）
        default_models = [
            # ========== 智谱AI 免费模型 ==========
            ("glm-4.7-flash", "https://open.bigmodel.cn/api/paas/v4/chat/completions", "daily", 100000000),
            ("glm-4-flash", "https://open.bigmodel.cn/api/paas/v4/chat/completions", "daily", 100000000),
            ("glm-4-flash-250414", "https://open.bigmodel.cn/api/paas/v4/chat/completions", "daily", 100000000),

            # ========== 阿里云百炼（通义千问） ==========
            ("qwen3.5-flash", "https://dashscope.aliyuncs.com/compatible-mode/v1/chat/completions", "once", 1000000),
            ("qwen3.5-plus", "https://dashscope.aliyuncs.com/compatible-mode/v1/chat/completions", "once", 1000000),
        ]
        for name, base, qtype, limit in default_models:
            cursor.execute("SELECT 1 FROM model_config WHERE model_name=?", (name,))
            if not cursor.fetchone():
                cursor.execute('''
                    INSERT INTO model_config (model_name, api_base, quota_type, quota_limit, used_tokens, period_start, is_enabled)
                    VALUES (?, ?, ?, ?, 0, ?, 1)
                ''', (name, base, qtype, limit, datetime.now().isoformat()))

        # 添加缺失列
        cursor.execute("PRAGMA table_info(exam_details)")
        existing = [col[1] for col in cursor.fetchall()]
        if 'knowledge_point_id' not in existing:
            cursor.execute("ALTER TABLE exam_details ADD COLUMN knowledge_point_id TEXT")
        if 'question_type' not in existing:
            cursor.execute("ALTER TABLE exam_details ADD COLUMN question_type TEXT")

        cursor.execute('CREATE INDEX IF NOT EXISTS idx_details_class ON exam_details (class_name, subject, grade, chapter)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_details_knowledge ON exam_details (knowledge_point_id)')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS mslq_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                student_name TEXT,
                class_name TEXT,
                test_type TEXT,
                total_score INTEGER,
                details TEXT,
                create_time TEXT DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # 添加 API 用量统计表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS api_usage (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                request_time TEXT,
                model TEXT,
                prompt_tokens INTEGER,
                completion_tokens INTEGER,
                total_tokens INTEGER,
                duration_ms INTEGER,
                scene TEXT
            )
        ''')

        cursor.execute("SELECT COUNT(*) FROM knowledge_base")
        if cursor.fetchone()[0] == 0 and CHAPTERS:
            self._insert_chapter_skeleton(cursor)

        # 确保 knowledge_base 表有 last_access_time 字段
        cursor.execute("PRAGMA table_info(knowledge_base)")
        columns = [col[1] for col in cursor.fetchall()]
        if 'last_access_time' not in columns:
            cursor.execute("ALTER TABLE knowledge_base ADD COLUMN last_access_time TEXT")

        self._close(conn)

    def _insert_chapter_skeleton(self, cursor):
        skeleton_data = []
        for subject, grades in CHAPTERS.items():
            for grade, chapters in grades.items():
                for chapter, lessons in chapters.items():
                    for lesson in lessons:
                        skeleton_data.append((subject, grade, chapter, lesson, ""))
        cursor.executemany(
            "INSERT INTO knowledge_base (subject, grade, chapter, lesson, content) VALUES (?, ?, ?, ?, ?)",
            skeleton_data)

    def get_chapters_by_subject_grade(self, subject, grade):
        if subject not in CHAPTERS or grade not in CHAPTERS[subject]:
            return []
        return list(CHAPTERS[subject][grade].keys())

    def get_lessons_by_chapter(self, subject, grade, chapter):
        if subject not in CHAPTERS or grade not in CHAPTERS[subject] or chapter not in CHAPTERS[subject][grade]:
            return ["整个章节"]
        return CHAPTERS[subject][grade][chapter]

    def get_chapter_content(self, subject, grade, chapter, lesson="整个章节"):
        conn, cursor = self._connect()
        cursor.execute(
            "SELECT content FROM knowledge_base WHERE subject = ? AND grade = ? AND chapter = ? AND lesson = ?",
            (subject, grade, chapter, lesson))
        row = cursor.fetchone()
        self._close(conn)
        return row[0] if row else None

    def update_last_access(self, subject, grade, chapter, lesson):
        """更新知识点最后访问时间"""
        conn, cursor = self._connect()
        cursor.execute(
            "UPDATE knowledge_base SET last_access_time = ? WHERE subject = ? AND grade = ? AND chapter = ? AND lesson = ?",
            (datetime.now().isoformat(), subject, grade, chapter, lesson)
        )
        self._close(conn)

    def get_last_access(self, subject, grade):
        """获取最近访问的章节和课时"""
        conn, cursor = self._connect()
        cursor.execute(
            "SELECT chapter, lesson FROM knowledge_base WHERE subject = ? AND grade = ? AND last_access_time IS NOT NULL ORDER BY last_access_time DESC LIMIT 1",
            (subject, grade)
        )
        row = cursor.fetchone()
        self._close(conn)
        if row:
            return row['chapter'], row['lesson']
        return None, None

    def update_chapter_content(self, subject, grade, chapter, lesson, content):
        conn, cursor = self._connect()
        cursor.execute(
            "DELETE FROM knowledge_base WHERE subject = ? AND grade = ? AND chapter = ? AND lesson = ?",
            (subject, grade, chapter, lesson))
        cursor.execute(
            "INSERT INTO knowledge_base (subject, grade, chapter, lesson, content) VALUES (?, ?, ?, ?, ?)",
            (subject, grade, chapter, lesson, content))
        self._close(conn)

    def search_knowledge(self, keyword, subject=None, grade=None):
        """根据关键词搜索知识点内容"""
        conn, cursor = self._connect()
        sql = "SELECT subject, grade, chapter, lesson, content FROM knowledge_base WHERE content LIKE ?"
        params = [f"%{keyword}%"]
        if subject and subject != "全部":
            sql += " AND subject = ?"
            params.append(subject)
        if grade:
            sql += " AND grade = ?"
            params.append(grade)
        cursor.execute(sql, params)
        rows = cursor.fetchall()
        self._close(conn)
        return rows

    def save_chat_record(self, subject, grade, scene, question, answer):
        conn, cursor = self._connect()
        ask_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cursor.execute(
            "INSERT INTO ai_chat_records (ask_time, subject, grade, scene, question, answer) VALUES (?, ?, ?, ?, ?, ?)",
            (ask_time, subject, grade, scene, question, answer))
        self._close(conn)

    def get_chat_records(self, subject="全部"):
        conn, cursor = self._connect()
        sql = "SELECT * FROM ai_chat_records"
        params = []
        if subject != "全部":
            sql += " WHERE subject = ?"
            params.append(subject)
        sql += " ORDER BY ask_time DESC"
        cursor.execute(sql, params)
        result = cursor.fetchall()
        self._close(conn)
        return result

    def save_exam_record(self, subject, grade, chapter, lesson, total_num, correct_num):
        conn, cursor = self._connect()
        exam_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        accuracy = f"{(correct_num / total_num) * 100:.1f}%"
        cursor.execute(
            "INSERT INTO exam_records (exam_time, subject, grade, chapter, lesson, total_num, correct_num, accuracy) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (exam_time, subject, grade, chapter, lesson, total_num, correct_num, accuracy))
        self._close(conn)

    def get_exam_records(self):
        conn, cursor = self._connect()
        cursor.execute("SELECT * FROM exam_records ORDER BY exam_time DESC")
        result = cursor.fetchall()
        self._close(conn)
        return result

    def clear_all_records(self):
        conn, cursor = self._connect()
        cursor.execute("DELETE FROM ai_chat_records")
        cursor.execute("DELETE FROM exam_records")
        cursor.execute("DELETE FROM exam_details")
        self._close(conn)


# ---------------------- 2. AI核心类 ----------------------
class AIAssistant:
    def __init__(self, db):
        self.db = db
        self.api_key = ""
        self.model_name = "glm-4-flash-250414"
        self.class_name = "7/8/9年级*班"
        self.textbook_version = "人教版（人民教育出版社）六三制"
        self.area = "山东省临沂市沂水县"
        self.load_config()

    def load_config(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    config = json.load(f)
                    self.api_key = config.get("api_key", "")
                    self.model_name = config.get("model_name", "glm-4-flash-250414")
                    self.class_name = config.get("class_name", "7/8/9年级*班")
                    self.textbook_version = config.get("textbook_version", "人教版（人民教育出版社）六三制")
                    self.area = config.get("area", "山东省临沂市沂水县")
            except:
                pass

    def save_config(self):
        config = {
            "api_key": self.api_key,
            "model_name": self.model_name,
            "class_name": self.class_name,
            "textbook_version": self.textbook_version,
            "area": self.area
        }
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)

    def verify_config(self):
        if not self.api_key:
            return False, "请先填写API Key"
        if not requests:
            return False, "请先安装requests库"
        try:
            url = self.get_current_api_base()
            headers = {"Authorization": f"Bearer {self.api_key}", "Content-Type": "application/json"}
            payload = {
                "model": self.model_name,
                "messages": [{"role": "user", "content": "你好"}],
                "max_tokens": 10
            }
            response = requests.post(url, json=payload, headers=headers, timeout=90)
            result = response.json()
            if response.status_code == 200 and "choices" in result:
                return True, "AI配置成功！"
            else:
                error_msg = result.get('error', {}).get('message', f'HTTP {response.status_code}')
                if "model" in error_msg.lower():
                    return False, f"模型 {self.model_name} 未开通或不存在，请确认模型名称正确。"
                return False, f"配置失败：{error_msg}"
        except requests.exceptions.Timeout:
            return False, "网络超时，请检查网络连接"
        except Exception as e:
            return False, f"网络错误：{str(e)}"

    def get_current_api_base(self):
        conn, cursor = self.db._connect()
        cursor.execute("SELECT api_base FROM model_config WHERE model_name=?", (self.model_name,))
        row = cursor.fetchone()
        self.db._close(conn)
        if row and row['api_base']:
            return row['api_base']
        return "https://dashscope.aliyuncs.com/compatible-mode/v1/chat/completions"

    def _get_model_config(self, model_name):
        conn, cursor = self.db._connect()
        cursor.execute("SELECT * FROM model_config WHERE model_name=?", (model_name,))
        row = cursor.fetchone()
        self.db._close(conn)
        return row

    def _update_used_tokens(self, model_name, total_tokens):
        conn, cursor = self.db._connect()
        cursor.execute("SELECT quota_type, quota_limit, used_tokens, period_start FROM model_config WHERE model_name=?", (model_name,))
        row = cursor.fetchone()
        if not row:
            return
        quota_type, limit, used, period_start = row['quota_type'], row['quota_limit'], row['used_tokens'], row['period_start']
        now = datetime.now()
        need_reset = False
        if quota_type == 'daily':
            if period_start is None or now.date() > datetime.fromisoformat(period_start).date():
                need_reset = True
        elif quota_type == 'monthly':
            if period_start is None or now.month != datetime.fromisoformat(period_start).month:
                need_reset = True
        elif quota_type == 'once':
            need_reset = False  # 一次性额度永不重置
        if need_reset:
            used = 0
            period_start = now.isoformat()
        used += total_tokens
        cursor.execute('''
            UPDATE model_config SET used_tokens=?, period_start=? WHERE model_name=?
        ''', (used, period_start, model_name))
        self.db._close(conn)

    def _check_quota(self, model_name):
        row = self._get_model_config(model_name)
        if not row:
            return True, 100, 0
        limit = row['quota_limit']
        used = row['used_tokens']
        if limit <= 0:
            return True, 100, 0
        remaining = limit - used
        if remaining <= 0:
            return False, 0, 0
        percent = (remaining / limit) * 100
        return True, percent, remaining

    def ai_request(self, subject, grade, scene, user_prompt, auto_save=True, temperature=0.3):

        # 检查当前模型额度
        ok, percent, remaining = self._check_quota(self.model_name)
        if not ok:
            return False, f"当前模型 {self.model_name} 免费额度已用完，请切换模型或联系管理员。"

        if not self.api_key:
            return False, "请先配置AI密钥"
        if not requests:
            return False, "请先安装requests库"

        # 根据学科动态生成答疑提示词
        subject_prompts = {
                "语文": f"""你是初中语文老师。回答问题时要像在课堂上自然讲解，避免使用“首先、其次”等序号。可以这样组织：
        先带着学生关注原文中的关键细节，引用一两句原话。然后分析写法，比如这里用了什么手法，这个手法有什么效果。接着谈谈作者的情感，注意情感的复杂性，不要简单说“喜爱”或“讨厌”。再联系作者生平或作品整体主题，把意思往深处引。最后抛出一个开放性的问题，让学生自己思考，问题要具体、有代入感。
        整个回答用平实的陈述句和问句，段落之间空一行。不要用“嗯”、“你看”等语气词，也不要使用任何Markdown符号或列表标记。""",

                "数学": f"""你是初中数学老师。回答问题时要像在课堂上自然讲解，避免使用“首先、其次”等序号。可以这样组织：
        先点明题目考查的核心概念或公式。然后给出一种解题思路，分步写清楚，每步说明依据。再指出常见的易错点，并解释为什么容易错。最后可以提一个类似的变式，让学生课后思考。
        整个回答用平实的陈述句，不要用“嗯”、“你看”等语气词，也不要使用任何Markdown符号。""",

                "英语": f"""你是初中英语老师。回答问题时要像在课堂上自然讲解，避免使用“首先、其次”等序号。可以这样组织：
        先说明句子或对话发生的语境，推测说话人意图。然后解释关键语法点或词汇，给出两个例句。再指出学生容易混淆的类似结构。接着引导学生用该知识点造一个自己的句子。最后可以补充一点英语国家的文化背景。
        整个回答用平实的陈述句和问句，不要用“嗯”、“你看”等语气词，也不要使用任何Markdown符号。""",

                "物理": f"""你是初中物理老师。回答问题时要像在课堂上自然讲解，避免使用“首先、其次”等序号。可以这样组织：
        先描述问题中的物理现象，联系生活实际。然后明确指出涉及的物理定律或公式，解释每个符号的含义。逐步推导结论，强调逻辑链条。建议一个简单的小实验让学生自己验证。最后举例说明该原理在科技或生活中的应用。
        整个回答用平实的陈述句，不要用“嗯”、“你看”等语气词，也不要使用任何Markdown符号。""",

                "化学": f"""你是初中化学老师。回答问题时要像在课堂上自然讲解，避免使用“首先、其次”等序号。可以这样组织：
        先判断属于哪种反应类型。写出正确的化学方程式，并配平，标注条件。解释反应前后的颜色、沉淀、气体等现象。如果需要计算，分步讲解。最后说明该反应在生活或工业中的应用。
        整个回答用平实的陈述句，不要用“嗯”、“你看”等语气词，也不要使用任何Markdown符号。""",

                "生物": f"""你是初中生物老师。回答问题时要像在课堂上自然讲解，避免使用“首先、其次”等序号。可以这样组织：
        先指出涉及的器官、细胞或生态系统组成部分。解释各部分的功能及相互配合关系。用分步方式描述生理过程。讨论如果某个环节出问题会有什么后果。最后给出与知识点相关的健康生活建议。
        整个回答用平实的陈述句，不要用“嗯”、“你看”等语气词，也不要使用任何Markdown符号。""",

                "道德与法治": f"""你是初中道德与法治老师。回答问题时要像在课堂上自然讲解，避免使用“首先、其次”等序号。可以这样组织：
        先概括材料中的核心观点或法律条文。结合具体案例或生活情境，分析行为的是非对错。指出应该树立的正确价值观。提出学生可以做到的具体行动。最后抛出一个两难问题，引导学生思考。
        整个回答用平实的陈述句和问句，不要用“嗯”、“你看”等语气词，也不要使用任何Markdown符号。""",

                "历史": f"""你是初中历史老师。回答问题时要像在课堂上自然讲解，避免使用“首先、其次”等序号。可以这样组织：
        先交代事件发生的时间、地点、主要人物。分析事件发生的根本原因和直接导火索。简述事件经过，并分析对当时及后世的影响。引导学生从不同立场看待同一事件。最后讨论该历史事件对今天的启示。
        整个回答用平实的陈述句和问句，不要用“嗯”、“你看”等语气词，也不要使用任何Markdown符号。""",

                "地理": f"""你是初中地理老师。回答问题时要像在课堂上自然讲解，避免使用“首先、其次”等序号。可以这样组织：
        先指出地理位置和地形、气候特征。分析自然或人文现象的形成原因。建议学生查看地图或气温曲线图，培养读图能力。讨论人类活动如何影响该地环境，或环境如何影响人类。最后提出合理利用或保护该地资源的建议。
        整个回答用平实的陈述句，不要用“嗯”、“你看”等语气词，也不要使用任何Markdown符号。""",
            }

        default_subject_prompt = f"""你是初中{subject}老师。回答问题时要像在课堂上自然讲解，避免使用序号。先给出核心概念，再提供解题思路或分析方法，然后指出常见误区，最后鼓励学生自己尝试。用平实的陈述句，不要使用任何Markdown符号。"""

        subject_specific = subject_prompts.get(subject, default_subject_prompt)

        scene_prompts = {
                "答疑": f"""{subject_specific}
        同时必须遵守以下通用规则：
        - 禁止任何开场白（如“好的”、“同学你好”），直接开始回答。
        - 按照“知识点 → 解题思路 → 引导（不直接给答案）”的顺序组织内容。
        - 如果学生的问题过短（少于3个汉字）、纯表情或无意义，请回复：“⚠️ 请提出具体的学习问题，乱发信息不利于你的进步。”
        - 每段之间空一行，不要使用“答：”等前缀。
        - **输出格式要求**：禁止使用任何Markdown标记，包括但不限于：`-`、`*`、`#`、`**`、`1.`、`2.` 等。同时禁止使用 `·`、`•`、`→` 等符号作为列表或分隔符。直接使用普通文字和自然换行来组织内容。段落之间用空行分隔即可，不要添加任何多余的点或线。""",
                "出题": f"你是初中{subject}命题专家。直接输出题目，不要输出任何开场白或结束语。严格按照格式：【对应章节】【题目】【选项】【答案】【解析】，每部分单独成行，题目间空行分隔。选择题选项用A. B. C. D.。解析以“考点：”开头。",
                "生成章节内容": "你是初中教材专家。直接输出知识点正文，不要输出任何开场白或结束语。",
                "判题": f"你是初中{subject}老师。直接判断对错，不要客套话。正确只输出“正确”；错误输出“错误”并简述原因（≤20字）。"
            }

        url = self.get_current_api_base()
        headers = {"Authorization": f"Bearer {self.api_key}", "Content-Type": "application/json"}

        messages = [
            {"role": "system", "content": scene_prompts.get(scene, "你是 helpful 的助手")},
            {"role": "user", "content": user_prompt}
        ]

        if scene == "生成章节内容":
            max_tokens = 12000  # 输出设置最大值
        elif scene == "出题":
            max_tokens = 5000
        else:
            max_tokens = 3500

        payload = {
            "model": self.model_name,
            "messages": messages,
            "temperature": temperature,
            "max_tokens": max_tokens
        }

        for attempt in range(2):
            try:
                response = requests.post(url, json=payload, headers=headers, timeout=90)
                result = response.json()
                if response.status_code == 200 and "choices" in result:
                    answer = result["choices"][0]["message"]["content"]
                    if not answer:
                        return False, "AI返回内容为空"
                    if auto_save and scene not in ["判题", "生成章节内容"]:
                        self.db.save_chat_record(subject, grade, scene, user_prompt, answer)

                    # 新增：记录 API 用量
                    usage = result.get("usage", {})
                    if usage:
                        prompt_tokens = usage.get("prompt_tokens", 0)
                        completion_tokens = usage.get("completion_tokens", 0)
                        total_tokens = usage.get("total_tokens", 0)
                        conn, cursor = self.db._connect()
                        cursor.execute('''
                            INSERT INTO api_usage (request_time, model, prompt_tokens, completion_tokens, total_tokens, scene)
                            VALUES (?, ?, ?, ?, ?, ?)
                        ''', (datetime.now().isoformat(), self.model_name, prompt_tokens, completion_tokens, total_tokens, scene))
                        self.db._close(conn)
                        self._update_used_tokens(self.model_name, total_tokens)   # 新增

                    return True, answer
                else:
                    error_msg = result.get('error', {}).get('message', f'HTTP {response.status_code}')
                    if "model" in error_msg.lower():
                        return False, f"模型 {self.model_name} 调用失败：请确认模型名称是否正确或是否已开通。"
                    return False, f"AI调用失败: {error_msg}"
            except requests.exceptions.Timeout:
                if attempt == 1:
                    return False, "请求超时（60秒），请检查网络连接"
                else:
                    time.sleep(2)
            except Exception as e:
                return False, f"网络错误：{str(e)}"

        return False, "未知错误"




# ---------------------- 3. 主程序界面 ----------------------
class SmartClassroomApp:
    def __init__(self, root):
        self.root = root
        self.root.title(WINDOW_TITLE)
        self.root.geometry("1366x768+0+0")
        self.root.state('zoomed')
        self.root.minsize(1280, 700)
        self.root.configure(bg=COLORS["bg_main"])

        self.db = LearningDB(DB_FILE)
        self.ai = AIAssistant(self.db)
        self.current_frame = None

        self.is_loading_knowledge = False
        self.is_loading_exam = False
        self.is_grading = False

        self.exam_running = False
        self.current_exam_data = []
        self.current_exam_idx = 0
        self.exam_correct_num = 0
        self.exam_subject_var = tk.StringVar(value="语文")
        self.exam_grade_var = tk.StringVar(value="七年级")
        self.exam_chapter_var = tk.StringVar(value="")
        self.exam_lesson_var = tk.StringVar(value="整个章节")
        self.exam_num_var = tk.StringVar(value="5")
        self.exam_diff_var = tk.StringVar(value="中等")
        self.exam_qtype_var = tk.StringVar(value="选择题")
        self.exam_kp_var = tk.StringVar(value="随机")
        self.exam_radio_buttons = []

        # 学情分析变量
        self.current_weak_chapter = None
        self.current_weak_kp = None

        # 右键菜单
        self.right_menu = tk.Menu(self.root, tearoff=0)
        self.right_menu.add_command(label=" 朗读选中内容", command=self.speak_selected_text)

        # 初始化智能语音管理器（默认离线模式）
        # 加载保存的语音模式和音色
        saved_mode = self.load_tts_mode()
        saved_voice = self.load_tts_voice()
        if saved_voice is None:
            saved_voice = 'zh-CN-XiaoxiaoNeural'
        self.tts = SmartTTS(mode=saved_mode, voice=saved_voice)
        self.pending_tts_mode = saved_mode
        self.pending_tts_voice = saved_voice
        self.tts.fallback_callback = self.on_tts_fallback

        # 先加载字体配置（修改全局 FONTS）
        self.load_font_size()

        self.build_side_nav()
        self.build_main_content()

        # Win7 下删除所有 Emoji，并将搜索按钮替换为“搜索”，并处理按钮比例问题
        self.root.after(200, self._remove_all_emoji)
        self.root.after(250, self._adjust_buttons_for_win7)

        # 自定义事件，用于线程安全更新UI
        self.root.bind('<<UpdateChapterTree>>', self._on_update_chapter_tree)
        self.root.bind('<<UpdateRightPanel>>', self._on_update_right_panel)

        self.switch_frame("home")

        # 启动后台监控模型额度
        self.start_quota_monitor()

        # 在 __init__ 方法末尾添加
        self.task_queue = queue.Queue()
        self.root.after(100, self._process_task_queue)  # 每 100 毫秒检查一次队列

        # ========== 启用全局触摸滑动 ==========
        self.enable_touch_scroll_for_all()

    def start_quota_monitor(self):
        """每10分钟检查所有模型的额度，并处理周期重置"""
        def monitor():
            conn, cursor = self.db._connect()
            cursor.execute("SELECT model_name, quota_limit, used_tokens, quota_type, period_start FROM model_config WHERE quota_limit>0")
            rows = cursor.fetchall()
            warnings = []
            now = datetime.now()
            for r in rows:
                limit = r['quota_limit']
                used = r['used_tokens']
                quota_type = r['quota_type']
                period_start = r['period_start']
                need_reset = False
                if quota_type == 'daily' and period_start and now.date() > datetime.fromisoformat(period_start).date():
                    need_reset = True
                elif quota_type == 'monthly' and period_start and now.month != datetime.fromisoformat(
                        period_start).month:
                    need_reset = True
                elif quota_type == 'once':
                    need_reset = False
                if need_reset:
                    cursor.execute("UPDATE model_config SET used_tokens=0, period_start=? WHERE model_name=?", (now.isoformat(), r['model_name']))
                    used = 0
                remaining = limit - used
                if remaining <= 0:
                    warnings.append(f"❌ {r['model_name']} 额度已用完")
                elif remaining / limit < 0.1:
                    warnings.append(f"⚠️ {r['model_name']} 剩余额度不足10% ({remaining} tokens)")
            self.db._close(conn)
            if warnings:
                self.root.after(0, lambda: messagebox.showwarning("模型额度提醒", "\n".join(warnings)))
            self.root.after(600000, monitor)  # 10分钟后再次检查
        self.root.after(10000, monitor)  # 启动后10秒开始第一次检查

    def _process_task_queue(self):
        """处理来自子线程的任务队列（线程安全）"""
        try:
            while True:
                task = self.task_queue.get_nowait()
                func, args, kwargs = task
                func(*args, **kwargs)
        except queue.Empty:
            pass
        finally:
            self.root.after(100, self._process_task_queue)

    def _on_update_chapter_tree(self, event):
        rows = getattr(self, '_pending_chapter_rows', None)
        if rows is not None:
            self._update_chapter_tree(rows)
            self._pending_chapter_rows = None

    def _on_update_right_panel(self, event):
        data = getattr(self, '_pending_right_data', None)
        if data is not None:
            kp_rows, type_rows, wrong_rows, advice_text, chapter = data
            self._update_right_panel_ui(kp_rows, type_rows, wrong_rows, advice_text, chapter)
            self._pending_right_data = None

    # ---------------------- 语音功能（智能双模） ----------------------
    def stop(self):
        """停止当前朗读（离线或在线）"""
        if self.offline_engine:
            self.offline_engine.stop()
        try:
            import pygame
            if pygame.mixer.get_init():
                pygame.mixer.music.stop()
        except:
            pass

    def speak_text(self, text):
        """朗读文本（使用智能语音管理器）"""
        if not text:
            return
        if not hasattr(self, 'tts'):
            messagebox.showerror("错误", "语音管理器未初始化，请重启程序。")
            return
        self.tts.speak(text, self._on_speak_finished)

    def stop_tts(self):
        """停止当前朗读（离线或在线）"""
        if hasattr(self, 'tts'):
            self.tts.stop()

    def _on_speak_finished(self, success, msg):
        """朗读结束回调（可选，用于调试或错误提示）"""
        # 如果是用户主动停止，不弹窗
        if msg == "用户停止":
            return
        if not success:
            self.root.after(0, lambda: messagebox.showerror("朗读失败", msg))

    def speak_selected_text(self):
        """右键菜单：朗读选中的文字"""
        focused = self.root.focus_get()
        if not isinstance(focused, tk.Text) and not isinstance(focused, scrolledtext.ScrolledText):
            return
        try:
            selected = focused.get(tk.SEL_FIRST, tk.SEL_LAST).strip()
            if not selected:
                return
            self.speak_text(selected)
        except tk.TclError:
            pass

    def bind_right_click(self, widget):
        """为控件绑定右键菜单（朗读选中文字）"""
        def show_menu(event):
            try:
                has_selection = False
                try:
                    has_selection = widget.get(tk.SEL_FIRST, tk.SEL_LAST).strip() != ""
                except:
                    pass
                self.right_menu.entryconfig(0, state=tk.NORMAL if has_selection else tk.DISABLED)
                self.right_menu.post(event.x_root, event.y_root)
            except:
                pass
        widget.bind("<Button-3>", show_menu)

    def on_tts_fallback(self):
        """当在线语音失败自动降级时，弹出提示"""
        self.root.after(0, lambda: messagebox.showwarning("语音降级",
                                                          "在线语音不可用，已自动切换为离线语音。您可以在设置中重新尝试在线模式。"))
        self.save_tts_mode('offline')  # 保存降级后的模式

    def set_tts_mode(self, mode):
        """用户点击模式按钮时，只更新UI样式，不实际切换，等待确认"""
        if mode == 'offline':
            self.update_tts_buttons_style('offline')
            self.pending_tts_mode = 'offline'
            # 离线模式下，禁用所有在线音色按钮
            for btn in self.voice_buttons:
                btn.config(state=tk.DISABLED, bg=COLORS["bg_card"], fg=COLORS["text_secondary"])
        else:
            self.update_tts_buttons_style('online')
            self.pending_tts_mode = 'online'
            # 在线模式下，启用音色按钮
            for btn, code in zip(self.voice_buttons, self.voice_btn_code):
                btn.config(state=tk.NORMAL, bg=COLORS["bg_card"], fg=COLORS["text_main"])
            # 高亮当前选中的音色
            self.highlight_selected_voice(self.pending_tts_voice)

    def update_tts_buttons_style(self, active_mode):
        """更新按钮高亮样式"""
        if active_mode == 'offline':
            self.offline_tts_btn.config(bg=COLORS["primary"], fg="white", bd=0, relief=tk.FLAT)
            self.online_tts_btn.config(bg=COLORS["bg_card"], fg=COLORS["text_main"], bd=1, relief=tk.RAISED)
        else:
            self.online_tts_btn.config(bg=COLORS["primary"], fg="white", bd=0, relief=tk.FLAT)
            self.offline_tts_btn.config(bg=COLORS["bg_card"], fg=COLORS["text_main"], bd=1, relief=tk.RAISED)

    def get_voice_display(self, voice_code):
        """根据语音代码获取显示名称"""
        for display, code in self.voice_options.items():
            if code == voice_code:
                return display
        return "zh-CN-XiaoxiaoNeural (女声-自然)"

    def select_tts_voice(self, voice_code):
        """用户点击某个音色按钮时，高亮该按钮，并保存到待确认变量"""
        self.pending_tts_voice = voice_code
        self.highlight_selected_voice(voice_code)

    def apply_tts_settings(self):
        """确认应用语音设置，并朗读提示"""
        # 应用模式
        self.tts.set_mode(self.pending_tts_mode)
        # 应用音色
        self.tts.set_voice(self.pending_tts_voice)
        # 保存音色到文件
        self.save_tts_voice(self.pending_tts_voice)
        # 朗读提示
        self.save_tts_mode(self.pending_tts_mode)
        self.speak_text("设置成功")

    def on_online_voice_changed(self, event=None):
        """用户更改了在线音色"""
        display = self.voice_display_var.get()
        voice_code = self.voice_options.get(display, "zh-CN-XiaoxiaoNeural")
        self.tts.set_voice(voice_code)
        self.save_tts_voice(voice_code)
        messagebox.showinfo("音色切换", f"在线音色已切换为：{display}")

    def load_tts_voice(self):
        """从配置文件加载上次选择的在线音色"""
        config_file = "tts_config.json"
        if os.path.exists(config_file):
            try:
                with open(config_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    return data.get("online_voice", "zh-CN-XiaoxiaoNeural")
            except:
                pass
        return None

    def save_tts_voice(self, voice):
        """保存当前选择的在线音色"""
        config_file = "tts_config.json"
        try:
            with open(config_file, "w", encoding="utf-8") as f:
                json.dump({"online_voice": voice}, f)
        except:
            pass

    def highlight_selected_voice(self, voice_code):
        """高亮选中的音色按钮，其他按钮恢复默认"""
        for btn, code in zip(self.voice_buttons, self.voice_btn_code):
            if code == voice_code:
                btn.config(bg=COLORS["primary"], fg="white", bd=0, relief=tk.FLAT)
            else:
                btn.config(bg=COLORS["bg_card"], fg=COLORS["text_main"], bd=1, relief=tk.RAISED)

    def reset_tts_settings(self):
        """恢复默认设置：离线模式 + 默认音色"""
        self.pending_tts_mode = 'offline'
        self.pending_tts_voice = 'zh-CN-XiaoxiaoNeural'
        # 更新UI按钮样式
        self.update_tts_buttons_style('offline')
        # 高亮默认音色
        self.highlight_selected_voice(self.pending_tts_voice)
        # 禁用音色按钮（因为离线模式）
        for btn in self.voice_buttons:
            btn.config(state=tk.DISABLED, bg=COLORS["bg_card"], fg=COLORS["text_secondary"])
        # 立即应用设置
        self.tts.set_mode('offline')
        self.tts.set_voice(self.pending_tts_voice)
        self.save_tts_voice(self.pending_tts_voice)
        self.save_tts_mode('offline')
        self.speak_text("已恢复默认设置")


    # ---------------------- 导航栏 ----------------------
    def build_side_nav(self):
        self.nav_frame = tk.Frame(self.root, bg=COLORS["nav_bg"], width=240)
        self.nav_frame.pack(side=tk.LEFT, fill=tk.Y)
        self.nav_frame.pack_propagate(False)

        tk.Label(
            self.nav_frame, text="🤖 初中智慧课堂\nAI学习助手",
            font=FONTS["title"], bg=COLORS["nav_bg"], fg=COLORS["nav_fg"],
            justify="center"
        ).pack(pady=30)

        self.nav_config = [
            {"name": "🏠 首页", "frame": "home"},
            {"name": "📚 全科知识库", "frame": "knowledge"},
            {"name": "🤖 AI全科答疑", "frame": "ai_answer"},
            {"name": "📝 AI随堂答题", "frame": "ai_exam"},
            {"name": "📊 学情分析", "frame": "records"},
            {"name": "⚙️ 系统设置", "frame": "setting"},
        ]

        self.nav_buttons = []
        for nav_item in self.nav_config:
            btn = tk.Button(
                self.nav_frame, text=nav_item["name"],
                font=FONTS["nav"], bg=COLORS["nav_bg"], fg=COLORS["nav_fg"],
                bd=0, relief=tk.FLAT, cursor="hand2",
                width=22, height=2, anchor="w", padx=25,
                command=lambda f=nav_item["frame"]: self.switch_frame(f)
            )
            btn.pack(pady=8, fill=tk.X)
            self.nav_buttons.append(btn)
            btn.bind("<Enter>", lambda e, b=btn: b.configure(bg=COLORS["nav_hover"]))
            btn.bind("<Leave>", lambda e, b=btn: b.configure(bg=COLORS["nav_bg"]) if not hasattr(b, "is_active") or not b.is_active else None)

        tk.Label(
            self.nav_frame, text="初中智慧课堂AI学习助手 v1.5",
            font=FONTS["small"], bg=COLORS["nav_bg"], fg=COLORS["text_secondary"]
        ).pack(side=tk.BOTTOM, pady=25)

    # ---------------------- 主内容区域 ----------------------
    def build_main_content(self):
        self.content_container = tk.Frame(self.root, bg=COLORS["bg_main"])
        self.content_container.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        self.frames = {}
        self.frames["home"] = self.build_home_frame()
        self.frames["knowledge"] = self.build_knowledge_frame()
        self.frames["ai_answer"] = self.build_ai_answer_frame()
        self.frames["ai_exam"] = self.build_ai_exam_frame()
        self.frames["records"] = self.build_records_frame()
        self.frames["setting"] = self.build_setting_frame()

    def switch_frame(self, frame_name):
        for btn in self.nav_buttons:
            btn.is_active = False
            btn.configure(bg=COLORS["nav_bg"], fg=COLORS["nav_fg"])
        for idx, nav_item in enumerate(self.nav_config):
            if nav_item["frame"] == frame_name:
                self.nav_buttons[idx].is_active = True
                self.nav_buttons[idx].configure(bg=COLORS["primary"], fg="white")
                break
        if self.current_frame:
            self.current_frame.pack_forget()
        self.current_frame = self.frames[frame_name]
        self.current_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)
        self.update_speak_buttons_state()

    def update_speak_buttons_state(self):
        if hasattr(self, 'speak_knowledge_btn'):
            self.speak_knowledge_btn.config(state=tk.NORMAL if TTS_AVAILABLE else tk.DISABLED)
        if hasattr(self, 'speak_answer_btn'):
            self.speak_answer_btn.config(state=tk.NORMAL if TTS_AVAILABLE else tk.DISABLED)
        if hasattr(self, 'stop_speak_btn'):
            self.stop_speak_btn.config(state=tk.NORMAL if TTS_AVAILABLE else tk.DISABLED)

    # ---------------------- 首页 ----------------------
    def build_home_frame(self):
        frame = tk.Frame(self.content_container, bg=COLORS["bg_main"])

        tk.Label(
            frame, text="👋 欢迎使用初中智慧课堂AI学习助手",
            font=FONTS["title"], bg=COLORS["bg_main"], fg=COLORS["primary"]
        ).pack(pady=30)

        tk.Label(
            frame, text=f"适配{self.ai.area}{self.ai.textbook_version} · 初中全学科覆盖",
            font=FONTS["content"], bg=COLORS["bg_main"], fg=COLORS["text_secondary"]
        ).pack()

        if not self.ai.api_key:
            tk.Label(
                frame, text="⚠️ 未配置AI密钥，在线AI功能暂不可用，您可以直接使用离线知识库功能。",
                font=FONTS["content"], bg=COLORS["bg_main"], fg=COLORS["warning"]
            ).pack(pady=12)

        if not TTS_AVAILABLE:
            tk.Label(
                frame, text="⚠️ 当前系统不支持离线语音朗读功能，其他功能均可正常使用。",
                font=FONTS["content"], bg=COLORS["bg_main"], fg=COLORS["warning"]
            ).pack()

        card_frame = tk.Frame(frame, bg=COLORS["bg_main"])
        card_frame.pack(expand=True, pady=30)

        quick_cards = [
            {"name": "📚 全科知识库", "desc": "章节知识点全覆盖", "color": COLORS["primary"], "frame": "knowledge"},
            {"name": "🤖 AI全科答疑", "desc": "智能解答学习疑问", "color": COLORS["secondary"], "frame": "ai_answer"},
            {"name": "📝 AI随堂答题", "desc": "个性化精准出题", "color": COLORS["success"], "frame": "ai_exam"},
            {"name": "📊 学情分析", "desc": "班级学情与学习记录", "color": COLORS["warning"], "frame": "records"},
        ]

        for idx, card in enumerate(quick_cards):
            card_item = tk.Frame(card_frame, bg=card["color"], width=280, height=160, relief=tk.FLAT, bd=0)
            card_item.grid(row=idx // 2, column=idx % 2, padx=30, pady=20, sticky="nsew")
            card_item.pack_propagate(False)

            inner_frame = tk.Frame(card_item, bg=card["color"])
            inner_frame.place(relx=0.5, rely=0.5, anchor="center")

            tk.Label(inner_frame, text=card["name"], font=("Segoe UI Emoji", 15, "bold"), bg=card["color"], fg="white").pack(pady=8)
            tk.Label(inner_frame, text=card["desc"], font=FONTS["content"], bg=card["color"], fg="white").pack()

            card_item.bind("<Button-1>", lambda e, f=card["frame"]: self.switch_frame(f))
            for widget in inner_frame.winfo_children():
                widget.bind("<Button-1>", lambda e, f=card["frame"]: self.switch_frame(f))

        card_frame.grid_columnconfigure(0, weight=1)
        card_frame.grid_columnconfigure(1, weight=1)

        return frame

    # ---------------------- 知识库 ----------------------
    def build_knowledge_frame(self):
        frame = tk.Frame(self.content_container, bg=COLORS["bg_main"])

        if self.root.winfo_screenheight() >= 900:
            tk.Label(frame, text="📚 全科知识库", font=FONTS["title"], bg=COLORS["bg_main"],
                     fg=COLORS["text_main"]).pack(pady=8, anchor="w")

        filter_frame = tk.LabelFrame(frame, text="选择章节&课时", font=FONTS["content_bold"], bg=COLORS["bg_card"],
                                     fg=COLORS["primary"])
        filter_frame.pack(fill=tk.X, pady=8, padx=2)

        row1 = tk.Frame(filter_frame, bg=COLORS["bg_card"])
        row1.pack(fill=tk.X, padx=15, pady=10)

        # 科目、年级、单元、课时下拉框
        tk.Label(row1, text="科目：", font=FONTS["content_bold"], bg=COLORS["bg_card"]).pack(side=tk.LEFT)
        self.knowledge_subject_var = tk.StringVar(value="语文")
        subject_combo = ttk.Combobox(row1, textvariable=self.knowledge_subject_var, values=ALL_SUBJECTS[1:],
                                     font=FONTS["combo"], width=12, state="readonly")
        subject_combo.pack(side=tk.LEFT, padx=6)
        subject_combo.bind("<<ComboboxSelected>>", lambda e: self.on_subject_grade_changed())

        tk.Label(row1, text="年级：", font=FONTS["content_bold"], bg=COLORS["bg_card"]).pack(side=tk.LEFT, padx=12)
        self.knowledge_grade_var = tk.StringVar(value="七年级")
        grade_combo = ttk.Combobox(row1, textvariable=self.knowledge_grade_var, values=ALL_GRADES, font=FONTS["combo"],
                                   width=10, state="readonly")
        grade_combo.pack(side=tk.LEFT, padx=6)
        grade_combo.bind("<<ComboboxSelected>>", lambda e: self.on_subject_grade_changed())

        tk.Label(row1, text="单元：", font=FONTS["content_bold"], bg=COLORS["bg_card"]).pack(side=tk.LEFT, padx=12)
        self.knowledge_chapter_var = tk.StringVar(value="")
        self.chapter_combo = ttk.Combobox(row1, textvariable=self.knowledge_chapter_var, values=[], font=FONTS["combo"],
                                          width=28, state="readonly")
        self.chapter_combo.pack(side=tk.LEFT, padx=6)
        self.chapter_combo.bind("<<ComboboxSelected>>", lambda e: self.on_chapter_selected())

        tk.Label(row1, text="课时：", font=FONTS["content_bold"], bg=COLORS["bg_card"]).pack(side=tk.LEFT, padx=12)
        self.knowledge_lesson_var = tk.StringVar(value="整个章节")
        self.lesson_combo = ttk.Combobox(row1, textvariable=self.knowledge_lesson_var, values=[], font=FONTS["combo"],
                                         width=22, state="readonly")
        self.lesson_combo.pack(side=tk.LEFT, padx=6)
        self.lesson_combo.bind("<<ComboboxSelected>>", lambda e: self.on_lesson_selected())

        # 操作按钮行
        row2 = tk.Frame(filter_frame, bg=COLORS["bg_card"])
        row2.pack(fill=tk.X, padx=15, pady=8)

        # 左侧按钮组
        self.refresh_knowledge_btn = tk.Button(
            row2, text="📥 知识加载/刷新", font=FONTS["btn"],
            bg=COLORS["secondary"], fg="white", bd=0, relief=tk.FLAT, cursor="hand2",
            width=14, height=1, command=self.force_refresh_chapter
        )
        self.refresh_knowledge_btn.pack(side=tk.LEFT, padx=6)

        self.speak_knowledge_btn = tk.Button(
            row2, text="🔊 朗读全部", font=FONTS["btn"],
            bg=COLORS["success"], fg="white", bd=0, relief=tk.FLAT, cursor="hand2",
            width=12, height=1, command=self.speak_knowledge_content
        )
        self.speak_knowledge_btn.pack(side=tk.LEFT, padx=6)

        self.stop_speak_btn = tk.Button(
            row2, text="⏹️ 停止朗读", font=FONTS["btn"],
            bg=COLORS["danger"], fg="white", bd=0, relief=tk.FLAT, cursor="hand2",
            width=12, height=1, command=self.stop_tts
        )
        self.stop_speak_btn.pack(side=tk.LEFT, padx=6)

        # 字号调节按钮
        tk.Button(row2, text="A+", font=FONTS["btn"], bg=COLORS["warning"], fg="white", bd=0,
                  relief=tk.FLAT, cursor="hand2", width=4, command=lambda: self.change_font_size(2)).pack(side=tk.LEFT,
                                                                                                          padx=6)
        tk.Button(row2, text="A-", font=FONTS["btn"], bg=COLORS["warning"], fg="white", bd=0,
                  relief=tk.FLAT, cursor="hand2", width=4, command=lambda: self.change_font_size(-2)).pack(side=tk.LEFT,
                                                                                                           padx=6)

        # 继续学习按钮（放在字号调节按钮右侧）
        tk.Button(row2, text="📌 继续学习", font=FONTS["btn"], bg=COLORS["secondary"], fg="white", bd=0,
                  relief=tk.FLAT, cursor="hand2", width=10, command=self.continue_learning).pack(side=tk.LEFT, padx=6)
        # 添加功能用法提示标签（小字，灰色）
        #  tk.Label(row2, text="（跳转到上次学习的章节）", font=FONTS["small"], bg=COLORS["bg_card"],
        #         fg=COLORS["text_secondary"]).pack(side=tk.LEFT, padx=2)

        # 状态标签（朗读状态）
        self.knowledge_status_label = tk.Label(row2, text="", font=FONTS["content"], bg=COLORS["bg_card"],
                                               fg=COLORS["primary"])
        self.knowledge_status_label.pack(side=tk.LEFT, padx=6)
        self.knowledge_status_label.pack_forget()

        # ========== 右侧布局：仅搜索框 + 搜索按钮 ==========
        right_container = tk.Frame(row2, bg=COLORS["bg_card"])
        right_container.pack(side=tk.RIGHT)

        # 搜索按钮（在输入框右侧）
        tk.Button(right_container, text="🔍", font=FONTS["btn"], bg=COLORS["primary"], fg="white", bd=0,
                  relief=tk.FLAT, cursor="hand2", width=4, command=self.search_knowledge).pack(side=tk.RIGHT, padx=2)

        # 搜索输入框（在按钮左侧）
        self.search_entry = tk.Entry(right_container, font=FONTS["content"], width=15, fg="gray")
        self.search_entry.insert(0, "💡 输入关键词...")
        self.search_entry.bind("<FocusIn>", lambda e: self._clear_search_placeholder())
        self.search_entry.bind("<FocusOut>", lambda e: self._restore_search_placeholder())
        self.search_entry.pack(side=tk.RIGHT, padx=(0, 6))

        # 内容显示区
        self.knowledge_content = scrolledtext.ScrolledText(frame, font=FONTS["ai_output"], wrap=tk.WORD, padx=15,
                                                           pady=15)
        self.knowledge_content.pack(fill=tk.BOTH, expand=True, pady=8)

        # 返回顶部按钮（初始隐藏）
        self.top_btn = tk.Button(self.knowledge_content, text="↑ 顶部", font=FONTS["small"],
                                 bg=COLORS["primary"], fg="white", bd=0, relief=tk.FLAT,
                                 cursor="hand2", command=self.scroll_to_top)
        self.top_btn.place_forget()

        # 绑定滚动事件，控制按钮显示
        def on_scroll(event):
            self.update_top_btn_visibility()
        self.knowledge_content.bind("<MouseWheel>", on_scroll)
        self.knowledge_content.bind("<Button-4>", on_scroll)   # Linux 向上滚动
        self.knowledge_content.bind("<Button-5>", on_scroll)   # Linux 向下滚动
        # 内容大小改变时也更新
        self.knowledge_content.bind("<Configure>", on_scroll)

        self.bind_right_click(self.knowledge_content)

        # 返回顶部按钮
        self.top_btn = tk.Button(self.knowledge_content, text="↑ 顶部", font=FONTS["small"], bg=COLORS["primary"],
                                 fg="white", bd=0, relief=tk.FLAT, cursor="hand2",
                                 command=lambda: self.knowledge_content.yview_moveto(0))
        self.top_btn.place(relx=0.95, rely=0.95, anchor="se")
        self.top_btn.lower()

        self.refresh_chapter_options()
        self.root.after(500, self.auto_load_last_access)
        return frame

    # 辅助方法（用于搜索框占位符）
    def _clear_search_placeholder(self):
        if self.search_entry.get() == "💡 输入关键词...":
            self.search_entry.delete(0, tk.END)
            self.search_entry.config(fg="black")

    def _restore_search_placeholder(self):
        if self.search_entry.get().strip() == "":
            self.search_entry.delete(0, tk.END)
            self.search_entry.insert(0, "💡 输入关键词...")
            self.search_entry.config(fg="gray")

    def on_subject_grade_changed(self):
        self.refresh_chapter_options()

    def refresh_chapter_options(self):
        subject = self.knowledge_subject_var.get()
        grade = self.knowledge_grade_var.get()
        chapters = self.db.get_chapters_by_subject_grade(subject, grade)
        self.chapter_combo['values'] = chapters
        if chapters:
            self.knowledge_chapter_var.set(chapters[0])
            self.on_chapter_selected()
        else:
            self.knowledge_chapter_var.set("")
            self._update_knowledge_text("💡 请先选择有效的科目和年级。", is_placeholder=True)

    def on_chapter_selected(self):
        subject = self.knowledge_subject_var.get()
        grade = self.knowledge_grade_var.get()
        chapter = self.knowledge_chapter_var.get()
        if not chapter:
            return
        lessons = self.db.get_lessons_by_chapter(subject, grade, chapter)
        self.lesson_combo['values'] = lessons
        self.knowledge_lesson_var.set(lessons[0])
        self.on_lesson_selected()

    def on_lesson_selected(self):
        subject = self.knowledge_subject_var.get()
        grade = self.knowledge_grade_var.get()
        chapter = self.knowledge_chapter_var.get()
        lesson = self.knowledge_lesson_var.get()
        if not chapter:
            return

        content = self.db.get_chapter_content(subject, grade, chapter, lesson)
        if content and len(content) > 10:
            self._update_knowledge_text(content)
            self.db.update_last_access(subject, grade, chapter, lesson)
        else:
            full_name = chapter
            if lesson != "整个章节":
                full_name = f"{chapter} - {lesson}"
            placeholder_text = (
                f"💡 {full_name}\n\n"
                f"本地暂无知识点。\n\n"
                f"请点击【📥 知识加载/刷新】从AI获取（预计10-30秒）。\n\n"
                f"点击【📌 继续学习】可跳转到上次学习位置。\n\n"
                f"输入关键词可搜索已缓存内容。"
            )
            self._update_knowledge_text(placeholder_text, is_placeholder=True)

    def _update_knowledge_text(self, text, is_placeholder=False):
        """显示知识点内容，并自动清理 Markdown 和占位符"""
        # 如果不是占位符，则清理文本
        if not is_placeholder:
            text = self._clean_markdown(text)
        self.knowledge_content.config(state=tk.NORMAL)
        self.knowledge_content.delete(1.0, tk.END)
        self.knowledge_content.insert(tk.END, text)
        if is_placeholder:
            self.knowledge_content.config(fg="gray", font=FONTS["placeholder"])
        else:
            self.knowledge_content.config(fg="black", font=FONTS["ai_output"])
        self.knowledge_content.config(state=tk.DISABLED)
        self.root.after(100, self.update_top_btn_visibility)

    def force_refresh_chapter(self):
        if self.is_loading_knowledge:
            return
        if not self.ai.api_key:
            messagebox.showwarning("提示", "请先去【系统设置】配置AI API Key！")
            return

        subject = self.knowledge_subject_var.get()
        grade = self.knowledge_grade_var.get()
        chapter = self.knowledge_chapter_var.get()
        lesson = self.knowledge_lesson_var.get()
        if not chapter:
            messagebox.showwarning("提示", "请先选择具体章节")
            return

        full_name = chapter
        if lesson != "整个章节":
            full_name = f"{chapter} - {lesson}"

        self.is_loading_knowledge = True
        self.refresh_knowledge_btn.config(state=tk.DISABLED, text="⏳ 获取中...")
        loading_text = f" 正在连接AI生成《{full_name}》的内容...\n\n请稍候，这可能需要10-30秒。"
        self._update_knowledge_text(loading_text, is_placeholder=True)

        def on_knowledge_finish(success, result):
            self.is_loading_knowledge = False
            self.refresh_knowledge_btn.config(state=tk.NORMAL, text="📥 知识加载/刷新")
            if success:
                # ========== 新增：清理 Markdown 和占位符 ==========
                cleaned = self._clean_markdown(result)
                final_content = f"【{full_name}】\n\n{result}"
                self.db.update_chapter_content(subject, grade, chapter, lesson, final_content)
                self._update_knowledge_text(final_content)
            else:
                self._update_knowledge_text(f"⚠️ 获取失败：{result}\n\n请检查网络或API设置后重试。")

        def thread_func():
            chapter = self.knowledge_chapter_var.get()
            lesson = self.knowledge_lesson_var.get()
            custom_prompt = self._get_knowledge_prompt(subject, grade, chapter, lesson)
            success, result = self.ai.ai_request(subject, grade, "生成章节内容", custom_prompt, auto_save=False,
                                                 temperature=0.3)
            self.root.after(0, lambda: on_knowledge_finish(success, result))

        threading.Thread(target=thread_func, daemon=True).start()

    def speak_knowledge_content(self):
        """朗读当前显示的知识点内容（非占位符）"""
        text = self.knowledge_content.get(1.0, tk.END).strip()
        # 判断是否为占位符文本（灰色字体或包含提示关键词）
        is_placeholder = (self.knowledge_content.cget("fg") == "gray" or
                          text.startswith("💡") or
                          "暂无知识点" in text or
                          "请点击上方的【📥 知识加载/刷新】" in text)
        if not text or is_placeholder:
            messagebox.showinfo("提示", "没有可朗读的内容，请先加载或选择一个章节。")
            return
        self.speak_text(text)

    def search_knowledge(self):
        """全文搜索知识点"""
        keyword = self.search_entry.get().strip()
        if not keyword:
            messagebox.showwarning("提示", "请输入搜索关键词")
            return
        subject = self.knowledge_subject_var.get()
        grade = self.knowledge_grade_var.get()
        rows = self.db.search_knowledge(keyword, subject if subject != "全部" else None, grade)
        if not rows:
            messagebox.showinfo("搜索结果", "未找到匹配的知识点")
            return
        win = tk.Toplevel(self.root)
        win.title("搜索结果")
        win.geometry("700x500")
        win.configure(bg=COLORS["bg_main"])
        listbox = tk.Listbox(win, font=("微软雅黑", 12), selectmode=tk.SINGLE)
        listbox.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        items = []
        for r in rows:
            disp = f"{r['subject']} {r['grade']} {r['chapter']} - {r['lesson']}"
            items.append((disp, r['subject'], r['grade'], r['chapter'], r['lesson']))
            listbox.insert(tk.END, disp)

        def on_select():
            sel = listbox.curselection()
            if sel:
                _, s, g, c, l = items[sel[0]]
                self.knowledge_subject_var.set(s)
                self.knowledge_grade_var.set(g)
                self.refresh_chapter_options()
                self.knowledge_chapter_var.set(c)
                self.chapter_combo.set(c)
                self.on_chapter_selected()
                self.knowledge_lesson_var.set(l)
                self.lesson_combo.set(l)
                self.on_lesson_selected()
                win.destroy()

        btn_frame = tk.Frame(win, bg=COLORS["bg_main"])
        btn_frame.pack(pady=10)
        tk.Button(btn_frame, text="跳转", font=FONTS["btn"], bg=COLORS["primary"], fg="white",
                  command=on_select).pack(side=tk.LEFT, padx=10)
        tk.Button(btn_frame, text="取消", font=FONTS["btn"], bg=COLORS["danger"], fg="white",
                  command=win.destroy).pack(side=tk.LEFT, padx=10)

    def continue_learning(self):
        subject = self.knowledge_subject_var.get()
        grade = self.knowledge_grade_var.get()
        chapter, lesson = self.db.get_last_access(subject, grade)
        if chapter:
            self.knowledge_chapter_var.set(chapter)
            self.chapter_combo.set(chapter)
            self.on_chapter_selected()
            if lesson:
                self.knowledge_lesson_var.set(lesson)
                self.lesson_combo.set(lesson)
                self.on_lesson_selected()
            messagebox.showinfo("继续学习", f"已跳转到上次学习位置：{chapter} - {lesson}")
        else:
            messagebox.showinfo("继续学习", "暂无学习记录，请先浏览一些知识点。")

    def change_font_size(self, delta):
        """调节知识点显示区域的字体大小"""
        current = self.knowledge_content.cget("font")
        if isinstance(current, tuple):
            family, size = current[0], current[1]
        else:
            family = "微软雅黑"
            size = 16
        new_size = max(10, min(28, size + delta))
        self.knowledge_content.config(font=(family, new_size))

    def copy_knowledge_content(self):
        """复制当前显示的知识点内容到剪贴板"""
        text = self.knowledge_content.get(1.0, tk.END).strip()
        if text:
            self.root.clipboard_clear()
            self.root.clipboard_append(text)
            messagebox.showinfo("提示", "已复制到剪贴板")
        else:
            messagebox.showwarning("提示", "没有可复制的内容")

    def auto_load_last_access(self):
        """自动加载上次学习位置（仅在首次打开页面时）"""
        if not self.knowledge_chapter_var.get():
            subject = self.knowledge_subject_var.get()
            grade = self.knowledge_grade_var.get()
            chapter, lesson = self.db.get_last_access(subject, grade)
            if chapter:
                self.knowledge_chapter_var.set(chapter)
                self.chapter_combo.set(chapter)
                self.on_chapter_selected()
                if lesson:
                    self.knowledge_lesson_var.set(lesson)
                    self.lesson_combo.set(lesson)
                    self.on_lesson_selected()

    def update_top_btn_visibility(self):
        """根据滚动位置和内容高度决定是否显示返回顶部按钮"""
        try:
            # 获取当前滚动位置 (0~1)
            first, last = self.knowledge_content.yview()
            # 判断内容总高度是否超过可视区域
            total_lines = int(self.knowledge_content.index('end-1c').split('.')[0])
            first_visible = self.knowledge_content.index('@0,0')
            visible_lines = int(first_visible.split('.')[0]) if first_visible != '1.0' else total_lines
            need_scroll = total_lines > visible_lines
            # 如果内容需要滚动且滚动到底部附近（last > 0.85），则显示按钮，否则隐藏
            if need_scroll and last > 0.85:
                self.top_btn.place(relx=0.95, rely=0.95, anchor="se")
            else:
                self.top_btn.place_forget()
        except:
            self.top_btn.place_forget()

    def scroll_to_top(self):
        """滚动到顶部并隐藏按钮"""
        self.knowledge_content.yview_moveto(0)
        self.top_btn.place_forget()

    # ---------------------- AI全科答疑 ----------------------
    def build_ai_answer_frame(self):
        frame = tk.Frame(self.content_container, bg=COLORS["bg_main"])

        if self.root.winfo_screenheight() >= 900:
            tk.Label(frame, text="🤖 AI全科答疑", font=FONTS["title"], bg=COLORS["bg_main"],
                     fg=COLORS["text_main"]).pack(
                pady=8, anchor="w")

        filter_frame = tk.LabelFrame(frame, text="答疑设置", font=FONTS["content_bold"], bg=COLORS["bg_card"],
                                     fg=COLORS["primary"])
        filter_frame.pack(fill=tk.X, pady=8, padx=2)

        row1 = tk.Frame(filter_frame, bg=COLORS["bg_card"])
        row1.pack(fill=tk.X, padx=15, pady=10)

        tk.Label(row1, text="科目：", font=FONTS["content_bold"], bg=COLORS["bg_card"]).pack(side=tk.LEFT)
        self.answer_subject_var = tk.StringVar(value="语文")
        subject_combo = ttk.Combobox(row1, textvariable=self.answer_subject_var, values=ALL_SUBJECTS[1:],
                                     font=FONTS["combo"], width=8, state="readonly")
        subject_combo.pack(side=tk.LEFT, padx=6)

        tk.Label(row1, text="年级：", font=FONTS["content_bold"], bg=COLORS["bg_card"]).pack(side=tk.LEFT, padx=12)
        self.answer_grade_var = tk.StringVar(value="七年级")
        grade_combo = ttk.Combobox(row1, textvariable=self.answer_grade_var, values=ALL_GRADES, font=FONTS["combo"],
                                   width=10, state="readonly")
        grade_combo.pack(side=tk.LEFT, padx=6)

        tk.Label(frame, text="请输入你的问题：", font=FONTS["content_bold"], bg=COLORS["bg_main"]).pack(anchor="w",
                                                                                                       pady=6)
        self.answer_input = scrolledtext.ScrolledText(frame, font=FONTS["content"], height=3, wrap=tk.WORD, padx=15,
                                                      pady=15)
        self.answer_input.pack(fill=tk.X, pady=6)
        self.bind_right_click(self.answer_input)

        self.input_placeholder = "💡 例如：《七年级上册 第一单元 四季美景 - 雨的四季》有哪些写作手法？"
        self._setup_placeholder(self.answer_input, self.input_placeholder)

        btn_bar = tk.Frame(frame, bg=COLORS["bg_main"])
        btn_bar.pack(fill=tk.X, pady=6)

        self.send_answer_btn = tk.Button(
            btn_bar, text="📤 发送提问", font=FONTS["btn"],
            bg=COLORS["primary"], fg="white", bd=0, relief=tk.FLAT, cursor="hand2",
            width=14, height=1, command=self.send_ai_answer
        )
        self.send_answer_btn.pack(side=tk.LEFT, padx=6)

        clear_btn = tk.Button(
            btn_bar, text="🗑️ 清空输入", font=FONTS["btn"],
            bg=COLORS["warning"], fg="white", bd=0, relief=tk.FLAT, cursor="hand2",
            width=12, height=1, command=self._clear_input
        )
        clear_btn.pack(side=tk.LEFT, padx=6)

        self.speak_answer_btn = tk.Button(
            btn_bar, text="🔊 朗读答案", font=FONTS["btn"],
            bg=COLORS["success"], fg="white", bd=0, relief=tk.FLAT, cursor="hand2",
            width=14, height=1, command=self.speak_answer_content
        )
        self.speak_answer_btn.pack(side=tk.LEFT, padx=6)

        self.stop_speak_btn = tk.Button(
            btn_bar, text="⏹️ 停止朗读", font=FONTS["btn"],
            bg=COLORS["danger"], fg="white", bd=0, relief=tk.FLAT, cursor="hand2",
            width=14, height=1, command=self.stop_tts
        )
        self.stop_speak_btn.pack(side=tk.LEFT, padx=6)

        # 字号调节按钮
        tk.Button(btn_bar, text="A+", font=FONTS["btn"], bg=COLORS["warning"], fg="white", bd=0,
                  relief=tk.FLAT, cursor="hand2", width=4, command=lambda: self.change_answer_font_size(2)).pack(
            side=tk.LEFT, padx=6)
        tk.Button(btn_bar, text="A-", font=FONTS["btn"], bg=COLORS["warning"], fg="white", bd=0,
                  relief=tk.FLAT, cursor="hand2", width=4, command=lambda: self.change_answer_font_size(-2)).pack(
            side=tk.LEFT, padx=6)

        # 状态标签（朗读状态）
        self.answer_status_label = tk.Label(btn_bar, text="", font=FONTS["content"], bg=COLORS["bg_main"],
                                            fg=COLORS["primary"])
        self.answer_status_label.pack(side=tk.LEFT, padx=6)
        self.answer_status_label.pack_forget()

        tk.Label(frame, text="AI解答：", font=FONTS["content_bold"], bg=COLORS["bg_main"]).pack(anchor="w", pady=6)
        self.answer_output = scrolledtext.ScrolledText(frame, font=FONTS["ai_output"], wrap=tk.WORD, padx=15, pady=15)
        self.answer_output.pack(fill=tk.BOTH, expand=True, pady=6)
        self.answer_output.config(state=tk.DISABLED)
        self.bind_right_click(self.answer_output)
        self._set_output_placeholder()

        return frame

    def _setup_placeholder(self, text_widget, placeholder):
        text_widget.insert(1.0, placeholder)
        text_widget.config(fg="gray", font=FONTS["placeholder"])

        def on_focus_in(event):
            if text_widget.get(1.0, "end-1c") == placeholder:
                text_widget.delete(1.0, tk.END)
                text_widget.config(fg="black", font=FONTS["content"])

        def on_focus_out(event):
            if text_widget.get(1.0, "end-1c").strip() == "":
                text_widget.delete(1.0, tk.END)
                text_widget.insert(1.0, placeholder)
                text_widget.config(fg="gray", font=FONTS["placeholder"])

        text_widget.bind("<FocusIn>", on_focus_in)
        text_widget.bind("<FocusOut>", on_focus_out)

    def _clear_input(self):
        # 清空输入框
        self.answer_input.delete(1.0, tk.END)
        self.answer_input.insert(1.0, self.input_placeholder)
        self.answer_input.config(fg="gray", font=FONTS["placeholder"])
        # 直接清空解答区域
        self.answer_output.config(state=tk.NORMAL)
        self.answer_output.delete(1.0, tk.END)
        self.answer_output.insert(1.0, "💡 点击【📤 发送提问】后，AI的解答将显示在这里...")
        self.answer_output.config(fg="gray", font=FONTS["placeholder"], state=tk.DISABLED)

    def _set_output_placeholder(self):
        self.answer_output.config(state=tk.NORMAL)
        self.answer_output.insert(1.0, "💡 点击【📤 发送提问】后，AI的解答将显示在这里...")
        self.answer_output.config(fg="gray", font=FONTS["placeholder"], state=tk.DISABLED)

    def send_ai_answer(self):
        question = self.answer_input.get(1.0, tk.END).strip()
        if question == self.input_placeholder or not question:
            messagebox.showwarning("提示", "请输入有效的问题！")
            return

        # ---------- 新增：乱发信息检测 ----------
        # 去除空格和标点后，汉字字母数字少于3个，或者全是无意义字符
        import re
        clean_text = re.sub(r'[\s，。？！；：“”‘’《》【】、]', '', question)
        if len(clean_text) < 3 or not re.search(r'[\u4e00-\u9fa5a-zA-Z0-9]', clean_text):
            self._update_answer_text("⚠️ 请提出具体的学习问题（至少3个有效字符），乱发信息不利于你的进步。")
            return
        # 可选：检测是否为重复无意义内容（如“哈哈哈”）
        if re.fullmatch(r'([哈呀哦嗯]+)\1*', clean_text):
            self._update_answer_text("⚠️ 请认真提问，重复无意义内容不会被解答。")
            return
        # ------------------------------------

        subject = self.answer_subject_var.get()
        grade = self.answer_grade_var.get()

        if not self.ai.api_key:
            messagebox.showwarning("提示", "请先去【系统设置】配置AI API Key！")
            return

        self.send_answer_btn.config(state=tk.DISABLED, text="📤 发送中...")
        self._update_answer_text("正在思考中，请稍候...")

        def on_answer_finish(success, result):
            self.send_answer_btn.config(state=tk.NORMAL, text="📤 发送提问")
            self._update_answer_text(result)
            self.refresh_records()

        def thread_func():
            success, result = self.ai.ai_request(subject, grade, "答疑", question)
            self.root.after(0, lambda: on_answer_finish(success, result))

        threading.Thread(target=thread_func, daemon=True).start()

    def _update_answer_text(self, text):
        self.answer_output.config(state=tk.NORMAL)
        self.answer_output.delete(1.0, tk.END)
        self.answer_output.insert(tk.END, text)
        self.answer_output.config(fg="black", font=FONTS["ai_output"], state=tk.DISABLED)

    def speak_answer_content(self):
        text = self.answer_output.get(1.0, tk.END).strip()
        print(f"朗读内容：{text}")  # 调试输出
        if not TTS_AVAILABLE:
            messagebox.showerror("错误", "当前系统不支持离线语音朗读功能，其他功能均可正常使用。")
            return
        if not text or text == "💡 点击“发送提问”后，AI的解答将显示在这里...":
            messagebox.showinfo("提示", "没有可朗读的内容，请先发送一个问题。")
            return
        self.speak_text(text)

    def change_answer_font_size(self, delta):
        """调节AI解答区域的字体大小"""
        current = self.answer_output.cget("font")
        if isinstance(current, tuple):
            family, size = current[0], current[1]
        else:
            family = "微软雅黑"
            size = 16
        new_size = max(10, min(28, size + delta))
        self.answer_output.config(font=(family, new_size))

    # ---------------------- AI随堂答题 ----------------------
    def reset_exam_content(self):
        self.exam_running = False
        self.current_exam_idx = 0
        self.exam_correct_num = 0
        self.current_exam_data = []
        self.exam_progress_label.config(text="")
        self._set_exam_placeholder()
        self._set_result_placeholder()
        for widget in self.exam_option_inner.winfo_children():
            widget.destroy()
        qtype = self.exam_qtype_var.get()
        if qtype == "选择题":
            placeholder_text = "💡 生成题目后，选项将显示在这里..."
        else:
            placeholder_text = "💡 生成题目后，输入框将显示在这里..."
        self.option_placeholder = tk.Label(self.exam_option_inner, text=placeholder_text,
                                           font=FONTS["placeholder"], fg="gray", bg=COLORS["bg_main"], anchor="w")
        self.option_placeholder.pack(pady=20, anchor="w")

        self.option_title_label.config(text="💡 请选择答案：")
        self.option_title_label.pack(fill=tk.X, pady=4)
        self.submit_exam_btn.config(state=tk.DISABLED)
        self.next_exam_btn.config(state=tk.DISABLED)

    def build_ai_exam_frame(self):
        frame = tk.Frame(self.content_container, bg=COLORS["bg_main"])

        # 根据屏幕高度决定是否显示标题（1080p显示，1366隐藏）
        if self.root.winfo_screenheight() >= 900:
            tk.Label(frame, text="📝 AI随堂答题", font=FONTS["title"], bg=COLORS["bg_main"], fg=COLORS["text_main"]).pack(
                pady=4, anchor="w")

        # ========== 1. 顶部答题设置区域 ==========
        setting_frame = tk.LabelFrame(frame, text="答题设置", font=FONTS["content_bold"], bg=COLORS["bg_card"],
                                      fg=COLORS["primary"])
        setting_frame.pack(fill=tk.X, padx=10, pady=5)

        row1 = tk.Frame(setting_frame, bg=COLORS["bg_card"])
        row1.pack(fill=tk.X, padx=15, pady=8)

        tk.Label(row1, text="科目：", font=FONTS["content_bold"], bg=COLORS["bg_card"]).pack(side=tk.LEFT)
        subject_combo = ttk.Combobox(row1, textvariable=self.exam_subject_var, values=ALL_SUBJECTS[1:],
                                     font=FONTS["combo"], width=12, state="readonly")
        subject_combo.pack(side=tk.LEFT, padx=6)
        subject_combo.bind("<<ComboboxSelected>>", lambda e: self.update_exam_chapter_options())

        tk.Label(row1, text="年级：", font=FONTS["content_bold"], bg=COLORS["bg_card"]).pack(side=tk.LEFT, padx=12)
        grade_combo = ttk.Combobox(row1, textvariable=self.exam_grade_var, values=ALL_GRADES, font=FONTS["combo"],
                                   width=10, state="readonly")
        grade_combo.pack(side=tk.LEFT, padx=6)
        grade_combo.bind("<<ComboboxSelected>>", lambda e: self.update_exam_chapter_options())

        tk.Label(row1, text="单元：", font=FONTS["content_bold"], bg=COLORS["bg_card"]).pack(side=tk.LEFT, padx=12)
        self.exam_chapter_combo = ttk.Combobox(row1, textvariable=self.exam_chapter_var, values=[], font=FONTS["combo"],
                                               width=28, state="readonly")
        self.exam_chapter_combo.pack(side=tk.LEFT, padx=6)
        self.exam_chapter_combo.bind("<<ComboboxSelected>>", lambda e: self.update_exam_lesson_options())

        tk.Label(row1, text="课时：", font=FONTS["content_bold"], bg=COLORS["bg_card"]).pack(side=tk.LEFT, padx=12)
        self.exam_lesson_combo = ttk.Combobox(row1, textvariable=self.exam_lesson_var, values=["整个章节"],
                                              font=FONTS["combo"], width=20, state="readonly")
        self.exam_lesson_combo.pack(side=tk.LEFT, padx=6)

        row2 = tk.Frame(setting_frame, bg=COLORS["bg_card"])
        row2.pack(fill=tk.X, padx=15, pady=8)

        tk.Label(row2, text="题量：", font=FONTS["content_bold"], bg=COLORS["bg_card"]).pack(side=tk.LEFT)
        num_combo = ttk.Combobox(row2, textvariable=self.exam_num_var, values=ALL_EXAM_NUMS, font=FONTS["combo"],
                                 width=6, state="readonly")
        num_combo.pack(side=tk.LEFT, padx=6)

        tk.Label(row2, text="难度：", font=FONTS["content_bold"], bg=COLORS["bg_card"]).pack(side=tk.LEFT, padx=12)
        diff_combo = ttk.Combobox(row2, textvariable=self.exam_diff_var, values=ALL_DIFFICULTIES, font=FONTS["combo"],
                                  width=7, state="readonly")
        diff_combo.pack(side=tk.LEFT, padx=6)

        tk.Label(row2, text="题型：", font=FONTS["content_bold"], bg=COLORS["bg_card"]).pack(side=tk.LEFT, padx=12)
        qtype_combo = ttk.Combobox(row2, textvariable=self.exam_qtype_var, values=ALL_QTYPES, font=FONTS["combo"],
                                   width=10, state="readonly")
        qtype_combo.pack(side=tk.LEFT, padx=6)
        qtype_combo.bind("<<ComboboxSelected>>", lambda e: self.reset_exam_content())

        tk.Label(row2, text="考点：", font=FONTS["content_bold"], bg=COLORS["bg_card"]).pack(side=tk.LEFT, padx=12)
        self.exam_kp_combo = ttk.Combobox(row2, textvariable=self.exam_kp_var, values=["随机"],
                                          font=FONTS["combo"], width=16, state="readonly")
        self.exam_kp_combo.pack(side=tk.LEFT, padx=6)

        self.generate_exam_btn = tk.Button(
            row2, text="📝 开始生成题目", font=FONTS["btn"],
            bg=COLORS["primary"], fg="white", bd=0, relief=tk.FLAT, cursor="hand2",
            width=18, height=1, command=self.generate_ai_exam
        )
        self.generate_exam_btn.pack(side=tk.LEFT, padx=20)

        # ========== 2. 底部按钮区域 ==========
        btn_frame = tk.Frame(frame, bg=COLORS["bg_main"])
        btn_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=8)
        btn_frame.grid_columnconfigure(0, weight=1)
        btn_frame.grid_columnconfigure(1, weight=1)

        button_container = tk.Frame(btn_frame, bg=COLORS["bg_main"])
        button_container.grid(row=0, column=0, columnspan=2)

        self.submit_exam_btn = tk.Button(button_container, text="✅ 提交答案", font=FONTS["btn"],
                                         bg=COLORS["primary"], fg="white", bd=0, relief=tk.FLAT, cursor="hand2",
                                         width=14, height=1, command=self.submit_exam_answer, state=tk.DISABLED)
        self.submit_exam_btn.pack(side=tk.LEFT, padx=20)

        self.next_exam_btn = tk.Button(button_container, text="➡️ 下一题", font=FONTS["btn"],
                                       bg=COLORS["success"], fg="white", bd=0, relief=tk.FLAT, cursor="hand2",
                                       width=14, height=1, command=self.load_next_exam_question, state=tk.DISABLED)
        self.next_exam_btn.pack(side=tk.LEFT, padx=20)

        # ========== 3. 中间内容区域 ==========
        middle_frame = tk.Frame(frame, bg=COLORS["bg_main"])
        middle_frame.pack(fill=tk.BOTH, expand=True, pady=6)

        self.exam_progress_label = tk.Label(middle_frame, text="", font=FONTS["content_bold"], bg=COLORS["bg_main"])
        self.exam_progress_label.pack(fill=tk.X, pady=2)

        self.exam_question_display = scrolledtext.ScrolledText(middle_frame, font=FONTS["content_bold"],
                                                               wrap=tk.WORD, padx=15, pady=15, height=5)
        self.exam_question_display.pack(fill=tk.X, pady=6)
        self._set_exam_placeholder()
        self.bind_right_click(self.exam_question_display)

        # 左右两栏：左侧占3份，右侧占2份
        split_frame = tk.Frame(middle_frame, bg=COLORS["bg_main"])
        split_frame.pack(fill=tk.BOTH, expand=True, pady=6)
        # 【关键设置】设置权重为 3:2，并强制最小宽度，防止左侧太窄
        split_frame.grid_columnconfigure(0, weight=3, minsize=500)  # 左侧最小600像素
        split_frame.grid_columnconfigure(1, weight=2, minsize=300)  # 右侧最小300像素
        split_frame.grid_rowconfigure(0, weight=1)
        # --- 左侧选项区 ---
        left_frame = tk.Frame(split_frame, bg=COLORS["bg_main"], relief=tk.GROOVE, bd=2)
        # 注意：这里必须加上 sticky="nsew"，否则填不满
        left_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5), pady=4)
        self.option_title_label = tk.Label(left_frame, text="💡 请选择答案：", font=FONTS["content_bold"],
                                           bg=COLORS["bg_main"], fg=COLORS["primary"], anchor="w")
        self.option_title_label.pack(fill=tk.X, pady=4)
        # 创建一个包裹 Frame 用于放置 Canvas 和滚动条
        canvas_frame = tk.Frame(left_frame, bg=COLORS["bg_main"])
        canvas_frame.pack(fill=tk.BOTH, expand=True)
        option_canvas = tk.Canvas(canvas_frame, bg=COLORS["bg_main"], highlightthickness=0)
        option_scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=option_canvas.yview)
        option_canvas.configure(yscrollcommand=option_scrollbar.set)
        # 滚动条靠右
        option_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        # 画布填满左侧
        option_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.exam_option_inner = tk.Frame(option_canvas, bg=COLORS["bg_main"])
        # 【核心修复】保存画布窗口ID，并绑定宽度自适应事件
        canvas_window_id = option_canvas.create_window((0, 0), window=self.exam_option_inner, anchor="nw")

        def on_option_configure(event):
            # 当画布大小改变时，强制让内部选项区域的宽度等于画布宽度
            # 这样文字就不会被遮挡了
            option_canvas.itemconfig(canvas_window_id, width=event.width)
            option_canvas.configure(scrollregion=option_canvas.bbox("all"))

        # 绑定画布大小改变事件
        option_canvas.bind("<Configure>", on_option_configure)
        # 绑定内容改变事件（更新滚动范围）
        self.exam_option_inner.bind("<Configure>",
                                    lambda e: option_canvas.configure(scrollregion=option_canvas.bbox("all")))
        # 下面是占位符代码，保持不变
        qtype = self.exam_qtype_var.get()
        if qtype == "选择题":
            placeholder_text = "💡 生成题目后，选项将显示在这里..."
        else:
            placeholder_text = "💡 生成题目后，输入框将显示在这里..."
        self.option_placeholder = tk.Label(self.exam_option_inner, text=placeholder_text,
                                           font=FONTS["placeholder"], fg="gray", bg=COLORS["bg_main"], anchor="w")
        self.option_placeholder.pack(pady=20, anchor="w")
        self.exam_user_answer = tk.StringVar()
        self.exam_answer_entry = None
        self.exam_radio_buttons = []

        # 右侧解析区（自适应宽度）
        right_frame = tk.Frame(split_frame, bg=COLORS["bg_main"], relief=tk.GROOVE, bd=2)
        right_frame.grid(row=0, column=1, sticky="nsew", padx=5, pady=4)

        tk.Label(right_frame, text="📝 解析：", font=FONTS["content_bold"], bg=COLORS["bg_main"], anchor="w").pack(fill=tk.X, pady=4)
        self.exam_result_text = scrolledtext.ScrolledText(right_frame, font=FONTS["ai_output"], wrap=tk.WORD,
                                                          padx=15, pady=15)
        self.exam_result_text.pack(fill=tk.BOTH, expand=True, pady=6)
        self.bind_right_click(self.exam_result_text)
        self.exam_result_text.tag_configure("score", font=("微软雅黑", 16, "bold"), foreground=COLORS["success"])
        self.exam_result_text.tag_configure("wrong", font=("微软雅黑", 16, "bold"), foreground=COLORS["danger"])
        self.exam_result_text.tag_configure("knowledge", font=("微软雅黑", 14, "bold"), foreground=COLORS["primary"])
        self.exam_result_text.tag_configure("score_big", font=("微软雅黑", 18, "bold"), foreground=COLORS["success"])
        self._set_result_placeholder()

        # 初始化下拉框
        self.update_exam_chapter_options()
        # 如果 KNOWLEDGE_POINTS 未定义，请注释下一行
        # self.update_kp_combo_options()
        return frame

    def _set_result_placeholder(self):
        self.exam_result_text.config(state=tk.NORMAL)
        self.exam_result_text.delete(1.0, tk.END)
        self.exam_result_text.insert(1.0, "💡 提交答案后，AI解析将显示在这里...")
        self.exam_result_text.config(fg="gray", font=FONTS["placeholder"])
        self.exam_result_text.config(state=tk.DISABLED)

    def _clear_result_placeholder(self):
        self.exam_result_text.config(state=tk.NORMAL)
        current = self.exam_result_text.get(1.0, tk.END).strip()
        if current == "💡 提交答案后，AI解析将显示在这里...":
            self.exam_result_text.delete(1.0, tk.END)
            self.exam_result_text.config(fg="black", font=FONTS["ai_output"])
        self.exam_result_text.config(state=tk.DISABLED)

    def _set_exam_placeholder(self):
        self.exam_question_display.config(state=tk.NORMAL)
        self.exam_question_display.delete(1.0, tk.END)
        self.exam_question_display.insert(1.0,
                                          "💡 请点击上方【开始生成题目】按钮，AI将根据您的选择生成题目。\n\n预计等待：10-20秒。")
        self.exam_question_display.config(fg="gray", font=FONTS["placeholder"], state=tk.DISABLED)

    def _update_exam_question_text(self, text, is_placeholder=False):
        self.exam_question_display.config(state=tk.NORMAL)
        self.exam_question_display.delete(1.0, tk.END)
        self.exam_question_display.insert(tk.END, text)
        if is_placeholder:
            self.exam_question_display.config(fg="gray", font=FONTS["placeholder"])
        else:
            self.exam_question_display.config(fg="black", font=FONTS["content"])
        self.exam_question_display.config(state=tk.DISABLED)

    def update_exam_chapter_options(self):
        subject = self.exam_subject_var.get()
        grade = self.exam_grade_var.get()
        chapters = self.db.get_chapters_by_subject_grade(subject, grade)
        self.exam_chapter_combo['values'] = chapters
        if chapters:
            self.exam_chapter_var.set(chapters[0])
            self.update_exam_lesson_options()
        self.update_kp_combo_options()

    def update_exam_lesson_options(self):
        subject = self.exam_subject_var.get()
        grade = self.exam_grade_var.get()
        chapter = self.exam_chapter_var.get()
        lessons = self.db.get_lessons_by_chapter(subject, grade, chapter)
        self.exam_lesson_combo['values'] = lessons
        self.exam_lesson_var.set(lessons[0])
        self.update_kp_combo_options()

    def update_kp_combo_options(self):
        subject = self.exam_subject_var.get()
        grade = self.exam_grade_var.get()
        chapter = self.exam_chapter_var.get()
        lesson = self.exam_lesson_var.get().strip()  # 去除首尾空格

        filtered_names = ["随机"]
        self.kp_name_to_id = {}
        for kp_id, info in KNOWLEDGE_POINTS.items():
            # 科目、年级、章节匹配
            if info['subject'] == subject and info['grade'] == grade:
                if info['chapter'] == chapter or info['chapter'] == "" or info['chapter'] == "整个章节":
                    kp_lesson = info['lesson'].strip() if info['lesson'] else ""
                    if lesson == "整个章节":
                        # 显示该章节下所有考点
                        filtered_names.append(info['name'])
                        self.kp_name_to_id[info['name']] = kp_id
                    else:
                        # 显示当前课时的考点 或 该章节的综合考点
                        if kp_lesson == lesson or kp_lesson == "整个章节":
                            filtered_names.append(info['name'])
                            self.kp_name_to_id[info['name']] = kp_id
        # 去重（保持顺序）
        filtered_names = list(dict.fromkeys(filtered_names))
        self.exam_kp_combo['values'] = filtered_names
        # 如果当前选中的考点名称不在新列表中，重置为“随机”
        if self.exam_kp_var.get() not in filtered_names:
            self.exam_kp_var.set("随机")

    def parse_question_block(self, block):
        lines = [line.strip() for line in block.split("\n") if line.strip()]
        q_data = {"chapter": "", "question": "", "options": "", "answer": "", "analysis": ""}
        current_field = None
        for line in lines:
            if line.startswith("【对应章节】"):
                current_field = "chapter"
                q_data[current_field] = line.replace("【对应章节】", "").strip()
            elif line.startswith("【题目】"):
                current_field = "question"
                q_data[current_field] = line.replace("【题目】", "").strip()
            elif line.startswith("【选项】"):
                current_field = "options"
                q_data[current_field] = line.replace("【选项】", "").strip()
            elif line.startswith("【答案】"):
                current_field = "answer"
                q_data[current_field] = line.replace("【答案】", "").strip()
            elif line.startswith("【解析】"):
                current_field = "analysis"
                q_data[current_field] = line.replace("【解析】", "").strip()
            elif current_field:
                if current_field == "options":
                    q_data[current_field] += "\n" + line
                else:
                    q_data[current_field] += "\n" + line
        return q_data

    def _clean_analysis_text(self, text):
        if not text:
            return "暂无解析"
        cleaned = re.sub(r'(?m)^[A-Z][\.\、]\s*.*$', '', text)
        cleaned = re.sub(r'(?m)^[A-Z]\s+.*$', '', cleaned)
        cleaned = re.sub(r'\b[A-Z]\s*[\.\、]\s*', '', cleaned)
        cleaned = re.sub(r'\n\s*\n', '\n\n', cleaned).strip()
        return cleaned if cleaned else "暂无解析"

    def generate_ai_exam(self):
        if self.is_loading_exam:
            return
        if not self.ai.api_key:
            messagebox.showwarning("提示", "请先去【系统设置】配置AI API Key！")
            return

        exam_subject = self.exam_subject_var.get()
        exam_grade = self.exam_grade_var.get()
        exam_chapter = self.exam_chapter_var.get()
        exam_lesson = self.exam_lesson_var.get()
        exam_total_num = int(self.exam_num_var.get())
        q_type = self.exam_qtype_var.get()
        kp_choice = self.exam_kp_var.get()

        self.exam_running = False
        self.current_exam_idx = 0
        self.exam_correct_num = 0
        self.exam_user_answer.set("")

        self.exam_result_text.config(state=tk.NORMAL)
        self.exam_result_text.delete(1.0, tk.END)
        self.exam_result_text.insert(1.0, "⏳ 等待AI出题完成，提交答案后，解析将显示在这里...")
        self.exam_result_text.config(fg="gray", font=FONTS["placeholder"])
        self.exam_result_text.config(state=tk.DISABLED)

        self.submit_exam_btn.config(state=tk.DISABLED)
        self.next_exam_btn.config(state=tk.DISABLED)

        self.is_loading_exam = True
        self.generate_exam_btn.config(state=tk.DISABLED, text="🤖 AI出题中...")
        self._update_exam_question_text("💡 正在AI生成题目，请耐心等待（约10-20秒）...", is_placeholder=True)

        full_name = exam_chapter
        if exam_lesson != "整个章节":
            full_name = f"{exam_chapter} - {exam_lesson}"
        chapter_prompt = f"限定章节：{full_name}" if full_name else ""

        kp_prompt = ""

        selected_kp_id = None
        if kp_choice != "随机":
            # 从映射中获取考点编号
            kp_id = self.kp_name_to_id.get(kp_choice)
            if kp_id and kp_id in KNOWLEDGE_POINTS:
                kp_info = KNOWLEDGE_POINTS[kp_id]
                kp_name = kp_info["name"]
                kp_desc = kp_info["desc"]
                kp_prompt = f"限定考点：{kp_name}。{kp_desc}"
                selected_kp_id = kp_id

        if q_type == "选择题":
            format_part = """必须严格按照以下格式输出，只生成选择题：
    【对应章节】xxx
    【题目】xxx
    【选项】A. xxx B. xxx C. xxx D. xxx（选项可换行）
    【答案】xxx
    【解析】考点：简短说明（不超过20字）。详细解析内容...
    题目之间空行分隔。"""
        elif q_type == "填空题":
            format_part = """必须严格按照以下格式输出，只生成填空题：
    【对应章节】xxx
    【题目】xxx（用____表示填空处，绝对不能有选项）
    【答案】xxx
    【解析】考点：简短说明（不超过20字）。详细解析内容...
    题目之间空行分隔。"""
        elif q_type == "简答题":
            format_part = """必须严格按照以下格式输出，只生成简答题：
    【对应章节】xxx
    【题目】xxx
    【答案】详细标准答案，绝对不能有选项
    【解析】考点：简短说明（不超过20字）。详细解析内容...
    题目之间空行分隔。"""
        else:
            format_part = """必须严格按照以下格式输出：
    【对应章节】xxx
    【题目】xxx
    【答案】xxx
    【解析】考点：简短说明。详细解析内容...
    题目之间空行分隔。"""

        if kp_prompt:
            prompt = f"请生成{exam_total_num}道{exam_grade}{exam_subject}的{q_type}，{chapter_prompt}，{kp_prompt}，难度{self.exam_diff_var.get()}。{format_part}"
        else:
            prompt = f"请生成{exam_total_num}道{exam_grade}{exam_subject}的{q_type}，{chapter_prompt}，难度{self.exam_diff_var.get()}。{format_part}"

        def on_exam_generated(success, result):
            self.is_loading_exam = False
            self.generate_exam_btn.config(state=tk.NORMAL, text="📝 开始生成题目")
            if not success:
                self._update_exam_question_text(f"生成失败：{result}\n请检查配置后重试。")
                return
            blocks = result.strip().split("\n\n")
            self.current_exam_data = []
            for block in blocks:
                if not block.strip(): continue
                q_data = self.parse_question_block(block)
                if q_data.get("question"):
                    q_data['selected_kp_id'] = selected_kp_id
                    q_data['question_type'] = q_type
                    self.current_exam_data.append(q_data)
            if len(self.current_exam_data) == 0:
                self._update_exam_question_text("AI返回格式异常，无法解析题目，请重试。")
                return
            self.exam_running = True
            self.submit_exam_btn.config(state=tk.NORMAL)
            self.load_next_exam_question()

        def thread_func():
            success, result = self.ai.ai_request(exam_subject, exam_grade, "出题", prompt, auto_save=False,
                                                 temperature=0.8)
            self.root.after(0, lambda: on_exam_generated(success, result))

        threading.Thread(target=thread_func, daemon=True).start()

    def load_next_exam_question(self):
        if self.current_exam_idx >= len(self.current_exam_data):
            self.db.save_exam_record(self.exam_subject_var.get(), self.exam_grade_var.get(),
                                     self.exam_chapter_var.get(), self.exam_lesson_var.get(),
                                     len(self.current_exam_data), self.exam_correct_num)
            messagebox.showinfo("答题完成",
                                f"结束！\n总题数：{len(self.current_exam_data)}\n答对：{self.exam_correct_num}\n正确率：{(self.exam_correct_num / len(self.current_exam_data)) * 100:.1f}%")
            self.exam_running = False
            self.exam_progress_label.config(text="")
            self._set_exam_placeholder()
            self._set_result_placeholder()
            self.submit_exam_btn.config(state=tk.DISABLED)
            self.next_exam_btn.config(state=tk.DISABLED)
            for widget in self.exam_option_inner.winfo_children():
                widget.destroy()

            # 隐藏左侧标题（选项出现后标题不再需要）
            # self.option_title_label.pack_forget()

            qtype = self.exam_qtype_var.get()
            if qtype == "选择题":
                placeholder_text = "💡 生成题目后，选项将显示在这里..."
            else:
                placeholder_text = "💡 生成题目后，输入框将显示在这里..."
            self.option_placeholder = tk.Label(self.exam_option_inner, text=placeholder_text,
                                               font=FONTS["placeholder"], fg="gray", bg=COLORS["bg_main"], anchor="w")
            self.option_placeholder.pack(pady=20, anchor="w")

            self.option_title_label.config(text="💡 请选择答案：")
            self.option_title_label.pack(fill=tk.X, pady=4)
            self.refresh_records()
            return

        current_q = self.current_exam_data[self.current_exam_idx]
        self.exam_progress_label.config(text=f"第{self.current_exam_idx + 1}题 / 共{len(self.current_exam_data)}题")
        q_text = f"【对应章节】{current_q.get('chapter', '未指定')}\n\n【题目】{current_q['question']}"
        self._update_exam_question_text(q_text)

        self._set_result_placeholder()
        self.submit_exam_btn.config(text="✅ 提交答案", state=tk.NORMAL)
        self.next_exam_btn.config(state=tk.DISABLED)

        q_type = self.exam_qtype_var.get()
        for widget in self.exam_option_inner.winfo_children():
            widget.destroy()

        if q_type == "选择题":
            self.option_title_label.config(text="📌 请选择答案：")
            self.option_title_label.pack(fill=tk.X, pady=4)  # 改为 pack
            options_text = current_q.get('options', "")
            options = re.split(r'[A-Z]\s*[\.\、]\s*', options_text)[1:]
            self.exam_radio_buttons = []
            self.exam_user_answer.set("")
            for i, opt in enumerate(options):
                if opt.strip():
                    letter = chr(65 + i)
                    display_text = f"{letter}. {opt.strip()}"
                    rb = tk.Radiobutton(
                        self.exam_option_inner, text=display_text, variable=self.exam_user_answer,
                        value=chr(65 + i), font=FONTS["content"], bg=COLORS["bg_main"],
                        fg=COLORS["text_main"], anchor="w", cursor="hand2", justify="left",
                        wraplength=600, selectcolor="lightblue"
                    )
                    rb.pack(anchor="w", pady=4, fill=tk.X)
                    self.exam_radio_buttons.append(rb)
        else:
            self.option_title_label.config(text="✏️ 请在此输入答案：")
            self.option_title_label.pack(fill=tk.X, pady=4) # 改为 pack
            self.exam_answer_entry = scrolledtext.ScrolledText(self.exam_option_inner, font=FONTS["content"],
                                                               wrap=tk.CHAR, padx=12, pady=12)
            self.exam_answer_entry.insert(1.0, "💡 请在此输入您的答案...")
            self.exam_answer_entry.config(fg="gray", font=FONTS["placeholder"])

            def on_entry_focus_in(event):
                if self.exam_answer_entry.get(1.0, "end-1c") == "💡 请在此输入您的答案...":
                    self.exam_answer_entry.delete(1.0, tk.END)
                    self.exam_answer_entry.config(fg="black", font=FONTS["content"])

            def on_entry_focus_out(event):
                if self.exam_answer_entry.get(1.0, "end-1c").strip() == "":
                    self.exam_answer_entry.delete(1.0, tk.END)
                    self.exam_answer_entry.insert(1.0, "💡 请在此输入您的答案...")
                    self.exam_answer_entry.config(fg="gray", font=FONTS["placeholder"])

            self.exam_answer_entry.bind("<FocusIn>", on_entry_focus_in)
            self.exam_answer_entry.bind("<FocusOut>", on_entry_focus_out)
            self.exam_answer_entry.pack(fill=tk.BOTH, expand=True, pady=4)

    def submit_exam_answer(self):
        if not self.exam_running:
            return
        if self.is_grading:
            messagebox.showwarning("提示", "正在判题中，请稍候...")
            return

        current_q = self.current_exam_data[self.current_exam_idx]
        q_type = self.exam_qtype_var.get()
        subject = self.exam_subject_var.get()
        grade = self.exam_grade_var.get()

        user_answer = ""
        if q_type == "选择题":
            user_answer = self.exam_user_answer.get().strip()
            if not user_answer:
                messagebox.showwarning("提示", "请选择答案！")
                return
        else:
            user_answer = self.exam_answer_entry.get(1.0, tk.END).strip()
            if user_answer == "💡 请在此输入您的答案...":
                user_answer = ""
            if not user_answer:
                messagebox.showwarning("提示", "请输入答案！")
                return

        self.submit_exam_btn.config(state=tk.DISABLED, text="⏳ 判题中...")
        self.is_grading = True

        self.exam_result_text.config(state=tk.NORMAL)
        self.exam_result_text.delete(1.0, tk.END)
        self.exam_result_text.insert(1.0, "⏳ AI正在判题中，请稍候...")
        self.exam_result_text.config(fg="gray", font=FONTS["placeholder"])
        self.exam_result_text.config(state=tk.DISABLED)

        correct_answer = current_q.get('answer', '').strip()
        analysis = current_q.get('analysis', '')

        # 智能拆分考点与解析
        knowledge_point = ''
        clean_analysis = analysis
        match = re.search(r'考点[：:]\s*(.+?)(?:\n|$)', analysis)
        if match:
            full_content = match.group(1).strip()
            sentences = re.split(r'[。；！？]', full_content)
            if sentences:
                first_sentence = sentences[0].strip()
                if len(first_sentence) > 30:
                    knowledge_point = first_sentence[:30] + "..."
                else:
                    knowledge_point = first_sentence
                if len(sentences) > 1:
                    clean_analysis = '。'.join(sentences[1:]).strip()
                    if not clean_analysis:
                        clean_analysis = "（无详细解析）"
                else:
                    clean_analysis = "（详细解析见考点）"
            else:
                knowledge_point = full_content[:30]
                clean_analysis = "（无详细解析）"
        else:
            sentences = re.split(r'[。；！？]', analysis)
            if len(sentences) > 1:
                knowledge_point = sentences[0].strip()
                clean_analysis = '。'.join(sentences[1:]).strip()
            else:
                knowledge_point = analysis[:30] if analysis else "无"
                clean_analysis = "（无详细解析）"
        if not clean_analysis.strip():
            clean_analysis = "（无详细解析）"

        def get_full_answer(letter, options_text):
            if not options_text:
                return letter
            opts = re.split(r'[A-Z]\s*[\.\、]\s*', options_text)[1:]
            idx = ord(letter.upper()) - 65
            if 0 <= idx < len(opts):
                return f"{letter}. {opts[idx].strip()}"
            return letter

        praise_correct = [
            "太棒了！完全正确！", "厉害啦！答得真棒！", "完美！答案正确！", "太优秀了！继续保持！",
            "正确！做得很好！", "恭喜你答对了！", "真聪明！继续努力！", "了不起！完全正确！",
            "Excellent！完美！", "给你点赞！答得漂亮！", "思路清晰，完全正确！", "你真是个学霸！",
            "哇！太厉害了！", "又对一题，你真棒！", "学得真好，继续加油！", "完美解答，为你骄傲！"
        ]
        praise_wrong = [
            "答错了，没关系，再试试！", "哎呀，不对哦，看看解析吧！", "再想想，正确答案是这样的：",
            "别灰心，下次一定能对！", "加油！再复习一下这个知识点。", "差一点就对了，继续努力！",
            "哦豁，不对，我们来看看解析。", "再接再厉，下次一定行！", "错误是学习的好机会，看看解析吧。",
            "没关系，弄懂就好，下次就不会错啦！"
        ]

        def on_grade_finish(success, result):
            self.is_grading = False
            self.exam_result_text.config(state=tk.NORMAL)
            self.exam_result_text.config(fg="black")
            self.exam_result_text.delete(1.0, tk.END)

            if q_type == "选择题":
                is_correct = (user_answer.strip().upper() == correct_answer.strip().upper())
                full_correct = get_full_answer(correct_answer, current_q.get('options', ''))
            elif q_type == "填空题":
                is_correct = (success and "正确" in result)
                full_correct = correct_answer
            else:   # 简答题：兼容评分格式和自然语言格式
                if success:
                    # 尝试提取得分（格式：得分：X/10）
                    score_match = re.search(r'得分[:：]\s*(\d+)/10', result)
                    if score_match:
                        score = int(score_match.group(1))
                        is_correct = (score >= 6)
                        # 保留原始结果作为评语
                        analysis_result = result
                    else:
                        # 无得分格式，根据自然语言判断正确性
                        if any(keyword in result for keyword in ["完全正确", "答对了", "正确", "很棒", "不错", "优秀"]):
                            is_correct = True
                            analysis_result = f"得分：10/10\n理由：{result}"
                        elif any(keyword in result for keyword in ["错误", "不对", "答错了", "不正确"]):
                            is_correct = False
                            analysis_result = f"得分：0/10\n理由：{result}"
                        else:
                            # 无法判断时，默认错误
                            is_correct = False
                            analysis_result = f"得分：0/10\n理由：{result}"
                else:
                    is_correct = False
                    analysis_result = "判题失败"
                full_correct = correct_answer
                # 将处理后的结果赋值给 result，以便后续显示
                result = analysis_result

            if is_correct:
                self.exam_correct_num += 1
                praise = random.choice(praise_correct)
                self.exam_result_text.insert(tk.END, f"{praise}\n", "score")
                self.speak_text(praise)
            else:
                praise = random.choice(praise_wrong)
                self.exam_result_text.insert(tk.END, f"{praise}\n", "wrong")
                self.speak_text(praise)
                # 增加一个空行，拉大间距
                self.exam_result_text.insert(tk.END, "\n")

            # ========== 新增：显示简答题的评分和理由 ==========
            if q_type == "简答题" and 'analysis_result' in locals():
                # 提取“得分：X/10”部分
                score_match = re.search(r'(得分：\d+/10)', analysis_result)
                if score_match:
                    score_text = score_match.group(1)
                    # 去除得分部分，得到理由
                    reason_text = analysis_result.replace(score_text, '').strip()
                    # 插入得分（大字体）
                    self.exam_result_text.insert(tk.END, f"{score_text}\n", "score_big")
                    # 插入理由（普通字体）
                    if reason_text:
                        self.exam_result_text.insert(tk.END, f"{reason_text}\n\n")
                else:
                    self.exam_result_text.insert(tk.END, f"{analysis_result}\n\n")
            # ================================================

            if knowledge_point:
                self.exam_result_text.insert(tk.END, f"📌 考点：{knowledge_point}\n\n")
            if clean_analysis and clean_analysis != "（无详细解析）":
                self.exam_result_text.insert(tk.END, f"📖 解析：{clean_analysis}\n\n")
            self.exam_result_text.insert(tk.END, f"✅ 标准答案：{full_correct}")

            self.exam_result_text.config(state=tk.DISABLED)

            self._save_exam_detail(current_q, user_answer, is_correct, self.ai.class_name,
                                   subject, grade, self.exam_chapter_var.get(), self.exam_lesson_var.get())

            self.current_exam_idx += 1
            self.next_exam_btn.config(state=tk.NORMAL)

        def thread_func():
            try:
                if q_type == "选择题":
                    on_grade_finish(True, "")
                elif q_type == "填空题":
                    grade_prompt = f"题目：{current_q['question']}，标准答案：{correct_answer}，学生答案：{user_answer}。只返回正确/错误。"
                    success, result = self.ai.ai_request(subject, grade, "判题", grade_prompt, auto_save=False)
                    on_grade_finish(success, result)
                else:
                    grade_prompt = f"题目：{current_q['question']}，标准答案：{correct_answer}，学生答案：{user_answer}。请按照以下格式严格输出（不要输出其他任何内容）：得分：X/10 理由：xxx 示例：得分：8（学生得分）/10（满分） 理由：答案基本正确，但漏掉了一个关键步骤。"
                    success, result = self.ai.ai_request(subject, grade, "判题", grade_prompt, auto_save=False)
                    on_grade_finish(success, result)
            except Exception as e:
                self.root.after(0, lambda: on_grade_finish(False, f"判题异常：{str(e)}"))

        threading.Thread(target=thread_func, daemon=True).start()

    def _save_exam_detail(self, current_q, user_answer, is_correct, class_name, subject, grade, chapter, lesson):
        analysis = current_q.get('analysis', '')
        knowledge_point = ''
        knowledge_point_id = current_q.get('selected_kp_id', None)
        if not knowledge_point_id:
            match = re.search(r'考点[：:]\s*(.+?)(?:\n|$)', analysis)
            if match:
                knowledge_point = match.group(1).strip()
        else:
            kp_info = KNOWLEDGE_POINTS.get(knowledge_point_id, {})
            knowledge_point = kp_info.get('name', knowledge_point_id)

        question_type = current_q.get('question_type', self.exam_qtype_var.get())

        conn, cursor = self.db._connect()
        cursor.execute('''
            INSERT INTO exam_details 
            (exam_record_id, class_name, subject, grade, chapter, lesson, question, correct_answer, user_answer, is_correct, knowledge_point, knowledge_point_id, question_type)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (0, class_name, subject, grade, chapter, lesson,
              current_q['question'], current_q.get('answer', ''), user_answer, 1 if is_correct else 0,
              knowledge_point, knowledge_point_id, question_type))
        self.db._close(conn)

    # ---------------------- 学情分析模块 ----------------------
    def build_records_frame(self):
        frame = tk.Frame(self.content_container, bg=COLORS["bg_main"])

        tk.Label(frame, text="📊 学情分析", font=FONTS["title"], bg=COLORS["bg_main"],
                 fg=COLORS["text_main"]).pack(pady=6, anchor="w")

        # 设置 Notebook 标签页字体大小
        style = ttk.Style()
        style.configure("TNotebook.Tab", font=("微软雅黑", 12))

        notebook = ttk.Notebook(frame)
        notebook.pack(fill=tk.BOTH, expand=True, pady=5)

        # 子页1：班级学情分析
        analysis_frame = tk.Frame(notebook, bg=COLORS["bg_main"])
        notebook.add(analysis_frame, text="📊 班级学情分析")
        self._build_class_analysis_tab(analysis_frame)

        # 子页2：AI问答记录
        chat_frame = tk.Frame(notebook, bg=COLORS["bg_main"])
        notebook.add(chat_frame, text="💬 AI问答记录")
        self._build_chat_records_tab(chat_frame, skip_refresh=True)

        # 子页3：答题记录
        exam_rec_frame = tk.Frame(notebook, bg=COLORS["bg_main"])
        notebook.add(exam_rec_frame, text="📝 答题记录")
        self._build_exam_records_tab(exam_rec_frame, skip_refresh=True)

        # 子页4：学习动机测评 (MSLQ) - 内嵌优化版
        mslq_frame = tk.Frame(notebook, bg=COLORS["bg_main"])
        notebook.add(mslq_frame, text="📋 学习动机测评 (MSLQ)")
        self._build_mslq_tab_embedded(mslq_frame)

        self.refresh_records()
        return frame

    def _build_class_analysis_tab(self, parent):
        # 配置 Treeview 字体大小
        style = ttk.Style()
        style.configure("Treeview", font=("微软雅黑", 12), rowheight=25)
        style.configure("Treeview.Heading", font=("微软雅黑", 12, "bold"))

        # 顶部筛选栏
        filter_frame = tk.Frame(parent, bg=COLORS["bg_main"])
        filter_frame.pack(fill=tk.X, padx=10, pady=5)

        tk.Label(filter_frame, text="科目：", font=FONTS["content_bold"], bg=COLORS["bg_main"]).pack(side=tk.LEFT,
                                                                                                    padx=5)
        self.class_analysis_subject_var = tk.StringVar(value="语文")
        subject_combo = ttk.Combobox(filter_frame, textvariable=self.class_analysis_subject_var,
                                     values=ALL_SUBJECTS[1:], width=10, state="readonly",
                                     font=("微软雅黑", 12))
        subject_combo.pack(side=tk.LEFT, padx=5)

        tk.Label(filter_frame, text="年级：", font=FONTS["content_bold"], bg=COLORS["bg_main"]).pack(side=tk.LEFT,
                                                                                                    padx=10)
        self.class_analysis_grade_var = tk.StringVar(value="七年级")
        grade_combo = ttk.Combobox(filter_frame, textvariable=self.class_analysis_grade_var,
                                   values=ALL_GRADES, width=8, state="readonly",
                                   font=("微软雅黑", 12))
        grade_combo.pack(side=tk.LEFT, padx=5)

        refresh_btn = tk.Button(filter_frame, text="💡 生成教育建议", font=FONTS["btn"],
                                bg=COLORS["primary"], fg="white", bd=0, relief=tk.FLAT, cursor="hand2",
                                command=self.refresh_advice_only)  # 修改这里
        refresh_btn.pack(side=tk.LEFT, padx=10)

        clear_btn = tk.Button(filter_frame, text="🗑️ 清空当前班级数据", font=FONTS["btn"],
                              bg=COLORS["danger"], fg="white", bd=0, relief=tk.FLAT, cursor="hand2",
                              command=self.clear_class_data)
        clear_btn.pack(side=tk.RIGHT, padx=5)

        # 主内容区域：左右分割，左侧权重更大
        main_pane = tk.Frame(parent, bg=COLORS["bg_main"])
        main_pane.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        main_pane.grid_columnconfigure(0, weight=2, minsize=480)  # 左侧最小宽度 500 像素
        main_pane.grid_columnconfigure(1, weight=2)
        main_pane.grid_rowconfigure(0, weight=1)

        # 左侧：章节列表
        left_frame = tk.LabelFrame(main_pane, text="📊 章节正确率统计", font=FONTS["content_bold"],
                                   bg=COLORS["bg_card"], fg=COLORS["primary"])
        left_frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        left_frame.grid_rowconfigure(0, weight=1)
        left_frame.grid_columnconfigure(0, weight=1)

        tree_frame = tk.Frame(left_frame, bg=COLORS["bg_card"])
        tree_frame.pack(fill=tk.BOTH, expand=True)

        columns = ("chapter", "total", "correct", "rate")
        self.chapter_tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=15)
        self.chapter_tree.heading("chapter", text="章节")
        self.chapter_tree.heading("total", text="题目数")
        self.chapter_tree.heading("correct", text="正确数")
        self.chapter_tree.heading("rate", text="正确率")

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.chapter_tree.yview)
        self.chapter_tree.configure(yscrollcommand=vsb.set)
        self.chapter_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        def adjust_columns(event=None):
            # 强制更新界面以获取真实宽度
            tree_frame.update_idletasks()
            total_width = tree_frame.winfo_width()
            if total_width < 100:
                return
            # 调整比例：章节35%，题目数15%，正确数15%，正确率35%（确保正确率有足够空间）
            self.chapter_tree.column("chapter", width=int(total_width * 0.50), minwidth=150)
            self.chapter_tree.column("total", width=int(total_width * 0.20), minwidth=60)
            self.chapter_tree.column("correct", width=int(total_width * 0.15), minwidth=60)
            self.chapter_tree.column("rate", width=int(total_width * 0.15), minwidth=60)

        tree_frame.bind("<Configure>", adjust_columns)
        # 延迟调用并强制刷新，确保初始宽度正确
        self.root.after(100, lambda: (tree_frame.update_idletasks(), adjust_columns()))
        self.root.after(300, adjust_columns)  # 再次确保

        self.chapter_tree.bind("<Double-1>", self.on_chapter_double_click)

        # ========== 右侧详细分析 ==========
        right_frame = tk.LabelFrame(main_pane, text="🔍 详细分析", font=FONTS["content_bold"],
                                    bg=COLORS["bg_card"], fg=COLORS["primary"])
        right_frame.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)
        right_frame.grid_rowconfigure(0, weight=1)
        right_frame.grid_columnconfigure(0, weight=1)

        right_notebook = ttk.Notebook(right_frame)
        right_notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.right_notebook = right_notebook

        # 考点错误率页
        kp_frame = tk.Frame(right_notebook, bg=COLORS["bg_main"])
        right_notebook.add(kp_frame, text="📈 考点错误率")
        self.kp_tree = ttk.Treeview(kp_frame, columns=("point", "error_rate", "total"), show="headings", height=12)
        self.kp_tree.heading("point", text="考点")
        self.kp_tree.heading("error_rate", text="错误率")
        self.kp_tree.heading("total", text="总题数")
        self.kp_tree.column("point", width=220)
        self.kp_tree.column("error_rate", width=100)
        self.kp_tree.column("total", width=80)
        self.kp_tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 题型正确率页
        type_frame = tk.Frame(right_notebook, bg=COLORS["bg_main"])
        right_notebook.add(type_frame, text="📊 题型正确率")
        self.type_tree = ttk.Treeview(type_frame, columns=("type", "rate", "correct", "total"), show="headings",
                                      height=12)
        self.type_tree.heading("type", text="题型")
        self.type_tree.heading("rate", text="正确率")
        self.type_tree.heading("correct", text="正确数")
        self.type_tree.heading("total", text="总题数")
        self.type_tree.column("type", width=100)
        self.type_tree.column("rate", width=100)
        self.type_tree.column("correct", width=80)
        self.type_tree.column("total", width=80)
        self.type_tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 典型错题页
        wrong_frame = tk.Frame(right_notebook, bg=COLORS["bg_main"])
        right_notebook.add(wrong_frame, text="❌ 典型错题")
        self.wrong_text = scrolledtext.ScrolledText(wrong_frame, font=FONTS["content"], wrap=tk.WORD, height=12)
        self.wrong_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.bind_right_click(self.wrong_text)

        # 教学与学习建议页
        advice_frame = tk.Frame(right_notebook, bg=COLORS["bg_main"])
        right_notebook.add(advice_frame, text="💡 教学与学习建议")
        self.advice_text = scrolledtext.ScrolledText(advice_frame, font=FONTS["content"], wrap=tk.WORD, height=12)
        self.advice_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.bind_right_click(self.advice_text)

        # 底部按钮
        bottom_frame = tk.Frame(parent, bg=COLORS["bg_main"])
        bottom_frame.pack(fill=tk.X, pady=10)
        gen_btn = tk.Button(bottom_frame, text="📝 生成针对性练习", font=FONTS["btn"],
                            bg=COLORS["success"], fg="white", bd=0, relief=tk.FLAT, cursor="hand2",
                            command=self.generate_targeted_practice)
        gen_btn.pack(side=tk.LEFT, padx=10)

        # 绑定科目和年级的切换事件，自动加载章节数据（不生成建议）
        subject_combo.bind("<<ComboboxSelected>>", lambda e: self.load_class_data())
        grade_combo.bind("<<ComboboxSelected>>", lambda e: self.load_class_data())

        # 初始加载数据（不弹窗，不生成建议）
        self.load_class_data()

    def load_class_data(self):
        """加载章节正确率等数据（不生成教学建议）"""
        # 清空并显示加载中
        for item in self.chapter_tree.get_children():
            self.chapter_tree.delete(item)
        self.chapter_tree.insert("", tk.END, values=("加载中...", "", "", ""))

        self.kp_tree.delete(*self.kp_tree.get_children())
        self.type_tree.delete(*self.type_tree.get_children())
        self.wrong_text.config(state=tk.NORMAL)
        self.wrong_text.delete(1.0, tk.END)
        self.wrong_text.insert(tk.END, "⏳ 正在加载错题数据...")
        self.wrong_text.config(state=tk.DISABLED)
        # 建议区域显示提示，不自动生成
        self.advice_text.config(state=tk.NORMAL)
        self.advice_text.delete(1.0, tk.END)
        self.advice_text.insert(tk.END, "💡 教学与学习建议请点击【💡 生成教育建议】按钮更新。")
        self.advice_text.config(state=tk.DISABLED)

        class_name = self.ai.class_name
        subject = self.class_analysis_subject_var.get()
        grade = self.class_analysis_grade_var.get()

        # 直接在主线程中查询（数据量小，不卡顿）
        conn, cursor = self.db._connect()
        cursor.execute("""
            SELECT chapter, COUNT(*) as total, SUM(is_correct) as correct
            FROM exam_details
            WHERE class_name=? AND subject=? AND grade=?
            GROUP BY chapter
            ORDER BY chapter
        """, (class_name, subject, grade))
        rows = cursor.fetchall()
        self.db._close(conn)

        self._update_chapter_tree(rows)

    def _build_mslq_tab_embedded(self, parent):
        """内嵌的学习动机测评界面（左侧问卷，右侧历史记录可折叠）"""
        # 顶部统计和操作区
        top_frame = tk.Frame(parent, bg=COLORS["bg_main"])
        top_frame.pack(fill=tk.X, padx=20, pady=(8, 5))

        self.mslq_stats_label = tk.Label(top_frame, text="", font=FONTS["content_bold"], bg=COLORS["bg_main"],
                                         fg=COLORS["primary"])
        self.mslq_stats_label.pack(side=tk.LEFT, anchor="w")

        btn_frame = tk.Frame(top_frame, bg=COLORS["bg_main"])
        btn_frame.pack(side=tk.RIGHT)
        export_btn = tk.Button(btn_frame, text="📎 导出测评数据", font=FONTS["btn"], bg=COLORS["primary"], fg="white",
                               bd=0, relief=tk.FLAT, cursor="hand2", width=14, command=self.export_mslq_to_excel)
        export_btn.pack(side=tk.LEFT, padx=5)
        clear_mslq_btn = tk.Button(btn_frame, text="🗑️ 清空测评记录", font=FONTS["btn"], bg=COLORS["danger"],
                                   fg="white", bd=0, relief=tk.FLAT, cursor="hand2", width=14,
                                   command=self.clear_mslq_records)
        clear_mslq_btn.pack(side=tk.LEFT, padx=5)

        # 主内容区域：左右布局，但右侧可折叠
        main_frame = tk.Frame(parent, bg=COLORS["bg_main"])
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        main_frame.grid_columnconfigure(0, weight=1)  # 左侧问卷区域，会动态调整
        main_frame.grid_columnconfigure(1, weight=0)  # 右侧侧边栏，初始宽度0
        main_frame.grid_rowconfigure(0, weight=1)

        # ========= 左侧：填写测评问卷=========
        form_frame = tk.LabelFrame(main_frame, text="📝 填写测评问卷", font=FONTS["content_bold"],
                                   bg=COLORS["bg_card"], fg=COLORS["primary"])
        form_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 0), pady=0)
        form_frame.grid_rowconfigure(0, weight=1)
        form_frame.grid_columnconfigure(0, weight=1)

        # 创建Canvas和Scrollbar（问卷滚动区域）
        canvas = tk.Canvas(form_frame, bg=COLORS["bg_card"], highlightthickness=0)
        scrollbar = tk.Scrollbar(form_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=COLORS["bg_card"])

        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        def _configure_canvas(event):
            canvas.itemconfig(1, width=event.width)

        canvas.bind("<Configure>", _configure_canvas)

        # 滚轮支持
        def _on_enter(event):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)
            canvas.bind_all("<Button-4>", _on_mousewheel)
            canvas.bind_all("<Button-5>", _on_mousewheel)

        def _on_leave(event):
            canvas.unbind_all("<MouseWheel>")
            canvas.unbind_all("<Button-4>")
            canvas.unbind_all("<Button-5>")

        def _on_mousewheel(event):
            if event.num == 4 or (hasattr(event, 'delta') and event.delta > 0):
                canvas.yview_scroll(-1, "units")
            elif event.num == 5 or (hasattr(event, 'delta') and event.delta < 0):
                canvas.yview_scroll(1, "units")

        canvas.bind("<Enter>", _on_enter)
        canvas.bind("<Leave>", _on_leave)

        inner = scrollable_frame
        inner.grid_columnconfigure(0, weight=3)  # 题目列
        inner.grid_columnconfigure(1, weight=1)  # 选项列

        # 姓名、班级
        info_frame = tk.Frame(inner, bg=COLORS["bg_card"])
        info_frame.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 10))
        tk.Label(info_frame, text="👤 姓名：", font=FONTS["content_bold"], bg=COLORS["bg_card"]).pack(side=tk.LEFT)
        self.mslq_name_entry = tk.Entry(info_frame, font=FONTS["content"], width=15)
        self.mslq_name_entry.pack(side=tk.LEFT, padx=5)
        tk.Label(info_frame, text="🏫 班级：", font=FONTS["content_bold"], bg=COLORS["bg_card"]).pack(side=tk.LEFT,
                                                                                                    padx=10)
        self.mslq_class_entry = tk.Entry(info_frame, font=FONTS["content"], width=15)
        self.mslq_class_entry.insert(0, self.ai.class_name)
        self.mslq_class_entry.config(state='readonly')
        self.mslq_class_entry.pack(side=tk.LEFT, padx=5)

        # 列标题
        tk.Label(inner, text="题目", font=FONTS["content_bold"], bg=COLORS["bg_card"], fg=COLORS["primary"],
                 anchor="w").grid(row=1, column=0, sticky="w", pady=(0, 5), padx=5)

        title_frame = tk.Frame(inner, bg=COLORS["bg_card"])
        title_frame.grid(row=1, column=1, sticky="w", pady=(0, 5), padx=5)
        tk.Label(title_frame, text="评分", font=FONTS["content_bold"], bg=COLORS["bg_card"], fg=COLORS["primary"]).pack(
            side=tk.LEFT)
        tk.Label(title_frame, text=" (1低→5高)", font=FONTS["small"], bg=COLORS["bg_card"],
                 fg=COLORS["text_secondary"]).pack(side=tk.LEFT)

        # 18题
        questions = [
            "1. 在学习本课程时，我更喜欢学习那些能够引起我学习兴趣的内容，即使它比较难学。",
            "2. 本课程最让我满意的地方是我能够尽可能透彻地理解它的内容。",
            "3. 在学习本课程时，如果有选择的话，我会选择那些能让我学到更多东西的作业，即使这些作业不能保证我得到高分。",
            "4. 在本课程上拿到好分数是目前最令我满意的事。",
            "5. 现在对我来说最重要的事情就是成绩，所以我对本课程的主要期望是要拿个好成绩。",
            "6. 如果可能的话，我希望在本课程中能拿到比班里大部分同学更高的分数。",
            "7. 对我来说，学好本课程老师讲授的内容非常重要。",
            "8. 我对本课程所学习的内容很感兴趣。",
            "9. 理解本课程的学习内容对我来说很重要。",
            "10. 如果我学习得法，我就能够学会本课程的内容。",
            "11. 没有学好本课程是我自身的问题。",
            "12. 如果我足够努力，我就能够理解本课程的学习材料。",
            "13. 我相信我在本课程中能够获得一个漂亮的分数。",
            "14. 我确信自己能够理解本课程老师所呈现的最复杂的材料。",
            "15. 综合考虑本课程的难度、老师和我自己的能力，我想我可以在本课程中表现得很好。",
            "16. 考试时，我一直想着自己不会回答的题目。",
            "17. 考试时，我会去想考不好的后果。",
            "18. 考试时，我会感到紧张不安。"
        ]
        self.mslq_vars = []

        for i, q in enumerate(questions):
            lbl = tk.Label(inner, text=q, font=FONTS["content"], bg=COLORS["bg_card"], justify="left", anchor="w")
            lbl.grid(row=i + 2, column=0, sticky="w", pady=4, padx=5)

            var = tk.IntVar(value=3)
            self.mslq_vars.append(var)
            opt_frame = tk.Frame(inner, bg=COLORS["bg_card"])
            opt_frame.grid(row=i + 2, column=1, sticky="w", pady=4, padx=5)
            for val in [1, 2, 3, 4, 5]:
                rb = tk.Radiobutton(opt_frame, text=str(val), variable=var, value=val,
                                    bg=COLORS["bg_card"], activebackground=COLORS["bg_card"])
                rb.pack(side=tk.LEFT, padx=5)

        submit_btn = tk.Button(inner, text="✅ 提交测评", font=FONTS["btn"], bg=COLORS["success"], fg="white",
                               bd=0, relief=tk.FLAT, cursor="hand2", width=12, command=self.submit_mslq_embedded)
        submit_btn.grid(row=len(questions) + 2, column=0, columnspan=2, pady=15)

        # ========= 右侧：可折叠历史记录侧边栏 =========
        # 用于存储右侧frame的引用
        self.right_panel = None
        self.history_frame = None
        self.sidebar_visible = False
        self.sidebar_width = 400  # 展开时的宽度

        # 创建右侧容器（用于放置折叠/展开按钮和历史记录）
        right_container = tk.Frame(main_frame, bg=COLORS["bg_main"])
        right_container.grid(row=0, column=1, sticky="nsew", padx=(5, 0), pady=0)
        # 初始宽度为0（隐藏）
        right_container.columnconfigure(0, weight=1)
        right_container.rowconfigure(0, weight=1)

        # 折叠/展开按钮（竖条状，放在右侧容器的最左边）
        toggle_frame = tk.Frame(right_container, bg="#E5E7EB", width=30)
        toggle_frame.pack(side=tk.LEFT, fill=tk.Y)
        toggle_frame.pack_propagate(False)  # 固定宽度
        # 添加一个按钮
        toggle_btn = tk.Label(toggle_frame, text="◀", font=("Arial", 12, "bold"),
                              bg="#9CA3AF", fg="white", cursor="hand2")
        toggle_btn.pack(expand=True)
        toggle_btn.bind("<Button-1>", self._toggle_mslq_sidebar)

        # 历史记录面板（初始隐藏）
        self.history_panel = tk.Frame(right_container, bg=COLORS["bg_card"], width=self.sidebar_width)
        # 先不pack，等展开时才显示

        # 记录toggle_frame和history_panel的引用
        self.toggle_frame = toggle_frame
        self.right_container = right_container

        # 初始化历史记录树（先创建但不显示，等展开时再显示）
        self._create_mslq_history_tree()

        self.refresh_mslq_stats()
        self.refresh_mslq_tree()

    def _create_mslq_history_tree(self):
        """创建历史记录的树形视图（供折叠侧边栏使用）"""
        if hasattr(self, 'history_panel') and self.history_panel:
            # 清空原有内容
            for widget in self.history_panel.winfo_children():
                widget.destroy()
        else:
            self.history_panel = tk.Frame(self.right_container, bg=COLORS["bg_card"], width=self.sidebar_width)
        # 内部布局
        self.history_panel.pack_propagate(False)  # 固定宽度
        self.history_panel.config(width=self.sidebar_width)

        tree_frame = tk.Frame(self.history_panel, bg=COLORS["bg_card"])
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 树形视图
        self.mslq_tree = ttk.Treeview(tree_frame, columns=("name", "class", "type", "score", "time"),
                                      show="headings", height=15)
        self.mslq_tree.heading("name", text="姓名")
        self.mslq_tree.heading("class", text="班级")
        self.mslq_tree.heading("type", text="类型")
        self.mslq_tree.heading("score", text="总分")
        self.mslq_tree.heading("time", text="时间")

        # 设置列宽
        self.mslq_tree.column("name", width=80, minwidth=70)
        self.mslq_tree.column("class", width=80, minwidth=70)
        self.mslq_tree.column("type", width=60, minwidth=50)
        self.mslq_tree.column("score", width=60, minwidth=50)
        self.mslq_tree.column("time", width=150, minwidth=120)

        # 添加滚动条
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.mslq_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.mslq_tree.xview)
        self.mslq_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.mslq_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

    def _toggle_mslq_sidebar(self, event=None):
        """切换右侧历史记录侧边栏的展开/折叠状态"""
        if self.sidebar_visible:
            # 折叠：隐藏历史面板，改变按钮箭头方向
            self.history_panel.pack_forget()
            self.toggle_frame.children["!label"].config(text="▶")  # 箭头向右
            self.sidebar_visible = False
            # 让左侧问卷区域扩展（grid权重恢复）
            self.right_container.grid_configure(padx=(5, 0))
            # 调整左侧frame的padding
            for child in self.right_container.master.grid_slaves(row=0, column=0):
                if isinstance(child, tk.LabelFrame):
                    child.grid_configure(padx=(0, 0))
        else:
            # 展开：显示历史面板，改变按钮箭头方向
            self.history_panel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            self.toggle_frame.children["!label"].config(text="◀")  # 箭头向左
            self.sidebar_visible = True
            self.right_container.grid_configure(padx=(5, 0))
            # 刷新历史数据
            self.refresh_mslq_tree()

    def submit_mslq_embedded(self):
        name = self.mslq_name_entry.get().strip()
        if not name:
            messagebox.showwarning("提示", "请填写姓名！")
            return
        class_name = self.mslq_class_entry.get()
        total = sum(v.get() for v in self.mslq_vars)
        test_type = tk.messagebox.askyesno("测试类型", "这是前测吗？\n（是=前测，否=后测）")
        test_type_str = "pre" if test_type else "post"
        conn, cursor = self.db._connect()
        cursor.execute(
            "INSERT INTO mslq_records (student_name, class_name, test_type, total_score, details) VALUES (?, ?, ?, ?, ?)",
            (name, class_name, test_type_str, total, str([v.get() for v in self.mslq_vars])))
        self.db._close(conn)
        messagebox.showinfo("成功", "测评已保存！")
        self.refresh_mslq_stats()
        self.refresh_mslq_tree()
        # 清空表单
        self.mslq_name_entry.delete(0, tk.END)
        for var in self.mslq_vars:
            var.set(3)

    def refresh_mslq_stats(self):
        conn, cursor = self.db._connect()
        cursor.execute("SELECT COUNT(DISTINCT student_name) as count FROM mslq_records")
        row = cursor.fetchone()
        total_people = row['count'] if row else 0
        cursor.execute("SELECT COUNT(*) as total FROM mslq_records")
        total_records = cursor.fetchone()['total']
        self.db._close(conn)
        self.mslq_stats_label.config(text=f"📊 已参与人数：{total_people} 人  |  总测评次数：{total_records} 次")

    def refresh_mslq_tree(self):
        for item in self.mslq_tree.get_children():
            self.mslq_tree.delete(item)
        conn, cursor = self.db._connect()
        cursor.execute(
            "SELECT student_name, class_name, test_type, total_score, create_time FROM mslq_records ORDER BY create_time DESC")
        rows = cursor.fetchall()
        self.db._close(conn)
        for r in rows:
            type_str = "前测" if r['test_type'] == "pre" else "后测"
            self.mslq_tree.insert("", tk.END, values=(
            r['student_name'], r['class_name'], type_str, r['total_score'], r['create_time']))

    def clear_mslq_records(self):
        if messagebox.askyesno("确认清空", "此操作将永久删除所有学习动机测评记录，是否继续？"):
            conn, cursor = self.db._connect()
            cursor.execute("DELETE FROM mslq_records")
            self.db._close(conn)
            self.refresh_mslq_stats()
            self.refresh_mslq_tree()
            messagebox.showinfo("完成", "测评记录已清空")

    def export_mslq_to_excel(self):
        """导出MSLQ测评记录，文件名包含班级、测试类型、时间"""
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "MSLQ测评记录"
            ws.append(["姓名", "班级", "测试类型", "总分", "详细得分", "时间"])

            conn, cursor = self.db._connect()
            cursor.execute("SELECT * FROM mslq_records ORDER BY create_time DESC")
            rows = cursor.fetchall()
            self.db._close(conn)

            # 确定导出文件的班级和测试类型（取第一条记录，若无记录则用当前班级）
            class_name = self.ai.class_name
            test_type_display = ""
            if rows:
                class_name = rows[0]['class_name']
                test_type_display = "前测" if rows[0]['test_type'] == "pre" else "后测"

            for r in rows:
                ws.append([r['student_name'], r['class_name'], "前测" if r['test_type'] == "pre" else "后测",
                           r['total_score'], r['details'], r['create_time']])

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"MSLQ测评记录_{class_name}_{test_type_display}_{timestamp}.xlsx"
            wb.save(filename)
            messagebox.showinfo("导出成功", f"已保存到程序所在目录：{filename}")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    def refresh_advice_only(self):
        """仅刷新教学建议（带确认框），并跳转到建议标签页"""
        # 1. 检查是否选中章节
        selected = self.chapter_tree.selection()
        if not selected:
            messagebox.showinfo("提示", "请先双击选择一个章节。")
            return
        item = selected[0]
        chapter = self.chapter_data.get(item, "")
        if not chapter:
            messagebox.showinfo("提示", "请先双击选择一个章节。")
            return

        # 2. 检查该章节的题目数是否≥5
        class_name = self.ai.class_name
        subject = self.class_analysis_subject_var.get()
        grade = self.class_analysis_grade_var.get()
        conn, cursor = self.db._connect()
        cursor.execute("""
            SELECT COUNT(*) as total FROM exam_details
            WHERE class_name=? AND subject=? AND grade=? AND chapter=?
        """, (class_name, subject, grade, chapter))
        total_questions = cursor.fetchone()['total']
        self.db._close(conn)

        if total_questions < 5:
            messagebox.showinfo("提示", f"当前章节“{chapter}”答题数据不足5道，无法生成教学建议。\n请至少完成5道题目后再试。")
            return

        # 3. 确认对话框
        if not messagebox.askyesno("确认生成教育建议", "即将生成教育与教学建议，是否继续？"):
            return

        # 4. 跳转到教学与学习建议标签页（索引3）
        if hasattr(self, 'right_notebook'):
            self.right_notebook.select(3)

        # 5. 显示“正在生成...”并异步生成建议
        self.advice_text.config(state=tk.NORMAL)
        self.advice_text.delete(1.0, tk.END)
        self.advice_text.insert(tk.END, "⏳ 正在生成教学建议...")
        self.advice_text.config(state=tk.DISABLED)

        def generate_advice():
            # 重新查询该章节的考点数据（确保最新）
            conn, cursor = self.db._connect()
            cursor.execute("""
                SELECT knowledge_point_id, COUNT(*) as total, SUM(is_correct) as correct
                FROM exam_details
                WHERE class_name=? AND subject=? AND grade=? AND chapter=?
                GROUP BY knowledge_point_id
                ORDER BY (total - correct) DESC
            """, (class_name, subject, grade, chapter))
            kp_rows = cursor.fetchall()
            self.db._close(conn)

            total_q = sum(r['total'] for r in kp_rows) if kp_rows else 0
            advice_text = "📭 暂无足够数据（至少5道题）生成教学建议。"
            if kp_rows and total_q >= 5:
                weakest = kp_rows[0]
                error_rate = (weakest['total'] - weakest['correct']) / weakest['total'] * 100 if weakest['total'] > 0 else 0
                kp_id = weakest['knowledge_point_id']
                kp_name = KNOWLEDGE_POINTS.get(kp_id, {}).get('name', '该考点')
                teacher_prompt = f"我班学生在【{kp_name}】考点错误率高达{error_rate:.1f}%，请给出一条简短的、可操作的教学改进建议（不超过100字）。"
                success_t, advice_teacher = self.ai.ai_request("教学建议", "不限", "建议", teacher_prompt, auto_save=False, temperature=0.5)
                teacher_advice = advice_teacher if success_t else "建议通过典型例题讲解该考点，并布置针对性练习。"
                student_prompt = f"学生在【{kp_name}】考点错误率较高，请给出一条简短的学习建议（不超过100字），帮助学生自主复习。"
                success_s, advice_student = self.ai.ai_request("学习建议", "不限", "建议", student_prompt, auto_save=False, temperature=0.5)
                student_advice = advice_student if success_s else "建议多练习该类题目，并回顾教材相关知识点。"
                advice_text = f"📌 教学建议（教师）：{teacher_advice}\n\n📌 学习建议（学生）：{student_advice}"

            self.task_queue.put((self._update_advice_text, (advice_text,), {}))

        threading.Thread(target=generate_advice, daemon=True).start()

    def _update_advice_text(self, text):
        self.advice_text.config(state=tk.NORMAL)
        self.advice_text.delete(1.0, tk.END)
        self.advice_text.insert(tk.END, text)
        self.advice_text.config(fg="black", font=FONTS["content"])  # 强制设为黑色
        self.advice_text.config(state=tk.DISABLED)

    def _auto_select_first_and_refresh(self):
        items = self.chapter_tree.get_children()
        if items:
            self.chapter_tree.selection_set(items[0])
            self._refresh_right_panel(force_advice=True)

    def _update_chapter_tree(self, rows):
        """更新章节正确率统计树"""
        for item in self.chapter_tree.get_children():
            self.chapter_tree.delete(item)
        self.chapter_data = {}
        total_questions = sum(r['total'] for r in rows) if rows else 0

        if not rows or total_questions < 5:
            self.chapter_tree.insert("", tk.END, values=("暂无足够数据", 0, 0, "0%"))
            self.kp_tree.delete(*self.kp_tree.get_children())
            self.kp_tree.insert("", tk.END, values=("暂无足够数据（至少5道题）", "0%", 0))
            self.type_tree.delete(*self.type_tree.get_children())
            self.type_tree.insert("", tk.END, values=("暂无足够数据", "0%", 0, 0))
            self.wrong_text.config(state=tk.NORMAL)
            self.wrong_text.delete(1.0, tk.END)
            self.wrong_text.insert(tk.END, "📭 暂无足够数据。请至少完成5道题目后再查看典型错题。")
            self.wrong_text.config(fg="gray", font=FONTS["content"])
            self.wrong_text.config(state=tk.DISABLED)
            self.advice_text.config(state=tk.NORMAL)
            self.advice_text.delete(1.0, tk.END)
            self.advice_text.insert(tk.END, "📭 暂无足够数据。请至少完成5道题目后再生成教学建议。")
            self.advice_text.config(fg="gray", font=FONTS["content"])
            self.advice_text.config(state=tk.DISABLED)
            return

        for r in rows:
            rate = (r['correct'] / r['total']) * 100 if r['total'] > 0 else 0
            tag = 'weak' if rate < 60 else 'normal'
            item_id = self.chapter_tree.insert("", tk.END,
                                               values=(r['chapter'], r['total'], r['correct'], f"{rate:.1f}%"),
                                               tags=(tag,))
            self.chapter_data[item_id] = r['chapter']
        self.chapter_tree.tag_configure('weak', background='#FFCCCC')

        if rows:
            first_item = self.chapter_tree.get_children()[0]
            self.chapter_tree.selection_set(first_item)
            self.on_chapter_double_click(None)

    def on_chapter_double_click(self, event):
        self._refresh_right_panel(force_advice=False)

    def generate_targeted_practice(self):
        if not hasattr(self, 'current_weak_kp') or not self.current_weak_kp:
            messagebox.showwarning("提示", "请先双击选择一个章节，并确保有薄弱考点数据。")
            return
        subject = self.class_analysis_subject_var.get()
        grade = self.class_analysis_grade_var.get()
        kp_id = self.current_weak_kp
        kp_info = KNOWLEDGE_POINTS.get(kp_id, {})
        kp_name = kp_info.get('name', kp_id)

        self.switch_frame("ai_exam")
        self.exam_subject_var.set(subject)
        self.exam_grade_var.set(grade)
        if self.current_weak_chapter:
            self.exam_chapter_var.set(self.current_weak_chapter)
            self.update_exam_lesson_options()
        self.exam_kp_var.set(kp_name)
        self.root.after(300, self.generate_ai_exam)

    def clear_class_data(self):
        if messagebox.askyesno("确认清空",
                               f"此操作将永久删除当前班级“{self.ai.class_name}”的所有学情明细数据，是否继续？"):
            conn, cursor = self.db._connect()
            cursor.execute("DELETE FROM exam_details WHERE class_name=?", (self.ai.class_name,))
            self.db._close(conn)
            self.refresh_class_analysis()
            messagebox.showinfo("完成", "班级学情数据已清空")

    def _refresh_right_panel(self, force_advice=False):
        selected = self.chapter_tree.selection()
        if not selected:
            return
        item = selected[0]
        chapter = self.chapter_data.get(item, "")
        if not chapter:
            return

        class_name = self.ai.class_name
        subject = self.class_analysis_subject_var.get()
        grade = self.class_analysis_grade_var.get()

        # 显示加载提示
        self.kp_tree.delete(*self.kp_tree.get_children())
        self.kp_tree.insert("", tk.END, values=("加载中...", "", ""))
        self.type_tree.delete(*self.type_tree.get_children())
        self.type_tree.insert("", tk.END, values=("加载中...", "", "", ""))
        self.wrong_text.config(state=tk.NORMAL)
        self.wrong_text.delete(1.0, tk.END)
        self.wrong_text.insert(tk.END, "⏳ 正在加载错题...")
        self.wrong_text.config(state=tk.DISABLED)

        # 建议区域提示
        self.advice_text.config(state=tk.NORMAL)
        self.advice_text.delete(1.0, tk.END)
        if force_advice:
            self.advice_text.insert(tk.END, "⏳ 正在生成教学建议...")
        else:
            self.advice_text.insert(tk.END, "💡 教学与学习建议请点击【💡 生成教育建议】按钮更新。")
        self.advice_text.config(state=tk.DISABLED)

        def fetch_data():
            conn, cursor = self.db._connect()
            # 考点错误率
            cursor.execute("""
                SELECT knowledge_point_id, COUNT(*) as total, SUM(is_correct) as correct
                FROM exam_details
                WHERE class_name=? AND subject=? AND grade=? AND chapter=?
                GROUP BY knowledge_point_id
                ORDER BY (total - correct) DESC
            """, (class_name, subject, grade, chapter))
            kp_rows = cursor.fetchall()

            # 题型正确率
            cursor.execute("""
                SELECT question_type, COUNT(*) as total, SUM(is_correct) as correct
                FROM exam_details
                WHERE class_name=? AND subject=? AND grade=? AND chapter=?
                GROUP BY question_type
            """, (class_name, subject, grade, chapter))
            type_rows = cursor.fetchall()

            # 典型错题
            cursor.execute("""
                SELECT question, correct_answer, COUNT(*) as wrong_count
                FROM exam_details
                WHERE class_name=? AND subject=? AND grade=? AND chapter=? AND is_correct=0
                GROUP BY question
                ORDER BY wrong_count DESC
                LIMIT 3
            """, (class_name, subject, grade, chapter))
            wrong_rows = cursor.fetchall()
            self.db._close(conn)

            advice_text = None
            if force_advice:
                total_questions = sum(r['total'] for r in kp_rows) if kp_rows else 0
                if kp_rows and total_questions >= 5:
                    weakest = kp_rows[0]
                    error_rate = (weakest['total'] - weakest['correct']) / weakest['total'] * 100 if weakest['total'] > 0 else 0
                    kp_id = weakest['knowledge_point_id']
                    kp_name = KNOWLEDGE_POINTS.get(kp_id, {}).get('name', '该考点')
                    teacher_prompt = f"我班学生在【{kp_name}】考点错误率高达{error_rate:.1f}%，请给出一条简短的、可操作的教学改进建议（不超过100字）。"
                    success_t, advice_teacher = self.ai.ai_request("教学建议", "不限", "建议", teacher_prompt, auto_save=False, temperature=0.5)
                    teacher_advice = advice_teacher if success_t else "建议通过典型例题讲解该考点，并布置针对性练习。"
                    student_prompt = f"学生在【{kp_name}】考点错误率较高，请给出一条简短的学习建议（不超过100字），帮助学生自主复习。"
                    success_s, advice_student = self.ai.ai_request("学习建议", "不限", "建议", student_prompt, auto_save=False, temperature=0.5)
                    student_advice = advice_student if success_s else "建议多练习该类题目，并回顾教材相关知识点。"
                    advice_text = f"📌 教学建议（教师）：{teacher_advice}\n\n📌 学习建议（学生）：{student_advice}"
                else:
                    advice_text = "📭 暂无足够数据（至少5道题）生成教学建议。"

            self._pending_right_data = (kp_rows, type_rows, wrong_rows, advice_text, chapter)
            self.root.event_generate('<<UpdateRightPanel>>')

        threading.Thread(target=fetch_data, daemon=True).start()

    def _update_right_panel_ui(self, kp_rows, type_rows, wrong_rows, advice_text, chapter):
        # 考点错误率树
        self.kp_tree.delete(*self.kp_tree.get_children())
        if kp_rows:
            for kp in kp_rows:
                kp_id = kp['knowledge_point_id']
                if kp_id and kp_id in KNOWLEDGE_POINTS:
                    kp_name = KNOWLEDGE_POINTS[kp_id]['name']
                else:
                    kp_name = kp_id if kp_id else "未归类考点"
                error_rate = (kp['total'] - kp['correct']) / kp['total'] * 100 if kp['total'] > 0 else 0
                self.kp_tree.insert("", tk.END, values=(kp_name, f"{error_rate:.1f}%", kp['total']))
        else:
            self.kp_tree.insert("", tk.END, values=("暂无考点数据", "0%", 0))

        # 题型正确率树
        self.type_tree.delete(*self.type_tree.get_children())
        if type_rows:
            for t in type_rows:
                rate = t['correct'] / t['total'] * 100 if t['total'] > 0 else 0
                self.type_tree.insert("", tk.END, values=(t['question_type'], f"{rate:.1f}%", t['correct'], t['total']))
        else:
            self.type_tree.insert("", tk.END, values=("暂无题型数据", "0%", 0, 0))

        # 典型错题
        self.wrong_text.config(state=tk.NORMAL)
        self.wrong_text.delete(1.0, tk.END)
        if wrong_rows:
            for idx, w in enumerate(wrong_rows, 1):
                self.wrong_text.insert(tk.END,
                                       f"【错题{idx}】\n题目：{w['question'][:200]}...\n标准答案：{w['correct_answer']}\n")
                self.wrong_text.insert(tk.END, f"错误次数：{w['wrong_count']}\n\n")
        else:
            self.wrong_text.insert(tk.END, "📭 暂无典型错题。")
        self.wrong_text.config(fg="black")
        self.wrong_text.config(state=tk.DISABLED)

        # 建议
        self.advice_text.config(state=tk.NORMAL)
        self.advice_text.delete(1.0, tk.END)
        if advice_text:
            self.advice_text.insert(tk.END, advice_text)
        else:
            self.advice_text.insert(tk.END, "💡 教学与学习建议请点击【💡 生成教育建议】按钮更新。")
        self.advice_text.config(state=tk.DISABLED)

        self.current_weak_chapter = chapter
        if kp_rows:
            self.current_weak_kp = kp_rows[0]['knowledge_point_id']
        else:
            self.current_weak_kp = None

    # ---------------------- 记录模块（AI问答记录 + 答题记录） ----------------------
    def _build_chat_records_tab(self, parent, skip_refresh=False):
        toolbar = tk.Frame(parent, bg=COLORS["bg_main"])
        toolbar.pack(fill=tk.X, padx=15, pady=8)

        tk.Label(toolbar, text="学科筛选：", font=FONTS["content_bold"], bg=COLORS["bg_main"]).pack(side=tk.LEFT)
        self.record_subject_combo = ttk.Combobox(toolbar, values=ALL_SUBJECTS, width=12, state="readonly")
        self.record_subject_combo.pack(side=tk.LEFT, padx=6)
        self.record_subject_combo.current(0)

        refresh_btn = tk.Button(toolbar, text="🔄 刷新", font=FONTS["btn"],
                                bg=COLORS["primary"], fg="white", bd=0, relief=tk.FLAT, cursor="hand2",
                                command=self.refresh_records)
        refresh_btn.pack(side=tk.LEFT, padx=12)

        clear_btn = tk.Button(toolbar, text="🗑️ 清空问答记录", font=FONTS["btn"],
                              bg=COLORS["danger"], fg="white", bd=0, relief=tk.FLAT, cursor="hand2",
                              command=self.clear_chat_records)
        clear_btn.pack(side=tk.RIGHT, padx=6)

        self.record_content = scrolledtext.ScrolledText(parent, font=FONTS["content"], wrap=tk.WORD)
        self.record_content.pack(fill=tk.BOTH, expand=True, padx=15, pady=8)
        self.bind_right_click(self.record_content)
        self.record_subject_combo.bind("<<ComboboxSelected>>", lambda e: self.refresh_records())
        if not skip_refresh:
            self.refresh_records()

    def _build_exam_records_tab(self, parent, skip_refresh=False):
        toolbar = tk.Frame(parent, bg=COLORS["bg_main"])
        toolbar.pack(fill=tk.X, padx=15, pady=8)

        tk.Label(toolbar, text="答题历史", font=FONTS["content_bold"], bg=COLORS["bg_main"]).pack(side=tk.LEFT)

        refresh_btn = tk.Button(toolbar, text="🔄 刷新", font=FONTS["btn"],
                                bg=COLORS["primary"], fg="white", bd=0, relief=tk.FLAT, cursor="hand2",
                                command=self.refresh_records)
        refresh_btn.pack(side=tk.LEFT, padx=12)

        clear_btn = tk.Button(toolbar, text="🗑️ 清空答题记录", font=FONTS["btn"],
                              bg=COLORS["danger"], fg="white", bd=0, relief=tk.FLAT, cursor="hand2",
                              command=self.clear_exam_records)
        clear_btn.pack(side=tk.RIGHT, padx=6)

        self.exam_record_content = scrolledtext.ScrolledText(parent, font=FONTS["content"], wrap=tk.WORD)
        self.exam_record_content.pack(fill=tk.BOTH, expand=True, padx=15, pady=8)
        self.bind_right_click(self.exam_record_content)
        if not skip_refresh:
            self.refresh_records()

    def clear_chat_records(self):
        if messagebox.askyesno("确认清空", "此操作将永久删除所有AI问答记录，是否继续？"):
            conn, cursor = self.db._connect()
            cursor.execute("DELETE FROM ai_chat_records")
            self.db._close(conn)
            self.refresh_records()
            messagebox.showinfo("完成", "问答记录已清空")

    def clear_exam_records(self):
        if messagebox.askyesno("确认清空", "此操作将永久删除所有答题记录（包括班级学情明细），是否继续？"):
            conn, cursor = self.db._connect()
            cursor.execute("DELETE FROM exam_records")
            cursor.execute("DELETE FROM exam_details")
            self.db._close(conn)
            self.refresh_records()
            messagebox.showinfo("完成", "答题记录已清空")

    def refresh_records(self):
        # 问答记录
        subject = self.record_subject_combo.get()
        records = self.db.get_chat_records(subject)
        self.record_content.config(state=tk.NORMAL)
        self.record_content.delete(1.0, tk.END)
        if not records:
            self.record_content.insert(tk.END, "📭 暂无问答记录。\n\n请前往【AI全科答疑】模块进行提问。")
            self.record_content.config(fg="gray", font=FONTS["placeholder"])
        else:
            self.record_content.config(fg="black", font=FONTS["content"])
            for record in records:
                self.record_content.insert(tk.END, f"⏰ {record['ask_time']} | 📚 {record['subject']}\n", "time")
                self.record_content.insert(tk.END, f"❓ {record['question']}\n\n", "question")
                self.record_content.insert(tk.END, f"🤖 {record['answer']}\n", "answer")
                self.record_content.insert(tk.END, "-" * 80 + "\n\n")
            self.record_content.tag_config("time", foreground=COLORS["text_secondary"])
            self.record_content.tag_config("question", foreground=COLORS["primary"])
        self.record_content.config(state=tk.DISABLED)

        # 答题记录
        exam_records = self.db.get_exam_records()
        self.exam_record_content.config(state=tk.NORMAL)
        self.exam_record_content.delete(1.0, tk.END)
        if not exam_records:
            self.exam_record_content.insert(tk.END, "📭 暂无答题记录。\n\n请前往【AI随堂答题】模块完成测试。")
            self.exam_record_content.config(fg="gray", font=FONTS["placeholder"])
        else:
            self.exam_record_content.config(fg="black", font=FONTS["content"])
            for record in exam_records:
                r = dict(record)
                lesson = r.get('lesson', '')
                if lesson and lesson != "整个章节":
                    lesson = f" - {lesson}"
                self.exam_record_content.insert(tk.END, f"⏰ {r['exam_time']}\n")
                self.exam_record_content.insert(tk.END, f"📚 {r['subject']} | {r['grade']} | {r['chapter']}{lesson}\n")
                self.exam_record_content.insert(tk.END,
                                                f"📊 总题数：{r['total_num']} | 答对：{r['correct_num']} | 正确率：{r['accuracy']}\n")
                self.exam_record_content.insert(tk.END, "-" * 80 + "\n\n")
        self.exam_record_content.config(state=tk.DISABLED)

    def _update_detail_panels(self, kp_rows, type_rows, wrong_rows, advice_text, chapter):
        """更新右侧详细分析面板（考点、题型、错题、建议）"""
        # 更新考点错误率树
        self.kp_tree.delete(*self.kp_tree.get_children())
        if kp_rows:
            for kp in kp_rows:
                kp_id = kp['knowledge_point_id']
                if kp_id and kp_id in KNOWLEDGE_POINTS:
                    kp_name = KNOWLEDGE_POINTS[kp_id]['name']
                else:
                    kp_name = kp_id if kp_id else "未归类考点"
                error_rate = (kp['total'] - kp['correct']) / kp['total'] * 100 if kp['total'] > 0 else 0
                self.kp_tree.insert("", tk.END, values=(kp_name, f"{error_rate:.1f}%", kp['total']))
        else:
            self.kp_tree.insert("", tk.END, values=("暂无考点数据", "0%", 0))

        # 更新题型正确率树
        self.type_tree.delete(*self.type_tree.get_children())
        if type_rows:
            for t in type_rows:
                rate = t['correct'] / t['total'] * 100 if t['total'] > 0 else 0
                self.type_tree.insert("", tk.END, values=(t['question_type'], f"{rate:.1f}%", t['correct'], t['total']))
        else:
            self.type_tree.insert("", tk.END, values=("暂无题型数据", "0%", 0, 0))

        # 更新典型错题
        self.wrong_text.config(state=tk.NORMAL)
        self.wrong_text.delete(1.0, tk.END)
        if wrong_rows:
            for idx, w in enumerate(wrong_rows, 1):
                self.wrong_text.insert(tk.END,
                                       f"【错题{idx}】\n题目：{w['question'][:200]}...\n标准答案：{w['correct_answer']}\n")
                self.wrong_text.insert(tk.END, f"错误次数：{w['wrong_count']}\n\n")
        else:
            self.wrong_text.insert(tk.END, "📭 暂无典型错题。")
        self.wrong_text.config(state=tk.DISABLED)

        # 更新建议
        self.advice_text.config(state=tk.NORMAL)
        self.advice_text.delete(1.0, tk.END)
        self.advice_text.insert(tk.END, advice_text)
        self.advice_text.config(state=tk.DISABLED)

        self.current_weak_chapter = chapter
        if kp_rows:
            self.current_weak_kp = kp_rows[0]['knowledge_point_id']
        else:
            self.current_weak_kp = None



    # ---------------------- 系统设置 ----------------------
    def build_setting_frame(self):
        """系统设置界面 - 美化版，支持触摸滚动"""
        # 外层容器
        outer = tk.Frame(self.content_container, bg=COLORS["bg_main"])

        # 标题（固定，左对齐）
        tk.Label(outer, text="⚙️ 系统设置", font=FONTS["title"],
                 bg=COLORS["bg_main"], fg=COLORS["text_main"]).pack(anchor='w', pady=(10, 5))

        # ========== 创建可滚动区域（同时支持鼠标滚轮和触摸屏滑动） ==========
        canvas = tk.Canvas(outer, bg=COLORS["bg_main"], highlightthickness=0)
        scrollbar = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)

        scrollable_frame = tk.Frame(canvas, bg=COLORS["bg_main"])
        canvas_frame_id = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

        # 当 Canvas 大小改变时，让内部 Frame 宽度自适应
        def on_canvas_resize(event):
            canvas.itemconfig(canvas_frame_id, width=event.width)

        canvas.bind("<Configure>", on_canvas_resize)

        # 当内部 Frame 大小改变时，更新滚动区域
        def on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))

        scrollable_frame.bind("<Configure>", on_frame_configure)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # ----- 1. 鼠标滚轮支持（仅在鼠标位于 Canvas 区域内时生效）-----
        def _on_mousewheel(event):
            # Windows: event.delta, Linux: event.num
            if event.num == 4 or (hasattr(event, 'delta') and event.delta > 0):
                canvas.yview_scroll(-1, "units")
            elif event.num == 5 or (hasattr(event, 'delta') and event.delta < 0):
                canvas.yview_scroll(1, "units")

        def _bind_mousewheel(event):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)
            canvas.bind_all("<Button-4>", _on_mousewheel)
            canvas.bind_all("<Button-5>", _on_mousewheel)

        def _unbind_mousewheel(event):
            canvas.unbind_all("<MouseWheel>")
            canvas.unbind_all("<Button-4>")
            canvas.unbind_all("<Button-5>")

        canvas.bind("<Enter>", _bind_mousewheel)
        canvas.bind("<Leave>", _unbind_mousewheel)

        inner = scrollable_frame

        # ========== 1. AI 功能配置 ==========
        ai_frame = tk.LabelFrame(inner, text="🤖 AI 功能配置", font=FONTS["nav"],
                                 bg=COLORS["bg_card"], fg=COLORS["primary"])
        ai_frame.pack(fill=tk.X, padx=20, pady=10)
        ai_frame.grid_columnconfigure(0, weight=0)
        ai_frame.grid_columnconfigure(1, weight=1)

        row = 0
        # 选择模型

        def get_model_names():
            conn, cursor = self.db._connect()
            cursor.execute("SELECT model_name FROM model_config WHERE is_enabled=1 ORDER BY priority, model_name")
            rows = cursor.fetchall()
            self.db._close(conn)
            return [r['model_name'] for r in rows]

        # 选择模型（从数据库读取）
        def get_model_names():
            conn, cursor = self.db._connect()
            cursor.execute("SELECT model_name FROM model_config WHERE is_enabled=1 ORDER BY priority, model_name")
            rows = cursor.fetchall()
            self.db._close(conn)
            return [r['model_name'] for r in rows]

        model_list = get_model_names()
        tk.Label(ai_frame, text="选择模型：", font=FONTS["content_bold"], bg=COLORS["bg_card"]).grid(row=row, column=0, sticky="e", padx=10, pady=8)
        self.model_combo = ttk.Combobox(ai_frame, values=model_list, font=FONTS["combo"], width=30, state="readonly")
        self.model_combo.grid(row=row, column=1, sticky="w", padx=10, pady=8)
        # 设置当前选择的模型
        if self.ai.model_name in model_list:
            self.model_combo.set(self.ai.model_name)
        else:
            self.model_combo.set(model_list[0] if model_list else "")

        row += 1

        # API Key
        tk.Label(ai_frame, text="API Key：", font=FONTS["content_bold"], bg=COLORS["bg_card"]).grid(row=row, column=0, sticky="e", padx=10, pady=8)
        self.api_key_entry = tk.Entry(ai_frame, font=FONTS["combo"], width=0, show="*")
        self.api_key_entry.grid(row=row, column=1, sticky="ew", padx=10, pady=8)
        self.api_key_entry.insert(0, self.ai.api_key)
        row += 1

        # 当前班级名称
        tk.Label(ai_frame, text="当前班级名称：", font=FONTS["content_bold"], bg=COLORS["bg_card"]).grid(row=row, column=0, sticky="e", padx=10, pady=8)
        self.class_name_entry = tk.Entry(ai_frame, font=FONTS["combo"], width=20)
        self.class_name_entry.grid(row=row, column=1, sticky="w", padx=10, pady=8)
        self.class_name_entry.insert(0, self.ai.class_name)
        row += 1

        # 教材版本
        tk.Label(ai_frame, text="教材版本：", font=FONTS["content_bold"], bg=COLORS["bg_card"]).grid(row=row, column=0, sticky="e", padx=10, pady=8)
        self.textbook_entry = tk.Entry(ai_frame, font=FONTS["combo"], width=35)
        self.textbook_entry.grid(row=row, column=1, sticky="w", padx=10, pady=8)
        self.textbook_entry.insert(0, self.ai.textbook_version)
        row += 1

        # 地区
        tk.Label(ai_frame, text="地区：", font=FONTS["content_bold"], bg=COLORS["bg_card"]).grid(row=row, column=0, sticky="e", padx=10, pady=8)
        self.area_entry = tk.Entry(ai_frame, font=FONTS["combo"], width=35)
        self.area_entry.grid(row=row, column=1, sticky="w", padx=10, pady=8)
        self.area_entry.insert(0, self.ai.area)
        row += 1

        # 按钮行
        btn_ai = tk.Frame(ai_frame, bg=COLORS["bg_card"])
        btn_ai.grid(row=row, column=0, columnspan=2, pady=12)

        def save_ai():
            # 获取当前界面输入的值
            self.ai.api_key = self.api_key_entry.get().strip()
            self.ai.model_name = self.model_combo.get()
            self.ai.class_name = self.class_name_entry.get().strip()
            self.ai.textbook_version = self.textbook_entry.get().strip()
            self.ai.area = self.area_entry.get().strip()

            # 先保存配置
            self.ai.save_config()

            # 验证配置
            ok, msg = self.ai.verify_config()
            if ok:
                self.update_speak_buttons_state()
                messagebox.showinfo("保存成功", "✅ 配置已保存，测试连接成功！")
            else:
                # 验证失败，但配置已保存，提示失败原因
                messagebox.showwarning("保存成功但测试失败", f"⚠️ 配置已保存，但测试连接失败：\n{msg}")

        def test_conn():
            temp_key = self.api_key_entry.get().strip()
            temp_model = self.model_combo.get()
            old_key = self.ai.api_key
            old_model = self.ai.model_name
            self.ai.api_key = temp_key
            self.ai.model_name = temp_model
            ok, _ = self.ai.verify_config()  # 忽略详细消息
            self.ai.api_key = old_key
            self.ai.model_name = old_model
            if ok:
                messagebox.showinfo("测试结果", "✅ 连接成功")
            else:
                messagebox.showerror("测试结果", "❌ 连接失败，请检查API Key和模型名称")

        tk.Button(btn_ai, text="💾 保存并测试", font=FONTS["btn"], bg=COLORS["primary"], fg="white",bd=0, relief=tk.FLAT,
                  command=save_ai, width=14, cursor="hand2").pack(side=tk.LEFT, padx=10)
        tk.Button(btn_ai, text="🔌 连接测试", font=FONTS["btn"], bg=COLORS["secondary"], fg="white",bd=0, relief=tk.FLAT,
                  command=test_conn, width=12, cursor="hand2").pack(side=tk.LEFT, padx=10)
        row += 1

        # ===== AI 功能配置详细说明 =====
        desc_start = row
        tk.Label(ai_frame, text="📌 配置说明：", font=FONTS["content_bold"], bg=COLORS["bg_card"], fg=COLORS["primary"]).grid(row=desc_start, column=0, columnspan=2, sticky="w", padx=20, pady=(10,0))
        row += 1
        tk.Label(ai_frame, text="  • 选择模型：可从预设模型或自定义模型中选择，不同模型的速度和成本不同。",
                 font=FONTS["small"], bg=COLORS["bg_card"], fg=COLORS["text_secondary"], justify="left").grid(row=row, column=0, columnspan=2, sticky="w", padx=35)
        row += 1
        tk.Label(ai_frame, text="  • API Key：用于调用所选模型的AI服务。不同模型可能需要不同API Key，请根据模型配置填写。",
                 font=FONTS["small"], bg=COLORS["bg_card"], fg=COLORS["text_secondary"], justify="left").grid(row=row, column=0, columnspan=2, sticky="w", padx=35)
        row += 1
        tk.Label(ai_frame, text="  • 当前班级名称：用于学情分析和记录，区分不同班级的数据。",
                 font=FONTS["small"], bg=COLORS["bg_card"], fg=COLORS["text_secondary"], justify="left").grid(row=row, column=0, columnspan=2, sticky="w", padx=35)
        row += 1
        tk.Label(ai_frame, text="  • 教材版本和地区：AI生成题目和知识点时会参考本地教材和地域特点。",
                 font=FONTS["small"], bg=COLORS["bg_card"], fg=COLORS["text_secondary"], justify="left").grid(row=row, column=0, columnspan=2, sticky="w", padx=35)
        row += 1
        tk.Label(ai_frame, text="  ⚠️ 注意：请先保存并验证API Key有效性，否则AI功能无法使用。",
                 font=FONTS["small"], bg=COLORS["bg_card"], fg=COLORS["danger"], justify="left").grid(row=row, column=0, columnspan=2, sticky="w", padx=35, pady=(0,10))

        # ========== 模型管理与用量统计（左右布局，按钮移至底部居中） ==========
        model_frame = tk.LabelFrame(inner, text="📊 模型管理与用量统计", font=FONTS["nav"],
                                    bg=COLORS["bg_card"], fg=COLORS["primary"])
        model_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        # 使用 grid 布局：三行（左右栏、按钮行、说明行）
        model_frame.grid_rowconfigure(0, weight=1)      # 内容行（左右栏）可扩展
        model_frame.grid_rowconfigure(1, weight=0)      # 按钮行不扩展
        model_frame.grid_rowconfigure(2, weight=0)      # 说明行不扩展
        model_frame.grid_columnconfigure(0, weight=4)   # 左侧占3份
        model_frame.grid_columnconfigure(1, weight=1)   # 右侧占2份

        # ========== 左侧：模型列表 ==========
        left_frame = tk.Frame(model_frame, bg=COLORS["bg_card"])
        left_frame.grid(row=0, column=0, sticky="nsew", padx=(0,5), pady=0)

        # Treeview 列宽自适应
        columns = ("name", "quota_type", "limit", "used", "remaining")
        self.model_tree = ttk.Treeview(left_frame, columns=columns, show="headings", height=10)
        self.model_tree.heading("name", text="模型名称")
        self.model_tree.heading("quota_type", text="周期")
        self.model_tree.heading("limit", text="额度上限")
        self.model_tree.heading("used", text="已用tokens")
        self.model_tree.heading("remaining", text="剩余tokens")

        def adjust_tree_columns(event=None):
            total_width = self.model_tree.winfo_width()
            if total_width < 100:
                return
            # 名称列只占15%，周期10%，额度上限20%，已用20%，剩余35%
            self.model_tree.column("name", width=int(total_width * 0.30), minwidth=80)
            self.model_tree.column("quota_type", width=int(total_width * 0.10), minwidth=50)
            self.model_tree.column("limit", width=int(total_width * 0.20), minwidth=70)
            self.model_tree.column("used", width=int(total_width * 0.20), minwidth=70)
            self.model_tree.column("remaining", width=int(total_width * 0.20), minwidth=80)

        self.model_tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.model_tree.bind("<Configure>", adjust_tree_columns)
        self.root.after(100, adjust_tree_columns)

        # ========== 右侧：模型用量详情 ==========
        right_frame = tk.Frame(model_frame, bg=COLORS["bg_card"])
        right_frame.grid(row=0, column=1, sticky="nsew", padx=(5,0), pady=0)
        right_frame.grid_rowconfigure(0, weight=1)
        right_frame.grid_columnconfigure(0, weight=1)

        # 右侧滚动区域
        right_canvas = tk.Canvas(right_frame, bg=COLORS["bg_card"], highlightthickness=0)
        h_scroll = ttk.Scrollbar(right_frame, orient="horizontal", command=right_canvas.xview)
        v_scroll = ttk.Scrollbar(right_frame, orient="vertical", command=right_canvas.yview)
        right_canvas.configure(xscrollcommand=h_scroll.set, yscrollcommand=v_scroll.set)
        right_canvas.grid(row=0, column=0, sticky="nsew")
        h_scroll.grid(row=1, column=0, sticky="ew")
        v_scroll.grid(row=0, column=1, sticky="ns")

        right_inner = tk.Frame(right_canvas, bg=COLORS["bg_card"])
        canvas_win = right_canvas.create_window((0, 0), window=right_inner, anchor="nw")

        def on_right_resize(event):
            right_canvas.configure(scrollregion=right_canvas.bbox("all"))
            if event.widget == right_canvas:
                width = event.width
            else:
                width = right_canvas.winfo_width()
            if width > 10:
                right_canvas.itemconfig(canvas_win, width=width)

        right_inner.bind("<Configure>", on_right_resize)
        right_canvas.bind("<Configure>", on_right_resize)

        # 周期用量
        quota_frame = tk.LabelFrame(right_inner, text="📈 周期用量", font=FONTS["content_bold"],
                                    bg=COLORS["bg_card"], fg=COLORS["primary"])
        quota_frame.pack(fill=tk.X, padx=10, pady=5)
        self.model_detail_info = tk.Label(quota_frame, text="", font=FONTS["content"], bg=COLORS["bg_card"],
                                          justify="left", anchor="w", wraplength=450)
        self.model_detail_info.pack(fill=tk.X, padx=10, pady=5)

        # 按场景统计
        scene_frame = tk.LabelFrame(right_inner, text="📊 按场景统计", font=FONTS["content_bold"],
                                    bg=COLORS["bg_card"], fg=COLORS["primary"])
        scene_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        tree_container = tk.Frame(scene_frame, bg=COLORS["bg_card"])
        tree_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        tree_container.grid_rowconfigure(0, weight=1)
        tree_container.grid_columnconfigure(0, weight=1)

        tree_h_scroll = ttk.Scrollbar(tree_container, orient="horizontal")
        tree_v_scroll = ttk.Scrollbar(tree_container, orient="vertical")
        self.scene_tree = ttk.Treeview(tree_container, columns=("scene", "requests", "tokens"),
                                       show="headings", height=6,
                                       xscrollcommand=tree_h_scroll.set,
                                       yscrollcommand=tree_v_scroll.set)
        tree_h_scroll.config(command=self.scene_tree.xview)
        tree_v_scroll.config(command=self.scene_tree.yview)

        self.scene_tree.heading("scene", text="场景")
        self.scene_tree.heading("requests", text="请求次数")
        self.scene_tree.heading("tokens", text="总tokens")

        def adjust_scene_columns(event=None):
            total_width = self.scene_tree.winfo_width()
            if total_width < 200:
                return
            self.scene_tree.column("scene", width=int(total_width * 0.35), minwidth=60)
            self.scene_tree.column("requests", width=int(total_width * 0.20), minwidth=80)
            self.scene_tree.column("tokens", width=int(total_width * 0.45), minwidth=80)

        self.scene_tree.bind("<Configure>", adjust_scene_columns)
        self.root.after(100, adjust_scene_columns)

        self.scene_tree.grid(row=0, column=0, sticky="nsew")
        tree_h_scroll.grid(row=1, column=0, sticky="ew")
        tree_v_scroll.grid(row=0, column=1, sticky="ns")

        # ========== 底部按钮行（使用 pack 居中，减小间距） ==========
        btn_bottom = tk.Frame(model_frame, bg=COLORS["bg_card"])
        btn_bottom.grid(row=1, column=0, columnspan=2, sticky="ew", pady=8)
        # 内部使用 pack 居中
        btn_inner = tk.Frame(btn_bottom, bg=COLORS["bg_card"])
        btn_inner.pack(anchor="center")
        tk.Button(btn_inner, text="🔄 刷新列表", font=FONTS["btn"], bg=COLORS["secondary"], fg="white", bd=0, relief=tk.FLAT,
                  command=self.refresh_model_list, width=10, cursor="hand2").pack(side=tk.LEFT, padx=6)
        tk.Button(btn_inner, text="➕ 添加模型", font=FONTS["btn"], bg=COLORS["primary"], fg="white", bd=0, relief=tk.FLAT,
                  command=self.add_model_dialog, width=10, cursor="hand2").pack(side=tk.LEFT, padx=6)
        tk.Button(btn_inner, text="❌ 删除模型", font=FONTS["btn"], bg=COLORS["danger"], fg="white", bd=0, relief=tk.FLAT,
                  command=self.delete_selected_model, width=10, cursor="hand2").pack(side=tk.LEFT, padx=6)
        tk.Button(btn_inner, text="🗑️ 清除API记录", font=FONTS["btn"], bg=COLORS["warning"], fg="white", bd=0, relief=tk.FLAT,
                  command=self.clear_api_usage, width=12, cursor="hand2").pack(side=tk.LEFT, padx=6)
        tk.Button(btn_inner, text="🔄 重置用量", font=FONTS["btn"], bg=COLORS["success"], fg="white", bd=0, relief=tk.FLAT,
                  command=self.reset_selected_model_quota, width=10, cursor="hand2").pack(side=tk.LEFT, padx=6)


        # ========== 功能说明（放在最底部） ==========
        desc_frame = tk.Frame(model_frame, bg=COLORS["bg_card"])
        desc_frame.grid(row=2, column=0, columnspan=2, sticky="ew", pady=8)
        tk.Label(desc_frame, text="📌 功能说明：", font=FONTS["content_bold"], bg=COLORS["bg_card"], fg=COLORS["primary"]).pack(anchor="w", padx=(20,0))
        tk.Label(desc_frame, text="  • 左侧模型管理：支持添加自定义模型（名称、API地址、周期、额度），可删除模型。",
                 font=FONTS["small"], bg=COLORS["bg_card"], fg=COLORS["text_secondary"], justify="left", anchor="w").pack(anchor="w", padx=(35,0))
        tk.Label(desc_frame, text="  • 周期用量：显示当前模型本周期已用tokens和剩余额度，每日/每月自动重置。",
                 font=FONTS["small"], bg=COLORS["bg_card"], fg=COLORS["text_secondary"], justify="left", anchor="w").pack(anchor="w", padx=(35,0))
        tk.Label(desc_frame, text="  • 按场景统计：统计该模型在不同AI场景（答疑、出题等）的请求次数和token消耗。",
                 font=FONTS["small"], bg=COLORS["bg_card"], fg=COLORS["text_secondary"], justify="left", anchor="w").pack(anchor="w", padx=(35,0))
        tk.Label(desc_frame, text="  ⚠️ 注意：清除API记录将删除所有场景统计数据，但不影响模型额度。",
                 font=FONTS["small"], bg=COLORS["bg_card"], fg=COLORS["danger"], justify="left", anchor="w").pack(anchor="w", padx=(35,0))

        # 绑定事件
        self.model_tree.bind("<<TreeviewSelect>>", self.on_model_select)
        self.refresh_model_list()

        # ========== 智能语音设置 ==========
        voice_frame = tk.LabelFrame(inner, text="🔊 智能语音设置", font=FONTS["nav"],
                                    bg=COLORS["bg_card"], fg=COLORS["primary"])
        voice_frame.pack(fill=tk.X, padx=20, pady=10)

        # ---- 语音模式（离线/在线） ----
        mode_frame = tk.Frame(voice_frame, bg=COLORS["bg_card"])
        mode_frame.pack(pady=10)
        mode_inner = tk.Frame(mode_frame, bg=COLORS["bg_card"])
        mode_inner.pack(anchor="center")

        tk.Label(mode_inner, text="语音模式：", font=FONTS["content_bold"], bg=COLORS["bg_card"]).pack(side=tk.LEFT,
                                                                                                      padx=5)

        self.offline_tts_btn = tk.Button(
            mode_inner, text="离线语音", font=FONTS["btn"],
            bg=COLORS["primary"], fg="white", bd=0, relief=tk.FLAT, cursor="hand2",
            width=10, command=lambda: self.set_tts_mode('offline')
        )
        self.offline_tts_btn.pack(side=tk.LEFT, padx=5)

        self.online_tts_btn = tk.Button(
            mode_inner, text="在线语音", font=FONTS["btn"],
            bg=COLORS["bg_card"], fg=COLORS["text_main"], bd=1, relief=tk.RAISED, cursor="hand2",
            width=10, command=lambda: self.set_tts_mode('online')
        )
        self.online_tts_btn.pack(side=tk.LEFT, padx=5)

        # ---- 在线音色（按钮组） ----
        voice_type_frame = tk.Frame(voice_frame, bg=COLORS["bg_card"])
        voice_type_frame.pack(pady=10)
        voice_inner = tk.Frame(voice_type_frame, bg=COLORS["bg_card"])
        voice_inner.pack(anchor="center")

        tk.Label(voice_inner, text="在线音色：", font=FONTS["content_bold"], bg=COLORS["bg_card"]).pack(side=tk.LEFT,
                                                                                                       padx=5)

        # 定义音色列表
        voice_list = [
            ("女声-自然", "zh-CN-XiaoxiaoNeural"),
            ("女声-活泼", "zh-CN-XiaoyiNeural"),
            ("男声-温润", "zh-CN-YunxiNeural"),
            ("男声-新闻", "zh-CN-YunyangNeural"),
            ("东北话", "zh-CN-liaoning-XiaobeiNeural"),
        ]

        self.voice_buttons = []  # 存储按钮对象
        self.voice_btn_code = []  # 存储对应的音色代码
        for display, code in voice_list:
            btn = tk.Button(
                voice_inner, text=display, font=FONTS["btn"],
                bg=COLORS["bg_card"], fg=COLORS["text_main"], bd=1, relief=tk.RAISED, cursor="hand2",
                width=12,
                command=lambda c=code: self.select_tts_voice(c)
            )
            btn.pack(side=tk.LEFT, padx=5)
            self.voice_buttons.append(btn)
            self.voice_btn_code.append(code)

        # 根据当前模式设置音色按钮的启用/禁用状态（初始化时离线模式 -> 禁用）
        if self.tts.mode == 'offline':
            for btn in self.voice_buttons:
                btn.config(state=tk.DISABLED, bg=COLORS["bg_card"], fg=COLORS["text_secondary"])
        else:
            for btn in self.voice_buttons:
                btn.config(state=tk.NORMAL, bg=COLORS["bg_card"], fg=COLORS["text_main"])
            # 高亮当前选中的音色
            self.highlight_selected_voice(self.pending_tts_voice)

        # ---- 底部按钮（确认 + 恢复默认） ----
        bottom_frame = tk.Frame(voice_frame, bg=COLORS["bg_card"])
        bottom_frame.pack(pady=15)
        bottom_inner = tk.Frame(bottom_frame, bg=COLORS["bg_card"])
        bottom_inner.pack(anchor="center")

        self.apply_tts_btn = tk.Button(
            bottom_inner, text="✅ 确认切换", font=FONTS["btn"],
            bg=COLORS["success"], fg="white", bd=0, relief=tk.FLAT, cursor="hand2",
            width=12, command=self.apply_tts_settings
        )
        self.apply_tts_btn.pack(side=tk.LEFT, padx=10)

        self.reset_tts_btn = tk.Button(
            bottom_inner, text="🔄 恢复默认", font=FONTS["btn"],
            bg=COLORS["warning"], fg="white", bd=0, relief=tk.FLAT, cursor="hand2",
            width=12, command=self.reset_tts_settings
        )
        self.reset_tts_btn.pack(side=tk.LEFT, padx=10)

        # ---- 使用说明 ----
        usage_frame = tk.Frame(voice_frame, bg=COLORS["bg_card"])
        usage_frame.pack(fill=tk.X, padx=20, pady=5)
        tk.Label(usage_frame, text="📌 使用说明：", font=FONTS["content_bold"], bg=COLORS["bg_card"], fg=COLORS["primary"]).pack(anchor="w")
        tk.Label(usage_frame, text="  • 离线语音（默认）：无需网络，响应快，音质一般。",
                 font=FONTS["small"], bg=COLORS["bg_card"], fg=COLORS["text_secondary"], justify="left", anchor="w").pack(anchor="w", padx=(15,0))
        tk.Label(usage_frame, text="  • 在线语音：需联网，音质佳，可换音色。可能会有点慢（1-5秒），请耐心等待。",
                 font=FONTS["small"], bg=COLORS["bg_card"], fg=COLORS["text_secondary"], justify="left", anchor="w").pack(anchor="w", padx=(15,0))
        tk.Label(usage_frame, text="  • 如果在线语音网络不佳，程序会自动降级为离线语音（会弹出提示）。",
                 font=FONTS["small"], bg=COLORS["bg_card"], fg=COLORS["text_secondary"], justify="left", anchor="w").pack(anchor="w", padx=(15,0))
        tk.Label(usage_frame, text="  • 右键点击任意文本框中的文字，选择“朗读选中内容”即可朗读。",
                 font=FONTS["small"], bg=COLORS["bg_card"], fg=COLORS["text_secondary"], justify="left", anchor="w").pack(anchor="w", padx=(15,0))

        # ========== 3. 数据导入导出 ==========
        data_frame = tk.LabelFrame(inner, text="📂 数据导入导出", font=FONTS["nav"],
                                   bg=COLORS["bg_card"], fg=COLORS["primary"])
        data_frame.pack(fill=tk.X, padx=20, pady=10)

        data_frame.grid_columnconfigure(0, weight=1)

        # 第一行：导出全部数据、导出学情报告
        row1 = tk.Frame(data_frame, bg=COLORS["bg_card"])
        row1.grid(row=0, column=0, pady=5)
        tk.Button(row1, text="📤 导出全部数据", font=FONTS["btn"], bg=COLORS["primary"], fg="white", bd=0, relief=tk.FLAT,
                  command=self.export_all_data_to_excel, width=16, cursor="hand2").pack(side=tk.LEFT, padx=10)
        tk.Button(row1, text="📊 导出学情报告", font=FONTS["btn"], bg=COLORS["success"], fg="white", bd=0, relief=tk.FLAT,
                  command=self.export_class_report, width=16, cursor="hand2").pack(side=tk.LEFT, padx=10)

        # 第二行：导出错题本、导入章节/考点
        row2 = tk.Frame(data_frame, bg=COLORS["bg_card"])
        row2.grid(row=1, column=0, pady=5)
        tk.Button(row2, text="📘 导出错题本", font=FONTS["btn"], bg=COLORS["secondary"], fg="white", bd=0, relief=tk.FLAT,
                  command=self.export_wrong_questions, width=16, cursor="hand2").pack(side=tk.LEFT, padx=10)
        tk.Button(row2, text="📂 导入章节/考点", font=FONTS["btn"], bg=COLORS["primary"], fg="white", bd=0, relief=tk.FLAT,
                  command=self.import_chapters_kps_from_excel, width=16, cursor="hand2").pack(side=tk.LEFT, padx=10)

        # 第三行：备份数据库、恢复数据库
        row3 = tk.Frame(data_frame, bg=COLORS["bg_card"])
        row3.grid(row=2, column=0, pady=5)
        tk.Button(row3, text="💾 备份数据库", font=FONTS["btn"], bg=COLORS["warning"], fg="white", bd=0, relief=tk.FLAT,
                  command=self.backup_database, width=16, cursor="hand2").pack(side=tk.LEFT, padx=10)
        tk.Button(row3, text="🔄 恢复数据库", font=FONTS["btn"], bg=COLORS["danger"], fg="white", bd=0, relief=tk.FLAT,
                  command=self.restore_database, width=16, cursor="hand2").pack(side=tk.LEFT, padx=10)

        # 统一的功能说明（所有说明左对齐，缩进35像素，每条以 • 开头，备份恢复合并为一行）
        desc_frame = tk.Frame(data_frame, bg=COLORS["bg_card"])
        desc_frame.grid(row=3, column=0, pady=8, sticky="w")
        tk.Label(desc_frame, text="📌 功能说明：", font=FONTS["content_bold"], bg=COLORS["bg_card"], fg=COLORS["primary"]).pack(anchor="w", padx=(20,0))
        # 每条说明单独一行，缩进一致
        tk.Label(desc_frame, text="  • 导出全部数据：将 AI问答记录、答题记录、答题明细、MSLQ测评记录导出到一个Excel文件（多个工作表）。",
                 font=FONTS["small"], bg=COLORS["bg_card"], fg=COLORS["text_secondary"], justify="left", anchor="w").pack(anchor="w", padx=(35,0))
        tk.Label(desc_frame, text="  • 导出学情报告：基于当前班级、科目、年级，生成章节正确率、考点错误率、题型分析、典型错题报告。",
                 font=FONTS["small"], bg=COLORS["bg_card"], fg=COLORS["text_secondary"], justify="left", anchor="w").pack(anchor="w", padx=(35,0))
        tk.Label(desc_frame, text="  • 导出错题本：导出错误次数≥2的典型错题（含题目、标准答案、错误次数）。",
                 font=FONTS["small"], bg=COLORS["bg_card"], fg=COLORS["text_secondary"], justify="left", anchor="w").pack(anchor="w", padx=(35,0))
        tk.Label(desc_frame, text="  • 导入章节/考点：通过Excel文件自定义章节结构和考点（替换默认smart_classroom_data.xlsx内容，需重启）。",
                 font=FONTS["small"], bg=COLORS["bg_card"], fg=COLORS["text_secondary"], justify="left", anchor="w").pack(anchor="w", padx=(35,0))
        # 备份和恢复合并为一行
        tk.Label(desc_frame, text="  • 备份数据库 / 恢复数据库：备份当前所有数据为带时间戳的文件，或从备份文件恢复（恢复需重启）。",
                 font=FONTS["small"], bg=COLORS["bg_card"], fg=COLORS["text_secondary"], justify="left", anchor="w").pack(anchor="w", padx=(35,0))
        tk.Label(desc_frame, text="  ⚠️ 注意：导出功能需要至少有一道答题记录，否则会提示无数据。",
                 font=FONTS["small"], bg=COLORS["bg_card"], fg=COLORS["danger"], justify="left", anchor="w").pack(anchor="w", padx=(35,0))

        # ========== 4. 批量任务 ==========
        batch_frame = tk.LabelFrame(inner, text="⚙️ 批量任务", font=FONTS["nav"],
                                    bg=COLORS["bg_card"], fg=COLORS["primary"])
        batch_frame.pack(fill=tk.X, padx=20, pady=10)
        def batch_with_warning():
            if messagebox.askyesno("配额提醒", "批量生成知识点会消耗大量 API Token，请确保您的 API Key 有足够配额。\n\n是否继续？"):
                self.batch_generate_knowledge()
        tk.Button(batch_frame, text="⚙️ 批量生成章节知识点", font=FONTS["btn"], bg=COLORS["primary"], fg="white", bd=0, relief=tk.FLAT,
                  command=batch_with_warning, width=25, cursor="hand2").pack(pady=10)
        # 说明
        batch_desc = tk.Frame(batch_frame, bg=COLORS["bg_card"])
        batch_desc.pack(pady=5, anchor="w", fill=tk.X, padx=20)
        tk.Label(batch_desc, text="📌 功能说明：", font=FONTS["content_bold"], bg=COLORS["bg_card"], fg=COLORS["primary"]).pack(anchor="w")
        tk.Label(batch_desc, text="  • 为当前科目、年级下的所有章节/课时自动生成知识点内容（使用AI）。",
                 font=FONTS["small"], bg=COLORS["bg_card"], fg=COLORS["text_secondary"], justify="left", anchor="w").pack(anchor="w", padx=(15,0))
        tk.Label(batch_desc, text="  • 该操作会消耗较多API额度，请确保配额充足，生成过程会显示进度窗口。",
                 font=FONTS["small"], bg=COLORS["bg_card"], fg=COLORS["text_secondary"], justify="left", anchor="w").pack(anchor="w", padx=(15,0))
        tk.Label(batch_desc, text="  ⚠️ 注意：仅对尚无知识点的章节生成，已有内容会被跳过（避免重复消耗）。",
                 font=FONTS["small"], bg=COLORS["bg_card"], fg=COLORS["danger"], justify="left", anchor="w").pack(anchor="w", padx=(15,0))

        # ========== 5. 界面字体大小 ==========
        font_frame = tk.LabelFrame(inner, text="🔤 UI字体大小", font=FONTS["nav"],
                                   bg=COLORS["bg_card"], fg=COLORS["primary"])
        font_frame.pack(fill=tk.X, padx=20, pady=10)

        # 控件行
        control_frame = tk.Frame(font_frame, bg=COLORS["bg_card"])
        control_frame.pack(pady=5)
        tk.Label(control_frame, text="选择基础字体大小：", font=FONTS["content"], bg=COLORS["bg_card"]).pack(
            side=tk.LEFT)
        font_var = tk.IntVar(value=14)
        font_spin = tk.Spinbox(control_frame, from_=10, to=22, textvariable=font_var, width=5, font=FONTS["content"])
        font_spin.pack(side=tk.LEFT, padx=10)

        def set_font():
            new_size = font_var.get()
            self.apply_font_size(new_size)
            if messagebox.askyesno("重启确认", f"字体大小已设置为 {new_size} 号，需要重启程序才能生效。是否立即重启？"):
                python = sys.executable
                os.execl(python, python, *sys.argv)

        tk.Button(control_frame, text="🔄 应用并重启", font=FONTS["btn"], bg=COLORS["success"], fg="white", bd=0, relief=tk.FLAT,
                  command=set_font, cursor="hand2", width=12).pack(side=tk.LEFT, padx=10)

        # 提示说明（统一缩进格式）
        tip_frame = tk.Frame(font_frame, bg=COLORS["bg_card"])
        tip_frame.pack(pady=5, anchor="w", fill=tk.X, padx=20)
        tk.Label(tip_frame, text="📌 使用说明：", font=FONTS["content_bold"], bg=COLORS["bg_card"],
                 fg=COLORS["primary"]).pack(anchor="w")
        tk.Label(tip_frame, text="  • 默认字体大小为14号，调整后需要重启程序才能完全生效。",
                 font=FONTS["small"], bg=COLORS["bg_card"], fg=COLORS["text_secondary"], justify="left",
                 anchor="w").pack(anchor="w", padx=(15, 0))

        # ---- 关于信息 ----
        about_frame = tk.LabelFrame(inner, text="📌 关于", font=FONTS["nav"],
                                    bg=COLORS["bg_card"], fg=COLORS["primary"])
        about_frame.pack(fill=tk.X, padx=20, pady=10)
        inner_about = tk.Frame(about_frame, bg=COLORS["bg_card"])
        inner_about.pack(fill=tk.X, padx=10, pady=10)
        tk.Label(inner_about, text="    • 初中智慧课堂AI学习助手  v1.5",
                 font=FONTS["small"], bg=COLORS["bg_card"], fg=COLORS["text_secondary"], anchor="w").pack(anchor="w", pady=2)
        tk.Label(inner_about, text="    • 作者：邵长超",
                 font=FONTS["small"], bg=COLORS["bg_card"], fg=COLORS["text_secondary"], anchor="w").pack(anchor="w", pady=2)
        tk.Label(inner_about, text="    • 单位：沂水县第二实验中学",
                 font=FONTS["small"], bg=COLORS["bg_card"], fg=COLORS["text_secondary"], anchor="w").pack(anchor="w", pady=2)
        tk.Label(inner_about, text="    • 联系方式：977258816@qq.com",
                 font=FONTS["small"], bg=COLORS["bg_card"], fg=COLORS["text_secondary"], anchor="w").pack(anchor="w", pady=2)
        tk.Label(inner_about, text="     ",
                 font=FONTS["small"], bg=COLORS["bg_card"], fg=COLORS["text_secondary"], anchor="w").pack(anchor="w", pady=2)

        return outer

    # ==================== 设置扩展功能 ====================
    def export_all_data_to_excel(self):
        """导出所有业务表到 Excel（多工作表）"""

        # 确认对话框
        if not messagebox.askyesno("确认导出", "导出全部数据将生成一个包含所有记录的Excel文件，耗时可能较长。\n\n是否继续？"):
            return

        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror("错误", "请先安装 openpyxl 库：pip install openpyxl")
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"全量数据备份_{timestamp}.xlsx"
        wb = openpyxl.Workbook()
        # 删除默认 sheet
        wb.remove(wb.active)

        conn, cursor = self.db._connect()

        # 表定义：(sheet名称, 查询SQL, 列标题)
        tables = [
            ("AI问答记录", "SELECT * FROM ai_chat_records ORDER BY ask_time DESC",
             ["ID", "提问时间", "科目", "年级", "场景", "问题", "答案"]),
            ("答题记录", "SELECT * FROM exam_records ORDER BY exam_time DESC",
             ["ID", "考试时间", "科目", "年级", "章节", "课时", "总题数", "正确数", "正确率"]),
            ("答题明细", "SELECT * FROM exam_details ORDER BY create_time DESC",
             ["ID", "考试记录ID", "班级", "科目", "年级", "章节", "课时", "题目", "标准答案", "用户答案", "是否正确", "知识点", "知识点ID", "题型", "创建时间"]),
            ("MSLQ测评记录", "SELECT * FROM mslq_records ORDER BY create_time DESC",
             ["ID", "学生姓名", "班级", "测试类型", "总分", "详细得分", "创建时间"]),
        ]

        for sheet_name, sql, headers in tables:
            cursor.execute(sql)
            rows = cursor.fetchall()
            ws = wb.create_sheet(title=sheet_name)
            ws.append(headers)
            for row in rows:
                ws.append(list(row))
            # 自动调整列宽
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_len + 2, 30)
                ws.column_dimensions[col_letter].width = adjusted_width

        self.db._close(conn)
        wb.save(filename)
        messagebox.showinfo("导出成功", f"已保存到程序所在目录：{filename}")

    def export_class_report(self):
        """导出当前班级学情报告（带科目年级选择）"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("错误", "请先安装 openpyxl 库：pip install openpyxl")
            return

        # 弹出选择科目年级对话框（居中、大小合适）
        dialog = tk.Toplevel(self.root)
        dialog.title("选择年级和科目")
        dialog.geometry("350x250")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg=COLORS["bg_main"])
        # 窗口居中
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (350 // 2)
        y = (dialog.winfo_screenheight() // 2) - (250 // 2)
        dialog.geometry(f"+{x}+{y}")

        main_frame = tk.Frame(dialog, bg=COLORS["bg_main"])
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)

        tk.Label(main_frame, text="年级：", font=FONTS["content_bold"], bg=COLORS["bg_main"]).pack(anchor="w", pady=5)
        grade_var = tk.StringVar(value="七年级")
        grade_combo = ttk.Combobox(main_frame, textvariable=grade_var, values=ALL_GRADES, width=25, state="readonly")
        grade_combo.pack(fill=tk.X, pady=5)

        tk.Label(main_frame, text="科目：", font=FONTS["content_bold"], bg=COLORS["bg_main"]).pack(anchor="w", pady=5)
        subject_var = tk.StringVar(value="语文")
        subject_combo = ttk.Combobox(main_frame, textvariable=subject_var, values=ALL_SUBJECTS[1:], width=25, state="readonly")
        subject_combo.pack(fill=tk.X, pady=5)

        result = {"subject": None, "grade": None}
        def on_confirm():
            result["subject"] = subject_var.get()
            result["grade"] = grade_var.get()
            dialog.destroy()
        def on_cancel():
            dialog.destroy()

        btn_frame = tk.Frame(main_frame, bg=COLORS["bg_main"])
        btn_frame.pack(pady=15)
        tk.Button(btn_frame, text="确定", command=on_confirm, width=10, cursor="hand2").pack(side=tk.LEFT, padx=15)
        tk.Button(btn_frame, text="取消", command=on_cancel, width=10, cursor="hand2").pack(side=tk.LEFT, padx=15)

        self.root.wait_window(dialog)
        if result["subject"] is None:
            return

        subject = result["subject"]
        grade = result["grade"]
        class_name = self.ai.class_name

        conn, cursor = self.db._connect()
        try:
            # 1. 章节正确率
            cursor.execute("""
                SELECT chapter, COUNT(*) as total, SUM(is_correct) as correct
                FROM exam_details
                WHERE class_name=? AND subject=? AND grade=?
                GROUP BY chapter
                ORDER BY chapter
            """, (class_name, subject, grade))
            chapter_rows = cursor.fetchall()

            # 2. 考点错误率
            cursor.execute("""
                SELECT knowledge_point_id, COUNT(*) as total, SUM(is_correct) as correct
                FROM exam_details
                WHERE class_name=? AND subject=? AND grade=?
                GROUP BY knowledge_point_id
                ORDER BY (total - correct) DESC
            """, (class_name, subject, grade))
            kp_rows = cursor.fetchall()

            # 3. 题型正确率
            cursor.execute("""
                SELECT question_type, COUNT(*) as total, SUM(is_correct) as correct
                FROM exam_details
                WHERE class_name=? AND subject=? AND grade=?
                GROUP BY question_type
            """, (class_name, subject, grade))
            type_rows = cursor.fetchall()

            # 4. 典型错题
            cursor.execute("""
                SELECT question, correct_answer, COUNT(*) as wrong_count
                FROM exam_details
                WHERE class_name=? AND subject=? AND grade=? AND is_correct=0
                GROUP BY question
                HAVING wrong_count >= 2
                ORDER BY wrong_count DESC
                LIMIT 10
            """, (class_name, subject, grade))
            wrong_rows = cursor.fetchall()
        except Exception as e:
            messagebox.showerror("查询失败", f"读取数据时出错：{str(e)}")
            return
        finally:
            self.db._close(conn)

        total_questions = sum(r['total'] for r in chapter_rows) if chapter_rows else 0
        if total_questions == 0:
            messagebox.showinfo("提示", f"当前班级“{class_name}”在 {subject} {grade} 中没有答题数据，无法生成报告。")
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"学情报告_{class_name}_{subject}_{grade}_{timestamp}.xlsx"
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")

        ws1 = wb.create_sheet("章节正确率")
        ws1.append(["章节", "总题数", "正确数", "正确率"])
        for r in chapter_rows:
            rate = (r['correct'] / r['total']) * 100 if r['total'] > 0 else 0
            ws1.append([r['chapter'], r['total'], r['correct'], f"{rate:.1f}%"])
        for col in range(1, 5):
            cell = ws1.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        ws2 = wb.create_sheet("考点错误率")
        ws2.append(["考点", "总题数", "正确数", "错误率"])
        for kp in kp_rows:
            kp_id = kp['knowledge_point_id']
            if kp_id and kp_id in KNOWLEDGE_POINTS:
                kp_name = KNOWLEDGE_POINTS[kp_id]['name']
            else:
                kp_name = kp_id if kp_id else "未归类"
            error_rate = (kp['total'] - kp['correct']) / kp['total'] * 100 if kp['total'] > 0 else 0
            ws2.append([kp_name, kp['total'], kp['correct'], f"{error_rate:.1f}%"])
        for col in range(1, 5):
            cell = ws2.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill

        ws3 = wb.create_sheet("题型正确率")
        ws3.append(["题型", "总题数", "正确数", "正确率"])
        for t in type_rows:
            rate = t['correct'] / t['total'] * 100 if t['total'] > 0 else 0
            ws3.append([t['question_type'], t['total'], t['correct'], f"{rate:.1f}%"])
        for col in range(1, 5):
            cell = ws3.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill

        ws4 = wb.create_sheet("典型错题")
        ws4.append(["题目", "标准答案", "错误次数"])
        for w in wrong_rows:
            ws4.append([w['question'], w['correct_answer'], w['wrong_count']])
        for col in range(1, 4):
            cell = ws4.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill

        wb.save(filename)
        messagebox.showinfo("导出成功", f"学情报告已保存到程序目录：{filename}")

    def export_wrong_questions(self):
        """导出典型错题（带科目年级选择）"""
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("错误", "请先安装 openpyxl 库")
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("选择年级和科目")
        dialog.geometry("350x250")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg=COLORS["bg_main"])
        # 窗口居中
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (350 // 2)
        y = (dialog.winfo_screenheight() // 2) - (250 // 2)
        dialog.geometry(f"+{x}+{y}")

        main_frame = tk.Frame(dialog, bg=COLORS["bg_main"])
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)

        tk.Label(main_frame, text="年级：", font=FONTS["content_bold"], bg=COLORS["bg_main"]).pack(anchor="w", pady=5)
        grade_var = tk.StringVar(value="七年级")
        grade_combo = ttk.Combobox(main_frame, textvariable=grade_var, values=ALL_GRADES, width=25, state="readonly")
        grade_combo.pack(fill=tk.X, pady=5)

        tk.Label(main_frame, text="科目：", font=FONTS["content_bold"], bg=COLORS["bg_main"]).pack(anchor="w", pady=5)
        subject_var = tk.StringVar(value="语文")
        subject_combo = ttk.Combobox(main_frame, textvariable=subject_var, values=ALL_SUBJECTS[1:], width=25, state="readonly")
        subject_combo.pack(fill=tk.X, pady=5)

        result = {"subject": None, "grade": None}
        def on_confirm():
            result["subject"] = subject_var.get()
            result["grade"] = grade_var.get()
            dialog.destroy()
        def on_cancel():
            dialog.destroy()

        btn_frame = tk.Frame(main_frame, bg=COLORS["bg_main"])
        btn_frame.pack(pady=15)
        tk.Button(btn_frame, text="确定", command=on_confirm, width=10, cursor="hand2").pack(side=tk.LEFT, padx=15)
        tk.Button(btn_frame, text="取消", command=on_cancel, width=10, cursor="hand2").pack(side=tk.LEFT, padx=15)

        self.root.wait_window(dialog)
        if result["subject"] is None:
            return

        subject = result["subject"]
        grade = result["grade"]
        class_name = self.ai.class_name

        conn, cursor = self.db._connect()
        try:
            cursor.execute("""
                SELECT question, correct_answer, COUNT(*) as wrong_count
                FROM exam_details
                WHERE class_name=? AND subject=? AND grade=? AND is_correct=0
                GROUP BY question
                HAVING wrong_count >= 2
                ORDER BY wrong_count DESC
            """, (class_name, subject, grade))
            rows = cursor.fetchall()
        except Exception as e:
            messagebox.showerror("查询失败", f"读取错题数据时出错：{str(e)}")
            return
        finally:
            self.db._close(conn)

        if not rows:
            messagebox.showinfo("提示", f"当前班级“{class_name}”在 {subject} {grade} 中没有典型错题（错误次数≥2）。")
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"错题本_{class_name}_{subject}_{grade}_{timestamp}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "典型错题"
        ws.append(["题目", "标准答案", "错误次数"])
        for r in rows:
            ws.append([r['question'], r['correct_answer'], r['wrong_count']])
        for col in ws.columns:
            max_len = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            adjusted_width = min(max_len + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width

        wb.save(filename)
        messagebox.showinfo("导出成功", f"错题本已保存到程序目录：{filename}")

    def backup_database(self):
        """备份当前数据库"""
        import shutil

        # 确认对话框
        if not messagebox.askyesno("确认备份", "备份数据库将保存当前所有数据到带时间戳的备份文件。\n\n是否继续？"):
            return

        source = DB_FILE
        if not os.path.exists(source):
            messagebox.showerror("错误", "数据库文件不存在")
            return
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"smart_classroom_backup_{timestamp}.db"
        try:
            shutil.copy2(source, backup_name)
            messagebox.showinfo("备份成功", f"数据库已备份到：{backup_name}")
        except Exception as e:
            messagebox.showerror("备份失败", str(e))

    def restore_database(self):
        """恢复数据库（需重启）"""
        from tkinter import filedialog
        import shutil
        file_path = filedialog.askopenfilename(filetypes=[("数据库文件", "*.db")])
        if not file_path:
            return
        if not messagebox.askyesno("确认恢复", "恢复将覆盖当前所有数据，并自动重启程序。是否继续？"):
            return
        try:
            # 关闭当前数据库连接
            conn, _ = self.db._connect()
            self.db._close(conn)
            # 覆盖数据库文件
            shutil.copy2(file_path, DB_FILE)
            messagebox.showinfo("成功", "数据库已恢复，程序将自动重启。")
            # 重启程序
            python = sys.executable
            os.execl(python, python, *sys.argv)
        except Exception as e:
            messagebox.showerror("恢复失败", str(e))

    def import_chapters_kps_from_excel(self):
        """导入章节和考点数据（替换当前）"""
        from tkinter import filedialog
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("错误", "请先安装 openpyxl 库：pip install openpyxl")
            return

        file_path = filedialog.askopenfilename(filetypes=[("Excel文件", "*.xlsx *.xls")])
        if not file_path:
            return
        if not messagebox.askyesno("确认导入", "导入将覆盖现有的章节和考点数据，是否继续？\n（程序会提示重启）"):
            return

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            new_chapters = {}
            new_knowledge_points = {}

            # 读取章节表
            if "章节" in wb.sheetnames:
                ws = wb["章节"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0] or not row[1] or not row[2]:
                        continue
                    subject = str(row[0]).strip()
                    grade = str(row[1]).strip()
                    chapter = str(row[2]).strip()
                    lesson = str(row[3]).strip() if row[3] else "整个章节"
                    new_chapters.setdefault(subject, {}).setdefault(grade, {}).setdefault(chapter, []).append(lesson)
            else:
                messagebox.showerror("错误", "Excel文件中缺少'章节'工作表")
                return

            # 读取考点表（可选）
            if "考点" in wb.sheetnames:
                ws = wb["考点"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0] or not row[4]:
                        continue
                    subject = str(row[0]).strip()
                    grade = str(row[1]).strip()
                    chapter = str(row[2]).strip() if row[2] else ""
                    lesson = str(row[3]).strip() if row[3] else "整个章节"
                    kp_id = str(row[4]).strip()
                    kp_name = str(row[5]).strip() if row[5] else kp_id
                    kp_desc = str(row[6]).strip() if row[6] else ""
                    new_knowledge_points[kp_id] = {
                        "name": kp_name,
                        "desc": kp_desc,
                        "subject": subject,
                        "grade": grade,
                        "chapter": chapter,
                        "lesson": lesson
                    }

            # 更新全局变量
            global CHAPTERS, KNOWLEDGE_POINTS
            CHAPTERS = new_chapters
            KNOWLEDGE_POINTS = new_knowledge_points

            # 同步到数据库的 knowledge_base 表
            conn, cursor = self.db._connect()
            cursor.execute("DELETE FROM knowledge_base")  # 清空原有骨架
            skeleton_data = []
            for subject, grades in CHAPTERS.items():
                for grade, chapters in grades.items():
                    for chapter, lessons in chapters.items():
                        for lesson in lessons:
                            skeleton_data.append((subject, grade, chapter, lesson, ""))
            if skeleton_data:
                cursor.executemany(
                    "INSERT INTO knowledge_base (subject, grade, chapter, lesson, content) VALUES (?, ?, ?, ?, ?)",
                    skeleton_data)
            self.db._close(conn)

            # 刷新界面中的下拉框
            self.refresh_chapter_options()
            self.update_exam_chapter_options()
            self.update_kp_combo_options()

            messagebox.showinfo("成功", "章节和考点数据已导入，请重新启动程序以确保所有功能正常。")
        except Exception as e:
            messagebox.showerror("导入失败", str(e))

    def batch_generate_knowledge(self):
        """批量生成当前科目、年级下所有章节的知识点（后台任务，显示进度）"""
        if not self.ai.api_key:
            messagebox.showwarning("提示", "请先配置 AI API Key")
            return

        subject = self.knowledge_subject_var.get()
        grade = self.knowledge_grade_var.get()

        # 获取章节列表
        if subject not in CHAPTERS or grade not in CHAPTERS[subject]:
            messagebox.showwarning("提示", f"当前科目 {subject} 年级 {grade} 没有预定义的章节数据，请先导入章节。")
            return

        chapters_dict = CHAPTERS[subject][grade]
        total_tasks = sum(len(lessons) for lessons in chapters_dict.values())
        if total_tasks == 0:
            messagebox.showinfo("提示", "没有可生成的章节课时。")
            return

        # 二次确认配额消耗
        if not messagebox.askyesno("批量生成确认",
                                   f"即将为 {subject} {grade} 生成 {total_tasks} 个章节/课时的知识点。\n\n"
                                   "此操作会消耗大量 API Token，请确保您的 API Key 有足够配额。\n\n是否继续？"):
            return

        # 创建进度窗口
        progress_win = tk.Toplevel(self.root)
        progress_win.title("批量生成知识点")
        progress_win.geometry("500x180")
        progress_win.transient(self.root)
        progress_win.grab_set()

        tk.Label(progress_win, text=f"正在为 {subject} {grade} 生成知识点...", font=FONTS["content"]).pack(pady=10)
        progress_bar = ttk.Progressbar(progress_win, length=400, mode='determinate', maximum=total_tasks)
        progress_bar.pack(pady=10)
        status_label = tk.Label(progress_win, text="准备就绪", font=FONTS["small"])
        status_label.pack()

        def generate():
            count = 0
            for chapter, lessons in chapters_dict.items():
                for lesson in lessons:
                    full_name = chapter if lesson == "整个章节" else f"{chapter} - {lesson}"
                    status_label.config(text=f"正在生成：{full_name}")
                    progress_bar['value'] = count
                    progress_win.update()

                    # 检查是否已有内容（可选：跳过已有内容的章节，或强制覆盖）
                    existing = self.db.get_chapter_content(subject, grade, chapter, lesson)
                    if existing and len(existing) > 50:
                        # 已有内容，跳过（可改为注释掉跳过，强制重新生成）
                        count += 1
                        continue

                    custom_prompt = self._get_knowledge_prompt(subject, grade, chapter, lesson)
                    success, content = self.ai.ai_request(subject, grade, "生成章节内容", custom_prompt,
                                                          auto_save=False, temperature=0.3)
                    if success:
                        final_content = f"【{full_name}】\n\n{content}"
                        self.db.update_chapter_content(subject, grade, chapter, lesson, final_content)
                    else:
                        status_label.config(text=f"失败：{full_name} - {content}")
                        progress_win.update()
                        time.sleep(1)
                    count += 1
                    progress_bar['value'] = count
                    progress_win.update()

            progress_win.destroy()
            messagebox.showinfo("批量生成完成", f"已处理 {total_tasks} 个章节/课时。\n请刷新知识库页面查看内容。")

        threading.Thread(target=generate, daemon=True).start()

    def clear_api_usage(self):
        if not messagebox.askyesno("确认清除", "此操作将永久删除所有 API 用量统计记录，是否继续？"):
            return
        conn, cursor = self.db._connect()
        cursor.execute("DELETE FROM api_usage")
        self.db._close(conn)

        # 刷新当前选中模型的统计
        selected = self.model_tree.selection()
        if selected:
            model_name = self.model_tree.item(selected[0], "values")[0]
            self.refresh_selected_model_stats(model_name)
        else:
            # 没有选中模型时，清空右侧并显示提示
            self.model_detail_info.config(text="")
            for item in self.scene_tree.get_children():
                self.scene_tree.delete(item)
            # 添加占位符，明确告知无数据
            self.scene_tree.insert("", tk.END, values=("暂无数据", 0, 0))

        messagebox.showinfo("完成", "API 用量记录已清空")

    def reset_selected_model_quota(self):
        selected = self.model_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择一个模型")
            return
        model_name = self.model_tree.item(selected[0], "values")[0]
        if messagebox.askyesno("确认重置", f"将重置 {model_name} 当前周期的已用tokens为0，是否继续？"):
            conn, cursor = self.db._connect()
            cursor.execute("UPDATE model_config SET used_tokens=0, period_start=? WHERE model_name=?",
                           (datetime.now().isoformat(), model_name))
            self.db._close(conn)
            self.refresh_model_list()  # 刷新左侧列表
            self.refresh_selected_model_stats(model_name)  # 刷新右侧统计
            messagebox.showinfo("成功", f"{model_name} 用量已重置")

    def load_font_size(self):
        """加载保存的字体大小配置"""
        config_file = "font_config.json"
        if os.path.exists(config_file):
            try:
                with open(config_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    base_size = data.get("base_size", 14)
                    self.apply_font_size(base_size, save=False)
            except:
                pass

    def apply_font_size(self, base_size, save=True):
        global FONTS
        new_fonts = {
            "title": ("微软雅黑", base_size + 6, "bold"),
            "nav": ("Segoe UI Emoji", base_size, "bold"),
            "btn": ("Segoe UI Emoji", base_size - 1, "bold"),
            "content": ("微软雅黑", base_size),
            "content_bold": ("微软雅黑", base_size, "bold"),
            "small": ("微软雅黑", base_size - 3),
            "combo": ("微软雅黑", base_size - 1),
            "placeholder": ("微软雅黑", base_size - 1),
            "ai_output": ("微软雅黑", base_size + 2)
        }
        FONTS.update(new_fonts)

        if save:
            # 保存到配置文件
            with open("font_config.json", "w", encoding="utf-8") as f:
                json.dump({"base_size": base_size}, f)
            # 提示重启
            messagebox.showinfo("字体设置", f"字体大小已设置为 {base_size} 号，请重启程序生效。")

    def refresh_model_list(self):
        for item in self.model_tree.get_children():
            self.model_tree.delete(item)
        conn, cursor = self.db._connect()
        cursor.execute("SELECT model_name, quota_type, quota_limit, used_tokens FROM model_config WHERE is_enabled=1 ORDER BY priority, model_name")
        rows = cursor.fetchall()
        self.db._close(conn)
        for r in rows:
            if r['quota_limit'] == 0:
                remaining = "无限"
            else:
                remaining = r['quota_limit'] - r['used_tokens']
                if remaining < 0:
                    remaining = 0
            quota_type_cn = "一次性" if r['quota_type'] == 'once' else (
                "每日" if r['quota_type'] == 'daily' else "每月")
            self.model_tree.insert("", tk.END, values=(r['model_name'], quota_type_cn, r['quota_limit'], r['used_tokens'], remaining))
        # 清空右侧详情（不再设置标题）
        self.model_detail_info.config(text="")
        for item in self.scene_tree.get_children():
            self.scene_tree.delete(item)

    def on_model_select(self, event):
        selected = self.model_tree.selection()
        if not selected:
            return
        model_name = self.model_tree.item(selected[0], "values")[0]
        self.refresh_selected_model_stats(model_name)

    def refresh_selected_model_stats(self, model_name=None):
        if model_name is None:
            selected = self.model_tree.selection()
            if not selected:
                return
            model_name = self.model_tree.item(selected[0], "values")[0]
        conn, cursor = self.db._connect()
        cursor.execute("SELECT quota_type, quota_limit, used_tokens FROM model_config WHERE model_name=?",
                       (model_name,))
        row = cursor.fetchone()
        if row:
            quota_type_cn = "一次性" if row['quota_type'] == 'once' else (
                "每日" if row['quota_type'] == 'daily' else "每月")
            limit = row['quota_limit']
            used = row['used_tokens']
            remaining = limit - used if limit > 0 else "无限"
            if limit > 0:
                percent = (remaining / limit) * 100
                info_text = f"周期类型：{quota_type_cn}\n额度上限：{limit} tokens\n本周期已用：{used} tokens\n剩余额度：{remaining} tokens ({percent:.1f}%)"
            else:
                info_text = f"周期类型：{quota_type_cn}\n额度上限：无限制\n本周期已用：{used} tokens\n剩余额度：无限"
            self.model_detail_info.config(text=info_text)
        else:
            self.model_detail_info.config(text="未找到模型配置")
        cursor.execute("""
            SELECT scene, COUNT(*) as requests, SUM(total_tokens) as tokens
            FROM api_usage
            WHERE model=?
            GROUP BY scene
            ORDER BY tokens DESC
        """, (model_name,))
        scene_rows = cursor.fetchall()
        self.db._close(conn)
        for item in self.scene_tree.get_children():
            self.scene_tree.delete(item)
        if scene_rows:
            for s in scene_rows:
                scene_name = s['scene'] if s['scene'] else "未知"
                self.scene_tree.insert("", tk.END, values=(scene_name, s['requests'], s['tokens']))
        else:
            self.scene_tree.insert("", tk.END, values=("暂无数据", 0, 0))

    def add_model_dialog(self):
        """弹出添加模型对话框（居中，周期中文）"""
        dialog = tk.Toplevel(self.root)
        dialog.title("添加模型")
        dialog.geometry("650x550")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg=COLORS["bg_main"])

        # 窗口居中
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (650 // 2)
        y = (dialog.winfo_screenheight() // 2) - (550 // 2)
        dialog.geometry(f"+{x}+{y}")

        main_frame = tk.Frame(dialog, bg=COLORS["bg_main"])
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)

        # 模型名称
        tk.Label(main_frame, text="模型名称", font=FONTS["content_bold"], bg=COLORS["bg_main"], anchor="w").pack(fill=tk.X, pady=(0,5))
        name_entry = tk.Entry(main_frame, font=FONTS["content"], width=30)
        name_entry.pack(fill=tk.X, pady=(0,10))
        self._setup_placeholder_entry(name_entry, "例如：my-gpt")

        # API 地址
        tk.Label(main_frame, text="API 地址", font=FONTS["content_bold"], bg=COLORS["bg_main"], anchor="w").pack(fill=tk.X, pady=(0,5))
        api_entry = tk.Entry(main_frame, font=FONTS["content"], width=40)
        api_entry.pack(fill=tk.X, pady=(0,10))
        self._setup_placeholder_entry(api_entry, "https://dashscope.aliyuncs.com/compatible-mode/v1/chat/completions")

        # 周期类型（中文显示，存储值对应英文）
        tk.Label(main_frame, text="周期类型", font=FONTS["content_bold"], bg=COLORS["bg_main"], anchor="w").pack(fill=tk.X, pady=(0,5))
        quota_type_var = tk.StringVar(value="一次性")
        quota_combo = ttk.Combobox(main_frame, textvariable=quota_type_var, values=["一次性", "每日", "每月"],
                                   state="readonly", width=20)
        quota_combo.pack(anchor="w", pady=(0,10))

        # 额度上限
        tk.Label(main_frame, text="额度上限（tokens，0表示无限制）", font=FONTS["content_bold"], bg=COLORS["bg_main"], anchor="w").pack(fill=tk.X, pady=(0,5))
        limit_entry = tk.Entry(main_frame, font=FONTS["content"], width=15)
        limit_entry.pack(anchor="w", pady=(0,10))
        self._setup_placeholder_entry(limit_entry, "例如：1000000")

        # 功能说明
        desc_frame = tk.Frame(main_frame, bg=COLORS["bg_main"])
        desc_frame.pack(fill=tk.X, pady=(10,0))
        tk.Label(desc_frame, text="📌 功能说明：", font=FONTS["content_bold"], bg=COLORS["bg_main"], fg=COLORS["primary"]).pack(anchor="w")
        tk.Label(desc_frame, text="  • 模型名称：自定义名称，用于在系统中显示和选择。",
                 font=FONTS["small"], bg=COLORS["bg_main"], fg=COLORS["text_secondary"], justify="left", anchor="w").pack(anchor="w", padx=(15,0))
        tk.Label(desc_frame, text="  • API 地址：模型的接口地址，需兼容 OpenAI 格式。默认使用阿里云通义千问。",
                 font=FONTS["small"], bg=COLORS["bg_main"], fg=COLORS["text_secondary"], justify="left", anchor="w").pack(anchor="w", padx=(15,0))
        tk.Label(desc_frame, text="  • 周期类型：一次性（用尽不重置）、每日（每天 0 点重置）、每月（每月 1 日重置）",
                 font=FONTS["small"], bg=COLORS["bg_main"], fg=COLORS["text_secondary"], justify="left", anchor="w").pack(anchor="w", padx=(15,0))
        tk.Label(desc_frame, text="  • 额度上限：每周期可使用的最大 tokens 数，0 表示不限制。",
                 font=FONTS["small"], bg=COLORS["bg_main"], fg=COLORS["text_secondary"], justify="left", anchor="w").pack(anchor="w", padx=(15,0))

        def save():
            name = name_entry.get().strip()
            api_base = api_entry.get().strip()
            quota_type = "once" if quota_type_var.get() == "一次性" else (
                "daily" if quota_type_var.get() == "每日" else "monthly")
            try:
                limit = int(limit_entry.get().strip())
            except:
                messagebox.showerror("错误", "额度上限必须是数字")
                return
            if not name:
                messagebox.showerror("错误", "模型名称不能为空")
                return
            conn, cursor = self.db._connect()
            cursor.execute("SELECT 1 FROM model_config WHERE model_name=?", (name,))
            if cursor.fetchone():
                self.db._close(conn)
                messagebox.showerror("错误", "模型名称已存在")
                return
            cursor.execute('''
                INSERT INTO model_config (model_name, api_base, quota_type, quota_limit, used_tokens, period_start, is_enabled)
                VALUES (?, ?, ?, ?, 0, ?, 1)
            ''', (name, api_base, quota_type, limit, datetime.now().isoformat()))
            self.db._close(conn)
            self.refresh_model_list()
            self._refresh_model_combo()
            dialog.destroy()
            messagebox.showinfo("成功", f"模型 {name} 已添加")

        btn_save = tk.Button(main_frame, text="💾 保存", command=save, width=10, cursor="hand2", bg=COLORS["success"], fg="white", bd=0, relief=tk.FLAT)
        btn_save.pack(pady=15)

    def delete_selected_model(self):
        selected = self.model_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择一个模型")
            return
        item = selected[0]
        model_name = self.model_tree.item(item, "values")[0]
        if model_name in ["glm-4.7-flash", "glm-4-flash", "glm-4-flash-250414", "qwen3.5-flash", "qwen3.5-plus"]:
            messagebox.showwarning("提示", "预设模型不能删除")
            return
        if messagebox.askyesno("确认删除", f"确定要删除模型 {model_name} 吗？"):
            conn, cursor = self.db._connect()
            cursor.execute("DELETE FROM model_config WHERE model_name=?", (model_name,))
            self.db._close(conn)
            self.refresh_model_list()
            self._refresh_model_combo()
            messagebox.showinfo("成功", f"模型 {model_name} 已删除")

    def _refresh_model_combo(self):
        """刷新系统设置中的模型下拉框"""
        conn, cursor = self.db._connect()
        cursor.execute("SELECT model_name FROM model_config WHERE is_enabled=1 ORDER BY priority, model_name")
        rows = cursor.fetchall()
        self.db._close(conn)
        model_list = [r['model_name'] for r in rows]
        self.model_combo['values'] = model_list
        if self.ai.model_name not in model_list and model_list:
            self.model_combo.set(model_list[0])

    def _setup_placeholder_entry(self, entry, placeholder):
        """为 Entry 设置占位符"""
        entry.insert(0, placeholder)
        entry.config(fg="gray")

        def on_focus_in(event):
            if entry.get() == placeholder:
                entry.delete(0, tk.END)
                entry.config(fg="black")

        def on_focus_out(event):
            if entry.get().strip() == "":
                entry.delete(0, tk.END)
                entry.insert(0, placeholder)
                entry.config(fg="gray")

        entry.bind("<FocusIn>", on_focus_in)
        entry.bind("<FocusOut>", on_focus_out)

    def _get_knowledge_prompt(self, subject, grade, chapter, lesson):
        """根据科目和课时生成不同的章节内容prompt"""
        # 判断是“整个章节”还是具体课时
        if lesson == "整个章节":
            title = chapter
            specificity = "请生成该单元/章节的综合知识点，包括知识框架、核心概念、重点难点、典型例题、单元总结等。"
        else:
            title = f"{chapter} - {lesson}"
            specificity = "请针对该课时生成详细知识点。"

        base = f"你是初中{subject}教材专家，请为{self.ai.area}{self.ai.textbook_version}{grade}{subject}的《{title}》生成详细知识点。{specificity}"

        # 不同科目的个性化要求
        subject_requirements = {
            "语文": "包括：课文内容概括、作者简介、写作背景、重点字词、段落结构分析、修辞手法赏析、中心思想、课后练习题解析。用清晰的标题分段输出。",
            "数学": "包括：核心概念定义、公式（用LaTeX格式）、定理、性质、典型例题（带详细解题步骤）、常见易错点、练习题（2-3道）。用清晰的标题分段输出。",
            "英语": "包括：课文翻译、重点词汇（词性、例句）、核心语法点、句型结构分析、课文朗读技巧、课后习题答案。用清晰的标题分段输出。",
            "物理": "包括：基本概念、公式（用LaTeX格式）、单位、定律内容、实验原理与步骤、典型计算题（带解析）、生活应用实例。用清晰的标题分段输出。",
            "化学": "包括：基本概念、化学方程式、实验现象与结论、物质性质（物理/化学）、典型计算题、记忆口诀。用清晰的标题分段输出。",
            "生物": "包括：核心概念、结构图描述、生理过程步骤、分类特征、实验设计、易混淆点对比。用清晰的标题分段输出。",
            "道德与法治": "包括：核心观点、法律条文/道德规范、案例分析、生活实践指导、相关时政链接。用清晰的标题分段输出。",
            "历史": "包括：事件时间线、原因背景、主要人物、过程经过、历史意义、影响评价、相关地图/文献描述。用清晰的标题分段输出。",
            "地理": "包括：地理位置、自然特征（地形/气候/水文）、人文特征、成因分析、图表数据描述、典型例题。用清晰的标题分段输出。",
        }
        requirement = subject_requirements.get(subject, "包括：核心概念、原理、典型例题、易错点。用清晰的标题分段输出。")
        return f"{base}{requirement}。请确保输出内容完整，不要因为长度限制而提前截断。输出字数为4000字左右。不要输出任何标记，如#、*、-等），不要输出任何开场白（如'好的'、'以下是'）或结束语（如'希望对您有所帮助'），直接输出包含标题的知识点正文，也不要说出限定的字数。"

    def _clean_markdown(self, text):
        """彻底清理 Markdown 标记、序号、加粗等，只保留纯文本"""
        import re
        # 删除所有 # 号标题（### 等）及后面的空格
        text = re.sub(r'^#{1,6}\s*', '', text, flags=re.MULTILINE)
        # 删除加粗 **文本** -> 文本
        text = re.sub(r'\*\*(.*?)\*\*', r'\1', text, flags=re.DOTALL)
        # 删除斜体 *文本* -> 文本
        text = re.sub(r'\*(.*?)\*', r'\1', text, flags=re.DOTALL)
        # 删除无序列表前的 - 或 * 以及后面的空格
        text = re.sub(r'^[\-\*]\s+', '', text, flags=re.MULTILINE)
        # 删除有序列表前的数字加点（如 1. 2.）及后面的空格
        text = re.sub(r'^\d+\.\s*', '', text, flags=re.MULTILINE)
        # 删除可能残留的单个 # 符号（如行内）
        text = re.sub(r'#', '', text)
        # 删除各种占位符图标
        icons = '💡⚠️✅❌⏳🔍📌✏️🔊⏹️📤🗑️🔄💾📂📎🔊⏹️📤🗑️🔄💾📂📎'
        for icon in icons:
            text = text.replace(icon, '')
        # 将连续三个以上的换行压缩为两个换行
        text = re.sub(r'\n{3,}', '\n\n', text)
        # 删除类似 --- 的分隔线
        text = re.sub(r'---\s*\n', '', text)
        # 去掉开头和结尾的空白
        return text.strip()

    def _is_win7(self):
        """检测是否为 Windows 7"""
        import platform
        return platform.system() == 'Windows' and platform.release() == '7'

    def _remove_all_emoji(self):
        """Win7 下删除所有 Emoji 图标（只保留文字），并将搜索按钮替换为“搜索”"""
        if not self._is_win7():
            return

        # Emoji 正则（匹配常见 Emoji 范围）
        import re
        emoji_pattern = re.compile("["
                                   u"\U0001F600-\U0001F64F"  # 表情
                                   u"\U0001F300-\U0001F5FF"  # 符号
                                   u"\U0001F680-\U0001F6FF"  # 交通
                                   u"\U0001F1E0-\U0001F1FF"  # 国旗
                                   u"\U00002702-\U000027B0"
                                   u"\U000024C2-\U0001F251"
                                   u"\U0001F900-\U0001F9FF"  # 补充符号
                                   "]+", flags=re.UNICODE)

        def clean_text(text):
            """删除文本中的 Emoji，并删除 Emoji 后面的空格"""
            # 先删除 Emoji 后面可能紧跟的空格
            text = re.sub(r'[\U00010000-\U0010FFFF]\s*', '', text)
            # 再删除所有 Emoji
            return emoji_pattern.sub('', text).strip()

        def process_widget(widget):
            """递归处理控件"""
            # 处理 Label 和 Button
            if isinstance(widget, (tk.Label, tk.Button)):
                try:
                    current = widget.cget('text')
                    if current:
                        # 特殊处理搜索按钮（文本为 🔍 或包含 🔍）
                        if current == "🔍" or current.strip() == "🔍":
                            widget.config(text="搜索", width=6)
                        else:
                            new_text = clean_text(current)
                            if new_text != current:
                                widget.config(text=new_text)
                except tk.TclError:
                    pass
            # 递归子控件
            for child in widget.winfo_children():
                process_widget(child)

        process_widget(self.root)

    def _adjust_buttons_for_win7(self):
        """Win7 下调整按钮内边距和字体，解决按钮过小问题"""
        if not self._is_win7():
            return

        # 方法1：为所有 tk.Button 增加内边距
        def adjust(widget):
            if isinstance(widget, tk.Button):
                try:
                    # 增加左右内边距和上下内边距
                    widget.config(padx=10, pady=5)
                    # 如果字体是 Emoji 字体，替换为普通字体
                    font = widget.cget('font')
                    if isinstance(font, tuple) and 'Segoe UI Emoji' in font[0]:
                        widget.config(font=('微软雅黑', font[1], font[2] if len(font) > 2 else 'normal'))
                except:
                    pass
            for child in widget.winfo_children():
                adjust(child)

        adjust(self.root)

        # 方法2：使用 ttk 样式（如果使用了 ttk.Button）
        try:
            from tkinter import ttk
            style = ttk.Style()
            style.configure('TButton', padding=(10, 5), font=('微软雅黑', 11))
        except:
            pass

    def save_tts_mode(self, mode):
        """保存语音模式到配置文件"""
        config_file = "tts_config.json"
        try:
            if os.path.exists(config_file):
                with open(config_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
            else:
                data = {}
            data["mode"] = mode
            with open(config_file, "w", encoding="utf-8") as f:
                json.dump(data, f)
        except:
            pass

    def load_tts_mode(self):
        """从配置文件加载语音模式"""
        config_file = "tts_config.json"
        if os.path.exists(config_file):
            try:
                with open(config_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    return data.get("mode", "offline")
            except:
                pass
        return "offline"

    def enable_touch_scroll_for_all(self):
        """为所有支持滚动的控件添加触摸滑动支持（递归）"""

        def add_touch_support(widget):
            # 为 Text 控件添加触摸滑动
            if isinstance(widget, tk.Text):
                def on_touch_start(e):
                    widget._scroll_start_y = e.y

                def on_touch_move(e):
                    delta = e.y - widget._scroll_start_y
                    widget.yview_scroll(-delta, "units")
                    widget._scroll_start_y = e.y

                widget.bind("<Button-1>", on_touch_start)
                widget.bind("<B1-Motion>", on_touch_move)
            # 为 Canvas 控件添加触摸滑动
            elif isinstance(widget, tk.Canvas):
                def on_touch_start(e):
                    widget._scroll_start_y = e.y

                def on_touch_move(e):
                    delta = e.y - widget._scroll_start_y
                    widget.yview_scroll(-delta, "units")
                    widget._scroll_start_y = e.y

                widget.bind("<Button-1>", on_touch_start)
                widget.bind("<B1-Motion>", on_touch_move)
            # 为 ttk.Treeview 添加触摸滑动
            elif isinstance(widget, ttk.Treeview):
                def on_touch_start(e):
                    widget._scroll_start_y = e.y

                def on_touch_move(e):
                    delta = e.y - widget._scroll_start_y
                    widget.yview_scroll(-delta, "units")
                    widget._scroll_start_y = e.y

                widget.bind("<Button-1>", on_touch_start)
                widget.bind("<B1-Motion>", on_touch_move)
            # 递归处理子控件
            for child in widget.winfo_children():
                add_touch_support(child)

        add_touch_support(self.root)



# ---------------------- 程序入口 ----------------------
if __name__ == "__main__":
    root = tk.Tk()
    app = SmartClassroomApp(root)
    root.mainloop()